#!/usr/bin/env python3
"""
HSG17 T0-to-Host Validator — Clean build from scratch.

Full pipeline:
- Two required/strongly recommended uploads (allconnections + pre-classified cutsheet)
- 6-stage clean processor (Ingest → Normalize with DH block derivation → Enrich → Analyze → Format → Log)
- Professional 5-tab output (Summary + the 4 error categories the user already uses)
- Prominent "Block" (DH-xxx) column using the authoritative strategy from the Bootstrap Sequence
- Silent central logging (per DH block) so the Dashboard shows current state + deltas
- No extra buttons — logging happens automatically on successful Process
"""

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
from pathlib import Path
import re

from utils.auth import require_login
from utils.data_logger import log_errors
from utils.hsg17_t0_host import process_hsg17_t0_host, extract_counts_for_logging

require_login()

st.set_page_config(page_title="HSG17 T0-to-Host", page_icon="🖥️", layout="wide")
st.title("HSG17 T0-to-Host Validator")
st.caption("Clean implementation • DH block derivation from authoritative Bootstrap Sequence • Feeds central Dashboard")

st.markdown("""
**Inputs (from Batam folder):**
- `QFABT0toHOST_allconnections.xlsx` — the big raw connections file (PP labels + device/rack info)
- `rack_validation_merged_2614HC8023.xlsx` — the pre-classified cutsheet with the 4 error sheets you already use
""")

# ================== Uploaders ==================
st.markdown("### Input Files")
col1, col2 = st.columns(2)

with col1:
    allconn_file = st.file_uploader(
        "T0-to-Host All Connections (required)",
        type=["xlsx", "xlsm"],
        key="hsg17_allconn",
        help="QFABT0toHOST_allconnections.xlsx — used for PP enrichment and block derivation"
    )

with col2:
    cutsheet_file = st.file_uploader(
        "Pre-classified Cutsheet (strongly recommended)",
        type=["xlsx", "xlsm"],
        key="hsg17_cutsheet",
        help="rack_validation_merged_....xlsx — contains the 4 error category sheets"
    )

can_process = bool(allconn_file)

# ================== Process Button ==================
if st.button("🚀 Process Files", type="primary", disabled=not can_process, key="hsg17_process"):
    with st.spinner("Running clean HSG17 T0-to-Host pipeline (block derivation, enrichment, formatting)..."):
        try:
            all_bytes = allconn_file.getvalue()
            cuts_bytes = cutsheet_file.getvalue() if cutsheet_file else b""

            result_bytes, filename = process_hsg17_t0_host(
                all_bytes,
                cuts_bytes,
                source_filename=allconn_file.name if allconn_file else "HSG17"
            )

            if result_bytes:
                st.success("✅ Processing complete! (clean 6-stage pipeline + connected-component clustering + PP enrichment)")

                # ====================== PRE-DOWNLOAD PREVIEW (matches the QFAB_V2 pattern the user likes) ======================
                st.subheader("📊 Pre-Download Analysis")

                wb_preview = load_workbook(BytesIO(result_bytes))

                # Summary tab as table (very useful for the user)
                if "Summary" in wb_preview.sheetnames:
                    st.markdown("**Summary (by Block)**")
                    ws_sum = wb_preview["Summary"]
                    summary_data = []
                    for row in ws_sum.iter_rows(min_row=3, values_only=True):
                        if row and row[0]:
                            summary_data.append([str(c) if c is not None else "" for c in row])
                    if summary_data:
                        st.table(summary_data[:25])  # cap for UI sanity on first runs

                # Quick metrics per tab
                st.markdown("**Tab Row Counts**")
                tab_counts = {}
                for sheet_name in wb_preview.sheetnames:
                    if sheet_name != "Summary":
                        tab_counts[sheet_name] = max(0, wb_preview[sheet_name].max_row - 3)

                if tab_counts:
                    cols = st.columns(len(tab_counts))
                    for i, (tab, count) in enumerate(tab_counts.items()):
                        with cols[i]:
                            st.metric(tab.replace(" Errors", ""), count)

                # ====================== SILENT CENTRAL LOGGING (exact pattern user approved) ======================
                # No extra button. Happens automatically. Only warns on real failure.
                try:
                    counts = extract_counts_for_logging(result_bytes)

                    for entry in counts:
                        blk = entry["block"]
                        for cat in ["LLDP Mismatch + Link Down", "Optic Errors", "FEC_BER Errors", "Interface Down Errors"]:
                            cnt = entry.get(cat, 0) or 0
                            if cnt > 0:
                                log_errors(
                                    hall="HSG17",
                                    rack_type="T0-Host",
                                    building=blk,           # This becomes the "Block" in the Dashboard
                                    error_category=cat,
                                    count=int(cnt),
                                    source_file=filename,
                                    processed_by="HSG17_T0_Host_v1"
                                )
                except Exception as log_exc:
                    st.warning(f"Central logging encountered an issue (non-fatal): {log_exc}")

                # ====================== DOWNLOAD ======================
                st.download_button(
                    "📥 Download Formatted HSG17 Report",
                    data=result_bytes,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="hsg17_download"
                )

                st.caption("All processing is local. Rich features now active: connected-component clustering (orange/yellow rows by Cluster on LLDP tab), PP enrichment, prominent Block (DH-xxx) column. Summary + per-Block counts are silently logged for the Dashboard.")

        except Exception as e:
            st.error(f"Error during HSG17 processing: {e}")
            st.exception(e)

st.markdown("---")
st.caption("HSG17 clean build • Block strategy uses the authoritative DH sectors from your Bootstrap Sequence document • Same central log as JPB15/SYD20 so one Dashboard works for everything")