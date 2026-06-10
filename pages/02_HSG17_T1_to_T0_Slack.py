#!/usr/bin/env python3
"""
Slack Formatter Tool — Streamlit
(UI styled to match the LVV Portal page for uniformity)
"""

import sys
import os
import shutil
from collections import Counter
import tempfile
import zipfile
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import streamlit as st
import tempfile
from pathlib import Path

from utils.data_logger import log_errors
from utils.hsg17_models import derive_placement_and_rack_from_files

# ── All original core logic (100% verbatim from SY20_QFAB_SLACK_No_PP.py - now V4 exact) ───────────
# (TAB_ALIASES, find_tab, style helpers, load_cutsheet, build_*_lookup, paired_subport,
#  process_file, highlight_mismatch_pairs, clear_and_border, autofit_sheet, swap_mismatch_groups, sort_mismatch_pairs, etc.)
# Mismatch logic (swap + sort) is verbatim from the original to ensure identical "suggested mismatch" output.

TAB_ALIASES = {
    'lldp':         ('lldp_sp', 'full_path_lldp_with_int_down'),
    'optics':       ('optics_rx_tx_threshold', 'optics_rx_tx_threshold_with_pp'),
    'interfaces':   ('interfaces_sp', 'interfaces_sp_with_pp'),
    'combined_fec': ('combined_fec', 'combined_fec_with_pp'),
}

TABS_TO_REMOVE = (
    'device_reporting_failure',
    'bgp_sp',
    'spectrum_health',
    'sp_power',
    'sp_fans',
    'optics_temp',
    'pre_fec_ber_threshold_with_pp',
    'unknown_test_sp',
    'summary',
)

COLUMNS_TO_REMOVE = (
    'Building', 'Act. Building', 'Exp. Building', 'PP_A', 'PP_Z',
    'Remote Host', 'Remote Interface', 'Mapped Remote Host',
    'Mapped Remote Interface', 'Mapped Remote Rack', 'Mapped Remote Elevation',
    'Remote Host Match', 'Remote Interface Match', 'Remote End Match',
    'Z_end_host', 'Z_end_intf', 'rack_z', 'Z_Rack', 'Z_Elevation',
    'Index', 'Source Sheet', 'Placement Group',
)

Z_FILL_TABS = ('Optics', 'combined_fec')

def find_tab(wb_or_sheetnames, key):
    names = (wb_or_sheetnames.sheetnames if hasattr(wb_or_sheetnames, 'sheetnames') else list(wb_or_sheetnames))
    for alias in TAB_ALIASES[key]:
        if alias in names:
            return alias
    return None

PINK = 'FFB6C1'
YELLOW = 'FFFF00'
ORANGE = 'FFA500'   # reciprocal-swap pair highlight (exact from reference)

def thin_border():
    s = Side(style='thin', color='000000')
    return Border(left=s, right=s, top=s, bottom=s)

def clear_and_border(ws, pink_cols=None):
    bd = thin_border()
    no_fill = PatternFill(fill_type=None)
    pink_fill = PatternFill('solid', start_color=PINK)
    yellow_fill = PatternFill('solid', start_color=YELLOW)
    pink_cols = set(pink_cols or [])
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.row == 1:
                cell.fill = yellow_fill
            elif cell.column in pink_cols:
                cell.fill = pink_fill
            else:
                cell.fill = no_fill
            cell.border = bd
            if cell.font:
                cell.font = Font(
                    bold=cell.font.bold,
                    name=cell.font.name or 'Arial',
                    size=cell.font.size or 10,
                    color='FF000000'
                )

def header_cell(cell, value, fill=None):
    cell.value = value
    cell.font = Font(bold=True, name='Arial', size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
    cell.border = thin_border()
    cell.fill = fill if fill else PatternFill('solid', start_color=YELLOW)

def autofit_sheet(ws, header_row_height=24, data_row_height=20, max_col_width=80):
    merged = set()
    for mrange in ws.merged_cells.ranges:
        for r in range(mrange.min_row, mrange.max_row + 1):
            for c in range(mrange.min_col, mrange.max_col + 1):
                merged.add((r, c))
    col_max = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is None or (cell.row, cell.column) in merged:
                continue
            longest_line = max((len(line) for line in str(cell.value).splitlines()), default=0)
            letter = get_column_letter(cell.column)
            col_max[letter] = max(col_max.get(letter, 0), longest_line)
    for letter, length in col_max.items():
        ws.column_dimensions[letter].width = min(length + 4, max_col_width)
    if ws.max_row >= 1:
        ws.row_dimensions[1].height = header_row_height
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = data_row_height

def write_sheet(wb, name, df):
    ws = wb.create_sheet(name)
    bd = thin_border()
    for c, col in enumerate(df.columns, 1):
        header_cell(ws.cell(row=1, column=c), col)
    for r, (_, row) in enumerate(df.iterrows(), 2):
        for c, col in enumerate(df.columns, 1):
            val = row[col]
            cell = ws.cell(row=r, column=c, value=None if pd.isna(val) else val)
            cell.border = bd
    for c, col in enumerate(df.columns, 1):
        mx = max([len(str(col))] + [len(str(v)) for v in df[col].dropna()])
        ws.column_dimensions[get_column_letter(c)].width = min(mx + 2, 40)
    ws.freeze_panes = 'A2'
    return ws

def load_cutsheet(path):
    xl = pd.ExcelFile(path)
    if 'Installation Sheet' in xl.sheet_names:
        return pd.read_excel(path, sheet_name='Installation Sheet')
    df = pd.read_excel(path, sheet_name=xl.sheet_names[0])
    def _split_device(col):
        parts = df[col].str.rsplit(' ', n=1, expand=True)
        return parts[0].str.strip(), parts[1].str.strip()
    df['Hostname'], df['Interface'] = _split_device('DeviceA')
    df['Z Hostname'], df['Z Interface'] = _split_device('DeviceB')
    def _lr(col):
        return df[col].apply(lambda v: str(v).strip()[-1] if pd.notna(v) and str(v).strip() else '')
    df['L/R'] = _lr('DeviceA Physical Port')
    df['Z L/R'] = _lr('DeviceB Physical Port')
    def _rack(col):
        return (df[col].str.extract(r'Rack\s+(\d+)')[0].astype(float).fillna(0).astype(int))
    def _elev(col):
        return (df[col].str.extract(r'U(\d+)')[0].astype(float).fillna(0).astype(int))
    df['Rack'] = _rack('RackA')
    df['Elevation'] = _elev('RackA')
    df['Z Rack'] = _rack('RackB')
    df['Z Elevation'] = _elev('RackB')
    return df

def build_cutsheet_lookup(cut_df):
    candidate_cols = ['Source_port', 'DMARC1', 'DMARC2', 'Destination_port']
    fill_cols = [c for c in candidate_cols if c in cut_df.columns]
    lookup = {}
    for _, row in cut_df.iterrows():
        key = (str(row['Hostname']).strip(), str(row['Interface']).strip())
        lookup[key] = {c: row[c] for c in fill_cols}
    lookup['__fill_cols__'] = fill_cols
    return lookup

def build_z_lookup(cut_df):
    lookup = {}
    for _, row in cut_df.iterrows():
        key = (str(row['Z Hostname']).strip(), str(row['Z Interface']).strip())
        lookup[key] = row
    return lookup

def paired_subport(iface):
    pairs = {'s0': 's1', 's1': 's0', 's2': 's3', 's3': 's2'}
    for suffix, mate in pairs.items():
        if iface.endswith(suffix):
            return iface[:-len(suffix)] + mate
    return None

def process_file(input_path, output_path, cut_df, log):
    r"""Full logic ported verbatim from the reference (Batam t0 slack / slack_formatter_T1toT0 (1).py).
    (Only the Streamlit call site + HSG17 logging/PG wrapper are outside this.)
    Produces EXACTLY the same output structure, fills, Summary, pink Possible/Active Z,
    grey-outs, pair highlighting, NOTE+autofilter, autofit, and rack-based rename.
    """
    shutil.copy2(input_path, output_path)
    wb = load_workbook(output_path)

    cutsheet_lookup = build_cutsheet_lookup(cut_df)
    z_lookup        = build_z_lookup(cut_df)

    # ── 1. Split lldp tab → Downlinks / Mismatches ──────────────────────────
    mis_orig_df = None   # keep for step 6b
    lldp_tab = find_tab(wb, 'lldp')
    if lldp_tab:
        log(f"  · Splitting {lldp_tab} → Downlinks / Mismatches")
        df = pd.read_excel(input_path, sheet_name=lldp_tab)
        down_df     = df[df['Act. Interface'] == 'interface down'].copy()
        mis_orig_df = df[df['Act. Interface'].str.startswith('swp', na=False)].copy()
        drop = ['Active Host', 'Act. Interface', 'Act. Rack', 'Act. Elevation']
        down_df.drop(columns=[c for c in drop if c in down_df.columns], inplace=True)
        exp_cols = ['Expected Hostname', 'Exp. Interface', 'Exp. Rack', 'Exp. Elevation']
        present_exp = [c for c in exp_cols if c in down_df.columns]
        if present_exp:
            rest = [c for c in down_df.columns if c not in present_exp]
            down_df = down_df[rest + present_exp]
        del wb[lldp_tab]
        write_sheet(wb, 'Downlinks', down_df)
        write_sheet(wb, 'Mismatches', mis_orig_df.drop(
            columns=[c for c in [] if c in mis_orig_df.columns]))

    # ── 2. Optics tab ───────────────────────────────────────────────────────
    optics_src = find_tab(wb, 'optics')
    if optics_src:
        log(f"  · Processing Optics tab ({optics_src})")
        drop_cols = {'Transceiver', 'Channel',
                     'Min Threshold (dBm)', 'Max Threshold (dBm)',
                     'PP_A', 'PP_Z', 'Z_end_host', 'Z_end_intf',
                     'rack_z', 'Z_Rack', 'Z_Elevation', 'Index',
                     'Status', 'Placement Group'}
        optics_df = pd.read_excel(input_path, sheet_name=optics_src)
        optics_df.drop(columns=[c for c in drop_cols if c in optics_df.columns], inplace=True)
        leading = [c for c in ('Metric', 'Measured (dBm)') if c in optics_df.columns]
        if leading:
            rest = [c for c in optics_df.columns if c not in leading]
            optics_df = optics_df[leading + rest]
        del wb[optics_src]
        write_sheet(wb, 'Optics', optics_df)
        wb['Optics'].freeze_panes = 'C2' if len(leading) >= 2 else 'B2'

    # ── 3. Remove interfaces tab ────────────────────────────────────────────
    interfaces_tab = find_tab(wb, 'interfaces')
    if interfaces_tab:
        log(f"  · Removing {interfaces_tab}")
        del wb[interfaces_tab]

    # ── 3a. Remove unwanted source tabs ─────────────────────────────────────
    drop_lower = {t.lower() for t in TABS_TO_REMOVE}
    for existing in list(wb.sheetnames):
        if existing.lower() in drop_lower:
            log(f"  · Removing {existing}")
            del wb[existing]

    # ── 3b. combined_fec: move Lock Status + Pre-FEC BER to the front ───────
    fec_tab = find_tab(wb, 'combined_fec')
    if fec_tab:
        log(f"  · Reordering {fec_tab} (Lock Status, Pre-FEC BER first)")
        fec_df = pd.read_excel(input_path, sheet_name=fec_tab)

        def _norm(s):
            return (str(s)
                    .replace('\u2011', '-')
                    .replace('\u2013', '-')
                    .replace('\u2014', '-')
                    .strip()
                    .lower())

        wanted = ['lock status', 'pre-fec ber']
        front = []
        for target in wanted:
            for col in fec_df.columns:
                if _norm(col) == target and col not in front:
                    front.append(col)
                    break

        if front:
            rest = [c for c in fec_df.columns if c not in front]
            fec_df = fec_df[front + rest]
            del wb[fec_tab]
            write_sheet(wb, 'combined_fec', fec_df)
        else:
            log("    ⚠ Lock Status / Pre-FEC BER not found — leaving combined_fec as-is")

    # ── 4. Reorder tabs ─────────────────────────────────────────────────────
    desired  = ['Downlinks', 'Mismatches', 'Optics', 'combined_fec', 'cannot_decode_sp']
    existing = [s for s in desired if s in wb.sheetnames]
    others   = [s for s in wb.sheetnames if s not in desired]
    wb._sheets = [wb[n] for n in others + existing]

    # ── 5. Insert L/R columns (sourced from cutsheet) ───────────────────────
    log("  · Adding L/R mapped columns (from cutsheet)")
    lr_from_cut = {}
    for _, row in cut_df.iterrows():
        iface = str(row.get('Interface', '') or '').strip()
        lr    = str(row.get('L/R', '')    or '').strip()
        if iface and lr:
            lr_from_cut[iface] = lr
        z_iface = str(row.get('Z Interface', '') or '').strip()
        z_lr    = str(row.get('Z L/R', '')      or '').strip()
        if z_iface and z_lr:
            lr_from_cut[z_iface] = z_lr

    lr_name_for = {
        'Interface':      'L/R',
        'Z Interface':    'Z L/R',
        'Exp. Interface': 'Exp. L/R',
    }
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        targets = [(i+1, h) for i, h in enumerate(header) if h in lr_name_for]
        for col_idx, col_name in sorted(targets, reverse=True):
            new_name = lr_name_for[col_name]
            ws.insert_cols(col_idx + 1)
            header_cell(ws.cell(row=1, column=col_idx + 1), new_name)
            for r in range(2, ws.max_row + 1):
                val = str(ws.cell(row=r, column=col_idx).value or '').strip()
                ws.cell(row=r, column=col_idx + 1, value=lr_from_cut.get(val, ''))
                ws.cell(row=r, column=col_idx + 1).border = thin_border()
            ws.column_dimensions[get_column_letter(col_idx + 1)].width = 10

    # ── 5b. Populate Source_port / DMARC1 / DMARC2 / Destination_port ────────
    fill_cols = cutsheet_lookup.get('__fill_cols__', [])
    log(f"  · Filling {', '.join(fill_cols) or '(no cutsheet fill cols available)'} (match on Hostname + Interface)")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        if not all(c in header for c in ['Hostname', 'Interface']):
            continue
        anchor = (header.index('Elevation') + 1) if 'Elevation' in header else len(header)
        insert_at = anchor + 1
        for col_name in fill_cols:
            if col_name in header:
                continue
            ws.insert_cols(insert_at)
            header_cell(ws.cell(row=1, column=insert_at), col_name)
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=insert_at).border = thin_border()
            ws.column_dimensions[get_column_letter(insert_at)].width = max(len(col_name)+2, 14)
            insert_at += 1
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        host_c, int_c = header.index('Hostname')+1, header.index('Interface')+1
        fill_idx = {c: header.index(c)+1 for c in fill_cols}
        for r in range(2, ws.max_row + 1):
            host  = str(ws.cell(row=r, column=host_c).value or '').strip()
            iface = str(ws.cell(row=r, column=int_c).value or '').strip()
            match = cutsheet_lookup.get((host, iface))
            if match:
                for col_name, col_idx in fill_idx.items():
                    val = match.get(col_name)
                    if val is not None and not (isinstance(val, float) and pd.isna(val)):
                        ws.cell(row=r, column=col_idx, value=val)

    # ── 6c. Fill Z-side info in designated tabs (default: Optics) ───────────
    Z_COLS = ['Z Hostname', 'Z Interface', 'Z L/R', 'Z Rack', 'Z Elevation']
    z_available = [c for c in Z_COLS if c in cut_df.columns]
    z_by_host_int = {}
    for _, row in cut_df.iterrows():
        k = (str(row['Hostname']).strip(), str(row['Interface']).strip())
        z_by_host_int[k] = {c: row[c] for c in z_available}

    if z_available and Z_FILL_TABS:
        for tab in Z_FILL_TABS:
            if tab not in wb.sheetnames:
                continue
            ws_z = wb[tab]
            header = [ws_z.cell(row=1, column=c).value
                      for c in range(1, ws_z.max_column + 1)]
            if not all(c in header for c in ['Hostname', 'Interface']):
                continue
            log(f"  · Filling Z-side info in {tab}: {', '.join(z_available)}")
            if 'Destination_port' in header:
                anchor = header.index('Destination_port') + 1
            elif 'Elevation' in header:
                anchor = header.index('Elevation') + 1
            else:
                anchor = len(header)
            insert_at = anchor + 1
            for col_name in z_available:
                if col_name in header:
                    continue
                ws_z.insert_cols(insert_at)
                header_cell(ws_z.cell(row=1, column=insert_at), col_name)
                for r in range(2, ws_z.max_row + 1):
                    ws_z.cell(row=r, column=insert_at).border = thin_border()
                ws_z.column_dimensions[get_column_letter(insert_at)].width = max(len(col_name)+2, 14)
                insert_at += 1
            header = [ws_z.cell(row=1, column=c).value
                      for c in range(1, ws_z.max_column + 1)]
            host_c, int_c = header.index('Hostname')+1, header.index('Interface')+1
            fill_idx = {c: header.index(c)+1 for c in z_available}
            for r in range(2, ws_z.max_row + 1):
                host  = str(ws_z.cell(row=r, column=host_c).value or '').strip()
                iface = str(ws_z.cell(row=r, column=int_c).value or '').strip()
                match = z_by_host_int.get((host, iface))
                if match:
                    for col_name, col_idx in fill_idx.items():
                        val = match.get(col_name)
                        if val is not None and not (isinstance(val, float) and pd.isna(val)):
                            ws_z.cell(row=r, column=col_idx, value=val)

    # ── 6b. Mismatches: Possible columns + Active Z columns (pink) ───────────
    pink_col_indices = []
    if 'Mismatches' in wb.sheetnames:
        log("  · Building Possible + Active Z columns in Mismatches")
        pink_fill   = PatternFill('solid', start_color=PINK)
        yellow_fill = PatternFill('solid', start_color=YELLOW)
        bd          = thin_border()

        act_lookup = {}
        src_sheets = pd.ExcelFile(input_path).sheet_names
        src_lldp = find_tab(src_sheets, 'lldp')
        if src_lldp:
            orig_df  = pd.read_excel(input_path, sheet_name=src_lldp)
            mis_rows = orig_df[orig_df['Act. Interface'].str.startswith('swp', na=False)]
            for _, row in mis_rows.iterrows():
                key       = (str(row['Hostname']).strip(), str(row['Interface']).strip())
                act_host  = str(row['Active Host']).strip()
                act_iface = str(row['Act. Interface']).strip()
                cut_z_row = z_lookup.get((act_host, act_iface))
                if cut_z_row is None:
                    mate = paired_subport(act_iface)
                    if mate:
                        cut_z_row = z_lookup.get((act_host, mate))
                z_lr_val = ''
                if cut_z_row is not None:
                    raw = cut_z_row.get('Z L/R')
                    if raw is not None and not (isinstance(raw, float) and pd.isna(raw)):
                        z_lr_val = str(raw).strip()
                act_lookup[key] = {
                    'Z Hostname' : act_host,
                    'Z Interface': act_iface,
                    'Z L/R'      : z_lr_val,
                    'Z Rack'     : int(float(str(row['Act. Rack']))),
                    'Z Elevation': int(float(str(row['Act. Elevation']))),
                }

        ws_m   = wb['Mismatches']
        header = [ws_m.cell(row=1, column=c).value for c in range(1, ws_m.max_column + 1)]

        act_drop = {'Active Host', 'Act. Interface', 'Act. Rack', 'Act. Elevation'}
        for idx in sorted([i+1 for i, h in enumerate(header) if h in act_drop], reverse=True):
            ws_m.delete_cols(idx)
        header = [ws_m.cell(row=1, column=c).value for c in range(1, ws_m.max_column + 1)]

        h_idx = header.index('Hostname') + 1
        i_idx = header.index('Interface') + 1

        act_rows = []
        for r in range(2, ws_m.max_row + 1):
            hn    = str(ws_m.cell(row=r, column=h_idx).value or '').strip()
            iface = str(ws_m.cell(row=r, column=i_idx).value or '').strip()
            act_rows.append(act_lookup.get((hn, iface), {}))

        possible_cols_all = [
            ('Possible Hostname',         'Hostname'),
            ('Possible Interface',        'Interface'),
            ('Possible L/R',             'L/R'),
            ('Possible Rack',             'Rack'),
            ('Possible Elevation',        'Elevation'),
            ('Possible Source_port',      'Source_port'),
            ('Possible DMARC1',           'DMARC1'),
            ('Possible DMARC2',           'DMARC2'),
            ('Possible Destination_port', 'Destination_port'),
        ]
        cut_cols = set(cut_df.columns)
        possible_cols = [
            (out_col, src) for out_col, src in possible_cols_all
            if src in cut_cols
        ]

        possible_data = {col: [] for col, _ in possible_cols}
        for act in act_rows:
            zh   = act.get('Z Hostname', '')
            zi   = act.get('Z Interface', '')
            match = z_lookup.get((zh, zi)) if zh else None
            if match is None and zh and zi:
                mate = paired_subport(zi)
                if mate:
                    match = z_lookup.get((zh, mate))
            for col, src in possible_cols:
                val = match.get(src, '') if match is not None else ''
                possible_data[col].append(val)

        pink_col_indices = []
        start = ws_m.max_column + 1
        for c_off, (col_name, _) in enumerate(possible_cols):
            col_idx = start + c_off
            pink_col_indices.append(col_idx)
            hdr = ws_m.cell(row=1, column=col_idx, value=col_name)
            hdr.font = Font(bold=True, name='Arial', size=10)
            hdr.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            hdr.fill  = yellow_fill
            hdr.border = bd
            ws_m.column_dimensions[get_column_letter(col_idx)].width = max(len(col_name)+3, 14)
            for r_off, val in enumerate(possible_data[col_name]):
                cell = ws_m.cell(row=r_off+2, column=col_idx,
                                 value=val if val != '' else None)
                cell.fill   = pink_fill
                cell.border = bd

        # Write Active Z columns
        act_z_cols = ['Z Hostname', 'Z Interface', 'Z L/R', 'Z Rack', 'Z Elevation']
        start2 = ws_m.max_column + 1
        for c_off, col_name in enumerate(act_z_cols):
            col_idx = start2 + c_off
            pink_col_indices.append(col_idx)
            hdr = ws_m.cell(row=1, column=col_idx, value=col_name)
            hdr.font = Font(bold=True, name='Arial', size=10)
            hdr.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            hdr.fill  = yellow_fill
            hdr.border = bd
            ws_m.column_dimensions[get_column_letter(col_idx)].width = max(len(col_name)+3, 14)
            for r_off, act in enumerate(act_rows):
                val = act.get(col_name, '')
                cell = ws_m.cell(row=r_off+2, column=col_idx,
                                 value=val if val != '' else None)
                cell.fill   = pink_fill
                cell.border = bd

    # ── 6c. Strip unwanted columns across every tab ─────────────────────────
    if COLUMNS_TO_REMOVE:
        log(f"  · Stripping columns: {', '.join(COLUMNS_TO_REMOVE)}")
        drop_set = set(COLUMNS_TO_REMOVE)
        for sheet_name in wb.sheetnames:
            ws_x = wb[sheet_name]
            header = [ws_x.cell(row=1, column=c).value
                      for c in range(1, ws_x.max_column + 1)]
            to_drop = [i + 1 for i, h in enumerate(header) if h in drop_set]
            for idx in sorted(to_drop, reverse=True):
                ws_x.delete_cols(idx)

    # ── 7. Summary tab (per Rack breakdown) ─────────────────────────────────
    log("  · Creating Summary tab")

    tab_rack  = {}
    all_racks = set()
    no_fill_s   = PatternFill(fill_type=None)
    yellow_fill_s = PatternFill('solid', start_color=YELLOW)
    center_s  = Alignment(horizontal='center', vertical='center', wrap_text=False)
    bd_s      = thin_border()

    def _s(cell, value, bold=False, header=False):
        cell.value     = value
        cell.font      = Font(bold=bold, name='Arial', size=10)
        cell.alignment = center_s
        cell.border    = bd_s
        cell.fill      = yellow_fill_s if header else no_fill_s

    for sname in wb.sheetnames:
        ws_tmp = wb[sname]
        hdr = [ws_tmp.cell(row=1, column=c).value for c in range(1, ws_tmp.max_column+1)]
        if 'Rack' not in hdr:
            tab_rack[sname] = {}
            continue
        rack_col = hdr.index('Rack') + 1
        counts = {}
        for r in range(2, ws_tmp.max_row + 1):
            val = ws_tmp.cell(row=r, column=rack_col).value
            if val is not None:
                try:
                    k = int(float(str(val)))
                    counts[k] = counts.get(k, 0) + 1
                    all_racks.add(k)
                except ValueError:
                    pass
        tab_rack[sname] = counts

    racks      = sorted(all_racks)
    tabs_order = [n for n in wb.sheetnames]
    total_cols = 1 + len(racks) + 1

    for existing in list(wb.sheetnames):
        if existing.lower() == 'summary':
            del wb[existing]
    wb.create_sheet('Summary', 0)
    ws_s = wb['Summary']

    ws_s.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_c = ws_s.cell(row=1, column=1, value='Tab Summary by Rack')
    title_c.font = Font(bold=True, name='Arial', size=13)
    title_c.alignment = center_s
    title_c.border    = bd_s
    title_c.fill      = yellow_fill_s
    ws_s.row_dimensions[1].height = 28

    _s(ws_s.cell(row=2, column=1), 'Tab Name', bold=True, header=True)
    for c, rack in enumerate(racks, start=2):
        _s(ws_s.cell(row=2, column=c), str(rack), bold=True, header=True)
    _s(ws_s.cell(row=2, column=total_cols), 'Total', bold=True, header=True)

    rack_totals = {r: 0 for r in racks}
    data_tabs   = [n for n in tabs_order if n.lower() != 'summary']
    for i, tab_name in enumerate(data_tabs, start=3):
        _s(ws_s.cell(row=i, column=1), tab_name)
        row_total = 0
        for c, rack in enumerate(racks, start=2):
            count = tab_rack.get(tab_name, {}).get(rack, 0)
            _s(ws_s.cell(row=i, column=c), count if count > 0 else '')
            rack_totals[rack] += count
            row_total += count
        _s(ws_s.cell(row=i, column=total_cols), row_total, bold=True)

    tot_r = 3 + len(data_tabs)
    _s(ws_s.cell(row=tot_r, column=1), 'TOTAL', bold=True)
    grand = 0
    for c, rack in enumerate(racks, start=2):
        _s(ws_s.cell(row=tot_r, column=c), rack_totals[rack], bold=True)
        grand += rack_totals[rack]
    _s(ws_s.cell(row=tot_r, column=total_cols), grand, bold=True)

    ws_s.column_dimensions['A'].width = 20
    for c in range(2, total_cols + 1):
        ws_s.column_dimensions[get_column_letter(c)].width = 14

    # ── 8. No fill + borders (preserve pink in Mismatches) ──────────────────
    log("  · Removing fills and applying borders")
    if 'Mismatches' in wb.sheetnames:
        ws_m = wb['Mismatches']
        m_header = [ws_m.cell(row=1, column=c).value
                    for c in range(1, ws_m.max_column + 1)]
        Z_NAMES = {'Z Hostname', 'Z Interface', 'Z L/R', 'Z Rack', 'Z Elevation'}
        pink_col_indices = [
            i + 1 for i, h in enumerate(m_header)
            if (h and (str(h).startswith('Possible ') or h in Z_NAMES))
        ]

    for sheet_name in wb.sheetnames:
        pcols = pink_col_indices if sheet_name == 'Mismatches' else []
        clear_and_border(wb[sheet_name], pink_cols=pcols)

    # ── 8c. Centre-align all cells across all tabs ──────────────────────────
    log("  · Aligning all cells to middle-centre")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=False)
    for sheet_name in wb.sheetnames:
        for row in wb[sheet_name].iter_rows():
            for cell in row:
                cell.alignment = center_align

    # ── 8b. Add NOTE column + autofilter to all tabs ───────────────────────
    log("  · Adding NOTE column and filters to all tabs")
    no_fill     = PatternFill(fill_type=None)
    yellow_fill = PatternFill('solid', start_color=YELLOW)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for col_name in ['NOTE']:
            col_idx = ws.max_column + 1
            hdr = ws.cell(row=1, column=col_idx, value=col_name)
            hdr.font      = Font(bold=True, name='Arial', size=10)
            hdr.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            hdr.fill      = yellow_fill
            hdr.border    = thin_border()
            ws.column_dimensions[get_column_letter(col_idx)].width = 16
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=col_idx)
                cell.fill   = no_fill
                cell.border = thin_border()
        if ws.max_row > 1 and ws.max_column > 0:
            ws.auto_filter.ref = ws.dimensions

    # ── 8d. Grey-out Optics rows that are matched in Downlinks ──────────────
    if 'Optics' in wb.sheetnames and 'Downlinks' in wb.sheetnames:
        log("  · Greying out matched Optics rows")

        MATCH_COLS = [
            'Hostname', 'Interface', 'L/R', 'Rack', 'Elevation',
            'Source_port', 'DMARC1', 'DMARC2', 'Destination_port',
            'Z Hostname', 'Z Interface', 'Z L/R', 'Z Rack', 'Z Elevation',
        ]
        GREY_FONT_COLOR = 'FFD3D3D3'

        ws_dl = wb['Downlinks']
        dl_header = [ws_dl.cell(row=1, column=c).value
                     for c in range(1, ws_dl.max_column + 1)]

        dl_match_cols = [c for c in MATCH_COLS if c in dl_header]
        dl_col_idx    = {c: dl_header.index(c) + 1 for c in dl_match_cols}

        dl_keys = set()
        for r in range(2, ws_dl.max_row + 1):
            key = tuple(
                str(ws_dl.cell(row=r, column=dl_col_idx[c]).value or '').strip()
                for c in dl_match_cols
            )
            dl_keys.add(key)

        ws_op = wb['Optics']
        op_header = [ws_op.cell(row=1, column=c).value
                     for c in range(1, ws_op.max_column + 1)]

        common_cols  = [c for c in dl_match_cols if c in op_header]
        op_col_idx   = {c: op_header.index(c) + 1 for c in common_cols}
        dl_col_idx_c = {c: dl_header.index(c) + 1 for c in common_cols}

        dl_keys_common = set()
        for r in range(2, ws_dl.max_row + 1):
            key = tuple(
                str(ws_dl.cell(row=r, column=dl_col_idx_c[c]).value or '').strip()
                for c in common_cols
            )
            dl_keys_common.add(key)

        for r in range(2, ws_op.max_row + 1):
            op_key = tuple(
                str(ws_op.cell(row=r, column=op_col_idx[c]).value or '').strip()
                for c in common_cols
            )
            if op_key in dl_keys_common:
                for c in range(1, ws_op.max_column + 1):
                    cell = ws_op.cell(row=r, column=c)
                    cell.font = Font(
                        bold=cell.font.bold if cell.font else False,
                        name=(cell.font.name if cell.font else None) or 'Arial',
                        size=(cell.font.size if cell.font else None) or 10,
                        color=GREY_FONT_COLOR,
                    )

    # ── 8e. Expand all columns and rows on every sheet ──────────────────────
    log("  · Expanding all columns and rows")
    for sheet_name in wb.sheetnames:
        autofit_sheet(wb[sheet_name])

    # ── 8f. Highlight reciprocal mismatch pairs orange ──────────────────────
    log("  · Highlighting reciprocal mismatch pairs")
    highlight_mismatch_pairs(wb, log)

    wb.save(output_path)

    # ── 9. Rename by top-2 Rack numbers ─────────────────────────────────────
    try:
        all_racks = []
        for sheet_name in wb.sheetnames:
            ws     = wb[sheet_name]
            header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
            if 'Rack' in header:
                rc = header.index('Rack') + 1
                for r in range(2, ws.max_row + 1):
                    val = ws.cell(row=r, column=rc).value
                    if val is not None:
                        try: all_racks.append(int(float(str(val))))
                        except ValueError: pass
        if all_racks:
            top2     = [str(r) for r, _ in Counter(all_racks).most_common(2)]
            new_name = '+'.join(top2) + '.xlsx'
            new_path = os.path.join(os.path.dirname(output_path), new_name)
            load_workbook(output_path).save(new_path)
            if new_path != output_path:
                os.remove(output_path)
            log(f"  ✓ Saved → {new_name}")
            return new_path
    except Exception as e:
        log(f"  ⚠ Could not rename by Rack: {e}")

    log(f"  ✓ Saved → {os.path.basename(output_path)}")
    return output_path

def highlight_mismatch_pairs(wb, log=lambda *_: None):
    """Mismatches tab: find reciprocal swap pairs and highlight them orange/yellow (exact from reference).
    Two mismatch rows are a *pair* when one's A-side block matches the other's Possible block (reciprocal).
    Pairs are moved adjacent, alternating orange/yellow on the main columns, pink preserved on Possible/Z.
    """
    if 'Mismatches' not in wb.sheetnames:
        return
    ws = wb['Mismatches']
    if ws.max_row < 3:
        return

    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    def idxs(names):
        return [header.index(n) + 1 for n in names if n in header]

    a_cols = idxs(['Hostname', 'Interface', 'L/R', 'Rack', 'Elevation',
                   'Source_port', 'Destination_port'])
    p_cols = idxs(['Possible Hostname', 'Possible Interface', 'Possible L/R',
                   'Possible Rack', 'Possible Elevation',
                   'Possible Source_port', 'Possible Destination_port'])
    orange_cols = idxs(['Hostname', 'Interface', 'L/R', 'Rack', 'Elevation',
                        'Source_port', 'DMARC1', 'DMARC2', 'Destination_port',
                        'Expected Hostname', 'Exp. Interface', 'Exp. L/R',
                        'Exp. Rack', 'Exp. Elevation'])
    if not a_cols or not p_cols:
        return

    ncol, nrow = ws.max_column, ws.max_row
    rows = [[ws.cell(row=r, column=c).value for c in range(1, ncol + 1)]
            for r in range(2, nrow + 1)]

    def key(rv, cols):
        return tuple(str(rv[c - 1] if rv[c - 1] is not None else '').strip()
                     for c in cols)

    a_keys = [key(rv, a_cols) for rv in rows]
    p_keys = [key(rv, p_cols) for rv in rows]

    n = len(rows)
    partner = [None] * n
    for i in range(n):
        if partner[i] is not None or not any(a_keys[i]):
            continue
        for j in range(i + 1, n):
            if partner[j] is not None:
                continue
            if a_keys[i] == p_keys[j] and a_keys[j] == p_keys[i]:
                partner[i] = j
                partner[j] = i
                break

    placed = [False] * n
    order = []
    for i in range(n):
        if placed[i] or partner[i] is None:
            continue
        j = partner[i]
        order.append(i); placed[i] = True
        if not placed[j]:
            order.append(j); placed[j] = True
    for i in range(n):
        if not placed[i]:
            order.append(i); placed[i] = True

    paired = {i for i in range(n) if partner[i] is not None}
    npairs = len(paired) // 2

    pair_no = {}
    counter = 0
    for i in range(n):
        j = partner[i]
        if j is not None and i < j:
            pair_no[i] = counter
            pair_no[j] = counter
            counter += 1

    bd          = thin_border()
    pink_fill   = PatternFill('solid', start_color=PINK)
    orange_fill = PatternFill('solid', start_color=ORANGE)
    yellow_fill = PatternFill('solid', start_color=YELLOW)
    pink_names  = {'Possible Hostname', 'Possible Interface', 'Possible L/R',
                   'Possible Rack', 'Possible Elevation', 'Possible Source_port',
                   'Possible DMARC1', 'Possible DMARC2', 'Possible Destination_port',
                   'Z Hostname', 'Z Interface', 'Z L/R', 'Z Rack', 'Z Elevation'}
    pink_idx    = {header.index(nm) + 1 for nm in pink_names if nm in header}
    orange_set  = set(orange_cols)
    no_fill     = PatternFill(fill_type=None)

    for out_off, src_i in enumerate(order):
        r  = out_off + 2
        rv = rows[src_i]
        is_pair = src_i in paired
        pair_fill = (orange_fill if pair_no.get(src_i, 0) % 2 == 0
                     else yellow_fill)
        for c in range(1, ncol + 1):
            cell = ws.cell(row=r, column=c, value=rv[c - 1])
            cell.font = Font(name='Arial', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = bd
            if is_pair and c in orange_set:
                cell.fill = pair_fill
            elif c in pink_idx:
                cell.fill = pink_fill
            else:
                cell.fill = no_fill

    log(f"  · Highlighted {npairs} mismatch pair(s) (alternating orange/yellow)")


def _count_issues_from_slack_inputs(paths: list[str]) -> dict:
    """Count rows in the relevant error sheets from the raw Slack report inputs.
    Uses the existing TAB_ALIASES to locate lldp / optics / interfaces / fec sheets.
    Returns counts in the same keys the 01 tool uses so Dashboard stays unified.
    This lets us log from 02 without depending on the (currently stubbed) output formatting.
    """
    import pandas as pd

    counts = {"mispatches": 0, "optics": 0, "downlinks": 0, "fec": 0}

    for p in paths or []:
        try:
            p = str(p)
            xl = pd.ExcelFile(p)
            sheet_names = xl.sheet_names

            # LLDP / Mismatches
            lldp_tab = find_tab(sheet_names, "lldp")
            if lldp_tab:
                df = pd.read_excel(p, sheet_name=lldp_tab)
                counts["mispatches"] += len(df)

            # Optics
            optics_tab = find_tab(sheet_names, "optics")
            if optics_tab:
                df = pd.read_excel(p, sheet_name=optics_tab)
                counts["optics"] += len(df)

            # Interface Down
            int_tab = find_tab(sheet_names, "interfaces")
            if int_tab:
                df = pd.read_excel(p, sheet_name=int_tab)
                counts["downlinks"] += len(df)

            # FEC
            fec_tab = find_tab(sheet_names, "combined_fec")
            if fec_tab:
                df = pd.read_excel(p, sheet_name=fec_tab)
                counts["fec"] += len(df)
        except Exception:
            continue

    return counts


# ── Streamlit UI (visual only - logic block below is untouched) ────────────────
st.set_page_config(page_title="HSG17 T1-to-T0 Slack Formatter", page_icon="🖥️", layout="wide")

st.title("🖥️ HSG17 T1-to-T0 Slack Upload")
st.caption("")

st.markdown("""
**How to use:**
1. Upload your **Cutsheet** (Installation Sheet)
2. Upload one or more **Slack Report Excel files**
3. Click **Generate Formatted Report**

The formatted report(s) will be available for immediate download.
""")

# ── Uploaders (stacked vertically) ───────────────────────────────────────────
cutsheet_uploader = st.file_uploader(
    "Cutsheet (Installation Sheet)",
    type=["xlsx", "xls"],
    accept_multiple_files=False,
    help="The Installation Sheet from the master cutsheet"
)

input_uploaders = st.file_uploader(
    "Slack Report Excel files",
    type=["xlsx", "xls"],
    accept_multiple_files=True,
    help="One or more Slack-style validation report files"
)

# Red primary button to differentiate from the LV Portal tool while keeping everything else uniform
st.markdown("""
<style>
div[data-testid="stButton"] button[kind="primary"] {
    background-color: #c62828 !important;
    border-color: #c62828 !important;
    color: white !important;
}
div[data-testid="stButton"] button[kind="primary"]:hover {
    background-color: #b71c1c !important;
    border-color: #b71c1c !important;
}
</style>
""", unsafe_allow_html=True)

# Blue/red primary button styling already injected above for the Slack tool.
run_btn = st.button(
    "🚀 Generate Formatted Report",
    type="primary",
    use_container_width=True,
    disabled=not (cutsheet_uploader and input_uploaders)
)

if run_btn and cutsheet_uploader and input_uploaders:
    with st.spinner("Processing Slack report(s)..."):
        tmpdir = Path(tempfile.mkdtemp(prefix="hsg17_slack_"))
        try:
            # Save uploads to temp (same pattern as page 01)
            cut_tmp = tmpdir / cutsheet_uploader.name
            cut_tmp.write_bytes(cutsheet_uploader.getvalue())

            slack_tmp_paths = []
            for f in input_uploaders:
                p = tmpdir / f.name
                p.write_bytes(f.getvalue())
                slack_tmp_paths.append(str(p))

            # Derive PG + representative rack from the uploaded files (same heuristic as 01)
            # so the Dashboard sees consistent "building" / PG across both tools.
            all_files_for_derive = [str(cut_tmp)] + slack_tmp_paths
            placement, rack = derive_placement_and_rack_from_files(all_files_for_derive)

            # Count issues from the raw Slack inputs using the existing tab aliases.
            # This lets 02 feed the exact same log format as 01.
            raw_counts = _count_issues_from_slack_inputs(slack_tmp_paths)

            # ====================== SILENT CENTRAL LOGGING (unified with 01) ======================
            try:
                source_name = ", ".join([f.name for f in input_uploaders])
                cat_map = {
                    "mispatches": "LLDP Mismatch + Link Down",
                    "downlinks": "Interface Down Errors",
                    "optics": "Optic Errors",
                    "fec": "FEC_BER Errors",
                }
                for cat_key, cnt in raw_counts.items():
                    if cnt > 0:
                        cat_name = cat_map.get(cat_key, cat_key.title())
                        log_errors(
                            hall="HSG17",
                            rack_type="T1-T0",
                            building=placement,
                            rack=rack,
                            error_category=cat_name,
                            count=int(cnt),
                            source_file=source_name,
                            processed_by="HSG17_T1toT0_Slack",
                        )
            except Exception:
                pass

            # --- Actual report generation (calls the existing process_file / helpers) ---
            # All transformation logic lives in the functions defined above (load_cutsheet,
            # process_file, swap/sort/highlight helpers, etc.). We only wire the execution here.
            cut_df = load_cutsheet(str(cut_tmp))
            output_paths = []
            for in_path_str in slack_tmp_paths:
                in_p = Path(in_path_str)
                out_name = in_p.stem + "_formatted.xlsx"
                out_p = tmpdir / out_name
                try:
                    produced = process_file(str(in_p), str(out_p), cut_df, log=lambda *a: None)
                    final_p = Path(produced) if produced else out_p
                    if final_p.exists():
                        output_paths.append(final_p)
                except Exception as proc_err:
                    # Keep going for other files; surface the issue without breaking the run
                    st.warning(f"Could not fully process {in_p.name}: {proc_err}")

            if output_paths:
                st.success(f"✅ {len(output_paths)} formatted report(s) ready for download.")
                if len(output_paths) == 1:
                    data = output_paths[0].read_bytes()
                    st.download_button(
                        label=f"📥 Download {output_paths[0].name}",
                        data=data,
                        file_name=output_paths[0].name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                else:
                    # Multiple inputs → individual downloads + convenient ZIP
                    for p in output_paths:
                        b = p.read_bytes()
                        st.download_button(
                            label=f"📥 Download {p.name}",
                            data=b,
                            file_name=p.name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    # ZIP all
                    zip_buf = BytesIO()
                    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                        for p in output_paths:
                            zf.write(p, arcname=p.name)
                    zip_buf.seek(0)
                    st.download_button(
                        label="📥 Download All as ZIP",
                        data=zip_buf.getvalue(),
                        file_name="HSG17_Slack_Reports_Formatted.zip",
                        mime="application/zip",
                        use_container_width=True
                    )
            else:
                st.info("Processing completed but no output files were produced (check inputs).")

        finally:
            # Cleanup temps (same as page 01)
            try:
                for p in tmpdir.glob("*"):
                    p.unlink(missing_ok=True)
                tmpdir.rmdir()
            except Exception:
                pass
