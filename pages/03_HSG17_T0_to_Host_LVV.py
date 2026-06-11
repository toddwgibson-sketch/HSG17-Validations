#!/usr/bin/env python3
"""
HSG17 T0-to-Host LVV Formatter — Streamlit page

Core formatter logic lives in:
    utils/t0_to_host_formatter.py
(exactly the logic from the reference lvv_portal_formatter_T0toHOST)

This page is the UI wrapper + HSG17-specific silent logging to the central
validation_error_log (Placement Group derivation + counts for the Dashboard).
"""

import sys
import tempfile
import zipfile
from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from utils.data_logger import log_errors
from utils.hsg17_models import derive_placement_and_rack_from_files
from utils.t0_to_host_formatter import format_report


# --- Streamlit UI -----------------------------------------------------------
st.set_page_config(page_title="HSG17 T0-to-Host LVV Formatter", page_icon="🖥️", layout="wide")

st.title("🖥️ HSG17 T0-to-Host LVV Portal Upload")
st.caption("")

st.markdown("""
**How to use:**
1. Upload your **LV Portal Validation Result(s)** (.xlsx / .xlsm) — the validationFailureResults file(s) containing T0-to-Host error sheets
2. Upload the **Combined Cutsheet / Allconnections** file(s)
3. Click **Generate Formatted Report**

The formatted report(s) will be available for immediate download.
The counts are silently logged for the HSG17 Dashboard (same as the other tools).
""")

# Pink primary button (distinct for T0-to-Host LVV)
st.markdown("""
<style>
div[data-testid="stButton"] button[kind="primary"] {
    background-color: #e91e63 !important;
    border-color: #e91e63 !important;
    color: white !important;
}
div[data-testid="stButton"] button[kind="primary"]:hover {
    background-color: #c2185b !important;
    border-color: #c2185b !important;
}
</style>
""", unsafe_allow_html=True)

# Larger labels (consistent with recent updates on other pages)
st.markdown(
    "<div style='font-size: 1.25rem; font-weight: 600; margin-bottom: 0.15rem;'>"
    "LV Portal Validation Result(s) (.xlsx / .xlsm)"
    "</div>",
    unsafe_allow_html=True,
)
validation_files = st.file_uploader(
    "LV Portal Validation Result(s) (.xlsx / .xlsm)",
    type=["xlsx", "xlsm"],
    accept_multiple_files=True,
    label_visibility="collapsed",
    help="The validation result export(s) containing T0-to-Host error sheets (LLDP, Optic, FEC, Interface Down, etc.)"
)

st.markdown(
    "<div style='font-size: 1.25rem; font-weight: 600; margin-bottom: 0.15rem;'>"
    "Combined Cutsheet(s) / Allconnections"
    "</div>",
    unsafe_allow_html=True,
)
cutsheet_files = st.file_uploader(
    "Combined Cutsheet(s) / Allconnections",
    type=["xlsx", "xlsm"],
    accept_multiple_files=True,
    label_visibility="collapsed",
    help="The QFABT0toHOST_allconnections.xlsx or equivalent combined cutsheet"
)

run_btn = st.button(
    "🚀 Generate Formatted Report",
    type="primary",
    use_container_width=True,
    disabled=not (validation_files and cutsheet_files)
)

if run_btn and validation_files and cutsheet_files:
    with st.spinner("Processing T0-to-Host validation report(s)..."):
        tmpdir = Path(tempfile.mkdtemp(prefix="hsg17_t0host_"))
        try:
            # Save uploads
            val_tmp_paths = []
            for f in validation_files:
                p = tmpdir / f.name
                p.write_bytes(f.getvalue())
                val_tmp_paths.append(str(p))

            cut_tmp_paths = []
            for f in cutsheet_files:
                p = tmpdir / f.name
                p.write_bytes(f.getvalue())
                cut_tmp_paths.append(str(p))

            # Use first cutsheet for the combined lookups (typical usage)
            combined_path = cut_tmp_paths[0]

            output_paths = []
            for in_path_str in val_tmp_paths:
                in_p = Path(in_path_str)
                try:
                    produced = format_report(in_path_str, combined_path)
                    final_p = Path(produced) if produced else in_p
                    if final_p.exists():
                        output_paths.append(final_p)
                except Exception as proc_err:
                    st.warning(f"Could not fully process {in_p.name}: {proc_err}")

            # Logging (unified with the other tools)
            try:
                source_name = ", ".join([f.name for f in validation_files])
                placement, rack = derive_placement_and_rack_from_files([combined_path] + val_tmp_paths)

                for out_p in output_paths:
                    # Count rows in the main output sheets for dashboard categories
                    try:
                        wb = load_workbook(out_p)
                        cat_counts = {}
                        if "Mismatch" in wb.sheetnames:
                            cat_counts["LLDP Mismatch + Link Down"] = wb["Mismatch"].max_row - 1
                        if "Optic Errors" in wb.sheetnames:
                            cat_counts["Optic Errors"] = wb["Optic Errors"].max_row - 1
                        if "Fec Errors" in wb.sheetnames:
                            cat_counts["FEC_BER Errors"] = wb["Fec Errors"].max_row - 1
                        if "Interface Down Errors" in wb.sheetnames:
                            cat_counts["Interface Down Errors"] = wb["Interface Down Errors"].max_row - 1

                        for cat_name, cnt in cat_counts.items():
                            if cnt > 0:
                                log_errors(
                                    hall="HSG17",
                                    rack_type="T0-Host",
                                    building=placement,
                                    rack=rack,
                                    error_category=cat_name,
                                    count=int(cnt),
                                    source_file=source_name,
                                    processed_by="HSG17_T0toHost_LVV",
                                )
                    except Exception:
                        pass
            except Exception:
                pass

            if output_paths:
                st.success(f"✅ {len(output_paths)} formatted T0-to-Host report(s) ready for download.")
                if len(output_paths) == 1:
                    data = output_paths[0].read_bytes()
                    st.download_button(
                        label=f"📥 Download {output_paths[0].name}",
                        data=data,
                        file_name=output_paths[0].name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                else:
                    for p in output_paths:
                        b = p.read_bytes()
                        st.download_button(
                            label=f"📥 Download {p.name}",
                            data=b,
                            file_name=p.name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    zip_buf = BytesIO()
                    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                        for p in output_paths:
                            zf.write(p, arcname=p.name)
                    zip_buf.seek(0)
                    st.download_button(
                        label="📥 Download All as ZIP",
                        data=zip_buf.getvalue(),
                        file_name="HSG17_T0Host_Reports_Formatted.zip",
                        mime="application/zip",
                        use_container_width=True
                    )
            else:
                st.info("Processing completed but no output files were produced (check inputs).")

        finally:
            try:
                for p in tmpdir.glob("*"):
                    p.unlink(missing_ok=True)
                tmpdir.rmdir()
            except Exception:
                pass
