#!/usr/bin/env python3

import sys
from pathlib import Path

# Ensure project root is on sys.path for "from utils.xxx" when this page
# is exec'd by st.navigation (fixes ImportError on Streamlit Cloud).
_here = Path(__file__).resolve()
_root = _here.parent.parent if _here.parent.name == "pages" else _here.parent
if str(_root) not in sys.path:
    sys.path.insert(0, str(_root))

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from pathlib import Path
from datetime import datetime
from zoneinfo import ZoneInfo
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from utils.hsg17_models import is_gpu_rack, get_rack_type, extract_filtered_counts_from_summary
from utils.data_logger import backup_log, save_daily_snapshot

SYDNEY_TZ = ZoneInfo("Australia/Sydney")

st.set_page_config(
    page_title="HSG17 Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.title("📊 HSG17 Dashboard")
st.caption("Captures Current Status/Changes -1")

# ====================== DASHBOARD STYLING (cards + panels) ======================
st.markdown("""
<style>
    .section-header {
        font-size: 1.25rem;
        font-weight: 600;
        margin-top: 1.4rem;
        margin-bottom: 0.5rem;
        color: #e2e8f0;
        border-left: 4px solid #22d3ee;
        padding-left: 10px;
    }
    /* Fancy metric cards */
    .hsg17-metric-card {
        border-radius: 14px;
        padding: 16px 18px;
        color: white;
        box-shadow: 0 10px 15px -3px rgb(0 0 0 / 0.3), 0 4px 6px -4px rgb(0 0 0 / 0.3);
        min-height: 118px;
        position: relative;
        overflow: hidden;
    }
    .hsg17-metric-card .label {
        font-size: 0.72rem;
        opacity: 0.85;
        font-weight: 500;
        letter-spacing: 0.6px;
        text-transform: uppercase;
    }
    .hsg17-metric-card .value {
        font-size: 1.95rem;
        font-weight: 700;
        line-height: 1.05;
        margin-top: 2px;
    }
    .hsg17-metric-card .icon {
        font-size: 1.9rem;
        opacity: 0.85;
    }
    .hsg17-metric-card .progress {
        margin-top: 10px;
        height: 3px;
        background: rgba(255,255,255,0.22);
        border-radius: 999px;
        position: relative;
    }
    .hsg17-metric-card .progress-fill {
        height: 3px;
        background: rgba(255,255,255,0.95);
        border-radius: 999px;
    }
    .hsg17-metric-card .progress-dot {
        position: absolute;
        top: -1.5px;
        width: 6px;
        height: 6px;
        background: white;
        border-radius: 999px;
        box-shadow: 0 0 0 2px rgba(255,255,255,0.4);
    }
    .progress {
        margin: 4px 0 8px;
        height: 3px;
        background: rgba(255,255,255,0.25);
        border-radius: 999px;
        position: relative;
    }
    .progress-fill {
        height: 3px;
        background: rgba(255,255,255,0.95);
        border-radius: 999px;
    }
    .progress-dot {
        position: absolute;
        top: -1.5px;
        width: 6px;
        height: 6px;
        background: white;
        border-radius: 999px;
        box-shadow: 0 0 0 2px rgba(255,255,255,0.4);
    }
    .hsg17-metric-card .sub {
        font-size: 0.62rem;
        opacity: 0.65;
        margin-top: 3px;
    }
    /* PG breakdown cards - full gradient like exec snapshot cards, with outline */
    .hsg17-pg-card {
        border-radius: 10px;
        padding: 10px 12px;
        margin-bottom: 4px;
        color: white;
        box-shadow: 0 10px 15px -3px rgb(0 0 0 / 0.3), 0 4px 6px -4px rgb(0 0 0 / 0.3);
    }
    .hsg17-pg-card .pg-pill {
        background-color:#0f172a; 
        border:2px solid #22d3ee; 
        border-radius:9999px; 
        padding:2px 8px; 
        font-size:0.7rem; 
        font-weight:500; 
        color:#e0f2fe; 
        letter-spacing:0.5px;
    }
    /* Rack table panels */
    .rack-panel {
        background: #1e2937;
        border: 1px solid #334155;
        border-radius: 10px;
        padding: 8px 12px;
        margin-bottom: 12px;
    }
    /* Bottom panels style */
    .dashboard-panel {
        background: #1e2937;
        border: 1px solid #334155;
        border-radius: 12px;
        padding: 12px;
        margin-bottom: 8px;
    }

    /* Tone down bright red / harsh colors on sidebar filters (multiselect tags, date inputs) */
    .stSidebar .stMultiSelect [data-baseweb="tag"] {
        background-color: #1e3a8a !important;  /* muted blue to match dark theme */
        color: #e0f2fe !important;
        border: 1px solid #22d3ee !important;
        border-radius: 999px !important;
    }
    .stSidebar .stMultiSelect [data-baseweb="tag"]:hover {
        background-color: #1e40af !important;
    }
    .stSidebar .stMultiSelect [data-baseweb="tag"] > span {
        color: #e0f2fe !important;
    }
    /* Date input focus to use cyan instead of red */
    .stSidebar .stDateInput > div > div > div > input:focus {
        border-color: #22d3ee !important;
        box-shadow: 0 0 0 1px #22d3ee !important;
    }
    /* Mute sidebar filter labels and header for consistency */
    .stSidebar .stMultiSelect label, .stSidebar .stDateInput label {
        color: #94a3b8 !important;
    }

    /* Cool gradient buttons for download error log and reset dashboard */
    .stDownloadButton button, .stButton button {
        background: linear-gradient(135deg, #1e3a8a, #0369a1) !important;
        color: #e0f2fe !important;
        border: 1px solid #22d3ee !important;
        border-radius: 8px !important;
        padding: 0.6rem 1rem !important;
        font-weight: 500 !important;
        transition: all 0.2s ease !important;
    }
    .stDownloadButton button:hover, .stButton button:hover {
        background: linear-gradient(135deg, #0369a1, #1e3a8a) !important;
        box-shadow: 0 0 0 2px #22d3ee !important;
        transform: translateY(-1px) !important;
    }
    .stButton button[kind="secondary"] {
        background: linear-gradient(135deg, #334155, #1e2937) !important;
        border: 1px solid #64748b !important;
    }
</style>
""", unsafe_allow_html=True)


# ====================== HELPER FUNCTIONS (moved up so we can use them for daily snapshots) ======================

def get_latest_snapshot(dataframe: pd.DataFrame) -> pd.DataFrame:
    """Most recent entry per (hall, building, rack, error_category).

    Rack is included in the key so that multiple racks reporting under the same
    Placement Group (e.g. GPU racks 2507 and 2508 both in PG11) each keep their
    own latest counts. The PG-level "Error Breakdown" cards sum across racks
    belonging to the same building; the GPU Rack Breakdown section renders a
    dedicated card per GPU rack.
    """
    if dataframe.empty:
        return dataframe
    group_keys = ['hall', 'building', 'error_category']
    if 'rack' in dataframe.columns:
        group_keys = ['hall', 'building', 'rack', 'error_category']
    return (
        dataframe.sort_values('timestamp')
        .groupby(group_keys, as_index=False)
        .last()
    )


def get_latest_with_deltas(dataframe: pd.DataFrame) -> pd.DataFrame:
    """Current vs previous run + delta for each (placement group + rack + category).

    Including rack in the key ensures 2507 vs 2508 (same PG) keep separate histories.
    Deltas are per-rack. PG cards later sum currents (and available deltas) across
    the racks that belong to the same building.
    """
    if dataframe.empty:
        return pd.DataFrame(columns=['hall', 'building', 'error_category', 'rack', 'current', 'previous', 'delta'])

    group_cols = ['hall', 'building', 'error_category']
    has_rack = 'rack' in dataframe.columns
    if has_rack:
        group_cols = ['hall', 'building', 'rack', 'error_category']

    records = []
    for keys, group in dataframe.sort_values('timestamp').groupby(group_cols):
        group = group.sort_values('timestamp')
        current = int(group.iloc[-1]['count'])

        if len(group) >= 2:
            previous = int(group.iloc[-2]['count'])
            delta = current - previous
        else:
            previous = None
            delta = None

        if pd.isna(delta):
            delta = None

        if has_rack:
            rack_val = keys[2]
            hall, bldg, _r, cat = keys
        else:
            rack_val = group.iloc[-1].get('rack', '') if 'rack' in group.columns else ''
            hall, bldg, cat = keys

        rack = str(rack_val) if pd.notna(rack_val) and str(rack_val).strip() != '' else ''

        records.append({
            'hall': hall,
            'building': bldg,
            'error_category': cat,
            'rack': rack,
            'current': current,
            'previous': previous,
            'delta': delta
        })
    return pd.DataFrame(records)


def generate_hsg17_summary_report(current_with_deltas: pd.DataFrame) -> bytes:
    """Generate a nicely formatted Excel report that matches the stakeholder example exactly.
    Only includes GPU racks (per-rack rows grouped by placement group / building),
    with group totals and group changes. Uses the exact same filtered data the dashboard cards use.
    """
    from collections import defaultdict

    if current_with_deltas is None or current_with_deltas.empty:
        gpu_df = pd.DataFrame()
    else:
        gpu_df = current_with_deltas[
            current_with_deltas['rack'].apply(lambda x: is_gpu_rack(x) if pd.notna(x) else False)
        ].copy()

    # Group by building (PG), collect per-rack per-category currents + group deltas
    groups = defaultdict(list)
    for _, row in gpu_df.iterrows():
        groups[row['building']].append({
            'rack': row['rack'],
            'category': row['error_category'],
            'current': int(row['current']) if pd.notna(row['current']) else 0,
            'delta': row.get('delta')
        })

    cat_order = ["LLDP Mismatch + Link Down", "Optic Errors", "FEC_BER Errors", "Interface Down Errors"]
    cat_headers = ["LLDP Mismatch Count", "Optic Errors Count", "FEC_BER Error Count", "Interface Down Count"]

    # Styles matching the example
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    blue_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    title_font = Font(name="Aptos Narrow", size=11)
    header_font = Font(name="Aptos Narrow", size=11, bold=True, color="FFFFFF")
    normal_font = Font(name="Aptos Narrow", size=11)
    red_bold_font = Font(name="Aptos Narrow", size=11, bold=True, color="FF0000")
    center_align = Alignment(horizontal="center", vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Title (merged B1:G2) - yellow banner
    ws.merge_cells('B1:G2')
    title_cell = ws['B1']
    title_cell.value = f"Validation Reporting Current as of the {datetime.now().strftime('%d/%m/%Y')}"
    title_cell.font = title_font
    title_cell.fill = yellow_fill
    title_cell.alignment = center_align
    for r in range(1, 3):
        for c in range(2, 8):
            ws.cell(row=r, column=c).border = thin_border
            ws.cell(row=r, column=c).fill = yellow_fill

    # Column widths (close match to example)
    ws.column_dimensions['A'].width = 16.43
    ws.column_dimensions['B'].width = 15.0
    ws.column_dimensions['C'].width = 16.86
    ws.column_dimensions['D'].width = 13.0
    ws.column_dimensions['E'].width = 13.0
    ws.column_dimensions['F'].width = 13.0
    ws.column_dimensions['G'].width = 13.14

    current_row = 3

    sorted_buildings = sorted(groups.keys())

    for bldg in sorted_buildings:
        items = groups[bldg]

        # Build per-rack data (default 0 for missing cats)
        rack_data = defaultdict(lambda: {cat: 0 for cat in cat_order})
        group_delta_sum = {cat: 0 for cat in cat_order}
        for it in items:
            r = it['rack']
            cat = it['category']
            if cat in rack_data[r]:
                rack_data[r][cat] = it['current']
            if pd.notna(it['delta']):
                group_delta_sum[cat] += int(it['delta'])

        # Unique sorted racks (numeric)
        sorted_racks = sorted(rack_data.keys(), key=lambda x: int(str(x).zfill(4)))

        # Header row for this group (blue)
        header_row = current_row
        ws.cell(row=current_row, column=2, value=bldg).font = header_font
        ws.cell(row=current_row, column=2).fill = blue_fill
        ws.cell(row=current_row, column=2).border = thin_border
        ws.cell(row=current_row, column=2).alignment = center_align

        for i, h in enumerate(cat_headers):
            cell = ws.cell(row=current_row, column=3 + i, value=h)
            cell.font = header_font
            cell.fill = blue_fill
            cell.border = thin_border
            cell.alignment = center_align

        total_h = ws.cell(row=current_row, column=7, value="Total Count")
        total_h.font = header_font
        total_h.fill = blue_fill
        total_h.border = thin_border
        total_h.alignment = center_align

        ws.row_dimensions[current_row].height = 30.0
        current_row += 1

        data_start_row = current_row

        # Per-rack data rows
        for rck in sorted_racks:
            ws.cell(row=current_row, column=2, value=rck).font = normal_font
            ws.cell(row=current_row, column=2).border = thin_border

            for i, cat in enumerate(cat_order):
                val = rack_data[rck][cat]
                cell = ws.cell(row=current_row, column=3 + i, value=val)
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

            # Total formula for the row
            tcell = ws.cell(row=current_row, column=7, value=f"=SUM(C{current_row}:F{current_row})")
            tcell.font = normal_font
            tcell.border = thin_border
            tcell.alignment = Alignment(horizontal="center")

            current_row += 1

        data_end_row = current_row - 1

        # Group Total row
        ws.cell(row=current_row, column=2, value=f"{bldg} Total").font = normal_font
        ws.cell(row=current_row, column=2).border = thin_border
        ws.cell(row=current_row, column=2).alignment = center_align

        for i in range(4):
            col_l = get_column_letter(3 + i)
            cell = ws.cell(row=current_row, column=3 + i, value=f"=SUM({col_l}{data_start_row}:{col_l}{data_end_row})")
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # Group total's Total Count column: sum the four category totals on *this row*
        # (not summing the per-rack G column values). Keeps row sums and column sums clean.
        gtotal = ws.cell(row=current_row, column=7, value=f"=SUM(C{current_row}:F{current_row})")
        gtotal.font = normal_font
        gtotal.border = thin_border
        gtotal.alignment = Alignment(horizontal="center")

        current_row += 1

        # Group Change row
        ws.cell(row=current_row, column=2, value=f"{bldg} Change").font = normal_font
        ws.cell(row=current_row, column=2).border = thin_border
        ws.cell(row=current_row, column=2).alignment = center_align

        for i, cat in enumerate(cat_order):
            d = group_delta_sum[cat]
            if d == 0:
                dstr = "(+0)"
                fnt = normal_font
            elif d > 0:
                dstr = f"(+{d})"
                fnt = red_bold_font
            else:
                dstr = f"({d})"
                fnt = normal_font
            cell = ws.cell(row=current_row, column=3 + i, value=dstr)
            cell.font = fnt
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # Total change column (match example style)
        tchange = ws.cell(row=current_row, column=7, value="(+0)")
        tchange.font = normal_font
        tchange.border = thin_border
        tchange.alignment = Alignment(horizontal="center")

        ws.row_dimensions[current_row].height = 8.25
        current_row += 1

        # Spacer row between groups
        current_row += 1

    # Write to bytes
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


DATA_FILE = Path(__file__).parent.parent / "data" / "validation_error_log.xlsx"

@st.cache_data(ttl=30)
def load_data():
    if not DATA_FILE.exists():
        return pd.DataFrame(columns=[
            "timestamp", "hall", "rack_type", "building", "rack",
            "error_category", "count", "source_file", "processed_by"
        ])
    df = pd.read_excel(DATA_FILE)
    # Upgrade schema if 'rack' column missing (for old logs without rack)
    if 'rack' not in df.columns:
        cols = list(df.columns)
        idx = cols.index('building') + 1 if 'building' in cols else len(cols)
        df.insert(idx, 'rack', '')
        df['rack'] = df['rack'].astype('object').fillna('').astype(str)
        try:
            df.to_excel(DATA_FILE, index=False)
            print(f"[DASHBOARD] Upgraded log schema with 'rack' column")
        except Exception as e:
            print(f"[DASHBOARD] Could not save upgraded log: {e}")
    df['timestamp'] = pd.to_datetime(df['timestamp'], errors='coerce')
    if not df.empty:
        # Treat stored timestamps as UTC (we log in UTC now).
        # Keep as UTC internally for consistency (snapshots, calculations).
        # UI display, LAST UPDATED (Sydney) ("time snapshot"), date filters, etc. are converted to hardcoded Sydney time.
        if getattr(df['timestamp'].dt, 'tz', None) is None:
            df['timestamp'] = df['timestamp'].dt.tz_localize('UTC')
    return df.dropna(subset=['timestamp'])

df = load_data()

# Keep UTC version for snapshots (consistent "right" timezone in snapshot files, independent of when/where dashboard is viewed)
hsg17_df_utc = df[df['hall'] == "HSG17"].copy()

# Hardcoded to Sydney time (Australia/Sydney) for all UI/display:
# - LAST UPDATED (Sydney) / time snapshot
# - Date range picker
# - Trend run times
# - Everything the user sees
hsg17_df = hsg17_df_utc.copy()
if not hsg17_df.empty:
    hsg17_df['timestamp'] = hsg17_df['timestamp'].dt.tz_convert(SYDNEY_TZ)

# Daily snapshot of the *full* current state (ignoring current sidebar filters)
# This gives a restore point for the "as of end of day" view the cards are showing.
# Runs once per calendar day on first load after midnight (using Sydney time for "today").
# We snapshot the UTC version so times in the .xlsx are always canonical UTC.
try:
    if not hsg17_df_utc.empty:
        full_latest = get_latest_snapshot(hsg17_df_utc)
        save_daily_snapshot(hsg17_df_utc, full_latest)

        # Also save the nice formatted stakeholder report for today (if not already)
        # Use Sydney time so the daily snapshot boundary matches the user's timezone.
        today_str = datetime.now(SYDNEY_TZ).strftime("%Y-%m-%d")
        snap_dir = Path(__file__).parent.parent / "data" / "snapshots"
        formatted = snap_dir / f"HSG17_Summary_Report_{today_str}.xlsx"
        if not formatted.exists():
            full_deltas = get_latest_with_deltas(hsg17_df_utc)
            report_bytes = generate_hsg17_summary_report(full_deltas)
            formatted.write_bytes(report_bytes)
            print(f"[SNAPSHOT] Saved daily formatted report: {formatted.name}")
except Exception as snap_err:
    print(f"[SNAPSHOT] Daily snapshot skipped/failed: {snap_err}")

if hsg17_df.empty:
    st.warning("No HSG17 data logged yet.")
    st.info("Run one of the HSG17 tools (01 T1-to-T0 LVV, 02 Slack, or 03 T0-to-Host LVV) to log issues. The dashboard shows current state + deltas per Placement Group.")
    st.stop()

# --- Sidebar Filters (more interactive development for the dashboard) ---
with st.sidebar:
    st.header("🔍 Filters")
    all_buildings = sorted(hsg17_df['building'].unique())
    selected_buildings = st.multiselect(
        "Placement Groups (Buildings)", 
        all_buildings, 
        default=all_buildings,
        help="Filter the view to specific Placement Groups (e.g. PG14 for rack 3110)"
    )
    all_cats = sorted(hsg17_df['error_category'].unique())
    selected_cats = st.multiselect(
        "Error Categories", 
        all_cats, 
        default=all_cats
    )

    min_date = hsg17_df['timestamp'].min().date()
    max_date = hsg17_df['timestamp'].max().date()
    date_range = st.date_input(
        "Consider logs from / to",
        value=(min_date, max_date),
        min_value=min_date,
        max_value=max_date,
        help="Only include log entries within this date range when computing current state and deltas."
    )

# Apply filters to the raw log data before computing snapshots/deltas
filtered_df = hsg17_df[
    hsg17_df['building'].isin(selected_buildings) &
    hsg17_df['error_category'].isin(selected_cats) &
    (hsg17_df['timestamp'].dt.date >= date_range[0]) &
    (hsg17_df['timestamp'].dt.date <= date_range[1])
].copy()

if filtered_df.empty:
    st.warning("No data matches the current filters. Adjust the Placement Groups, Categories or Date Range in the sidebar.")
    st.stop()

current = get_latest_snapshot(filtered_df)
current_with_deltas = get_latest_with_deltas(filtered_df)

if DATA_FILE.exists():
    st.markdown('<div class="dashboard-panel">', unsafe_allow_html=True)
    st.markdown("<div style='font-size:0.85rem; font-weight:600; color:#94a3b8; margin-bottom:4px;'>Data Management (unified — 01 LV Portal + 02 Slack + 03 T0-Host LVV)</div>", unsafe_allow_html=True)
    st.caption(f"Current HSG17 entries in log: **{len(hsg17_df)}**")
    # Summary report kept at top (the formatted one for stakeholders).
    # Error Log moved into Danger Zone for the restore/backup flow.
    try:
        report_bytes = generate_hsg17_summary_report(current_with_deltas)
        st.download_button(
            "📥 Download Summary Report",
            data=report_bytes,
            file_name=f"HSG17_Summary_Report_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
            help="Formatted GPU rack report (per the current filters)."
        )
    except Exception as e:
        st.warning(f"Could not generate summary report: {e}")

    st.markdown('</div>', unsafe_allow_html=True)

st.divider()

st.markdown('<div class="section-header">Executive Snapshot </div>', unsafe_allow_html=True)

col1, col2, col3, col4 = st.columns(4)

total_errors = int(current['count'].sum())
unique_blocks = current['building'].nunique()
active_rack_types = current['rack_type'].nunique()
last_ts = "—"
if not current.empty:
    ts = current['timestamp'].max()
    if pd.notna(ts):
        ts = pd.Timestamp(ts)
        if ts.tz is None:
            ts = ts.tz_localize('UTC')
        # Hardcoded Sydney TZ for the "time snapshot"
        local_ts = ts.tz_convert(SYDNEY_TZ)
        last_ts = local_ts.strftime("%Y-%m-%d %H:%M")

def _metric_card(label, value, icon, c1, c2, sub=""):
    return f'''
    <div class="hsg17-metric-card" style="background: linear-gradient(135deg, {c1}, {c2});">
        <div style="display:flex; justify-content:space-between; align-items:flex-start;">
            <div>
                <div class="label">{label}</div>
                <div class="value">{value}</div>
            </div>
            <div class="icon">{icon}</div>
        </div>
        <div class="progress">
            <div class="progress-fill" style="width: {min(92, 35 + (hash(str(value)) % 40))}%;"></div>
            <div class="progress-dot" style="left: {min(92, 35 + (hash(str(value)) % 40))}%;"></div>
        </div>
        <div class="sub">{sub}</div>
    </div>
    '''

with col1:
    st.markdown(_metric_card("TOTAL OPEN ISSUES (HSG17)", f"{total_errors:,}", "⚠️", "#0ea5e9", "#0369a1", "Current snapshot • filtered"), unsafe_allow_html=True)
with col2:
    st.markdown(_metric_card("PLACEMENT GROUPS WITH ISSUES", str(unique_blocks), "📍", "#f43f5e", "#9f1239", "Active PGs in view"), unsafe_allow_html=True)
with col3:
    st.markdown(_metric_card("RACK TYPES ACTIVE", str(active_rack_types), "🖥️", "#f59e0b", "#c2410f", "Unique rack types"), unsafe_allow_html=True)
with col4:
    st.markdown(_metric_card("Last Updated - Sydney Time", last_ts, "🕒", "#a855f7", "#6b21a8", "Latest processing run"), unsafe_allow_html=True)

st.divider()

st.markdown('<div class="section-header">Error Breakdown by Placement Group </div>', unsafe_allow_html=True)

CAT_COLORS = {
    "LLDP Mismatch + Link Down": "#e74c3c",
    "Optic Errors": "#3498db",
    "FEC_BER Errors": "#9b59b6",
    "Interface Down Errors": "#e67e22",
}
CAT_LABELS = {
    "LLDP Mismatch + Link Down": "LLDP Mismatch",
    "Optic Errors": "Optics",
    "FEC_BER Errors": "FEC BER",
    "Interface Down Errors": "Interface Down",
}

# Gradient palette for PG cards - cycle like the exec snapshot cards
GRADIENTS = [
    ("#0ea5e9", "#0369a1"),  # blue
    ("#f43f5e", "#9f1239"),  # rose
    ("#f59e0b", "#c2410f"),  # orange
    ("#a855f7", "#6b21a8"),  # purple
    ("#10b981", "#047857"),  # green
    ("#06b6d4", "#0e7490"),  # cyan
    ("#8b5cf6", "#5b21b6"),  # violet
]

if not current.empty:
    building_order = sorted(current['building'].unique())
    CARDS_PER_ROW = 5

    category_order = [c for c in CAT_LABELS.keys() if c in current['error_category'].unique()]

    for start_idx in range(0, len(building_order), CARDS_PER_ROW):
        row_buildings = building_order[start_idx : start_idx + CARDS_PER_ROW]
        cols = st.columns(CARDS_PER_ROW)

        for i, bldg in enumerate(row_buildings):
            bldg_deltas = current_with_deltas[current_with_deltas['building'] == bldg]

            # Sum across racks within the same PG (building). This allows PG11 to
            # correctly show combined numbers for 2507+2508 (and any future racks)
            # while still preserving per-rack latests for the GPU cards section.
            cat_current = {}
            cat_delta = {}
            for _, row in bldg_deltas.iterrows():
                cat = row['error_category']
                cat_current[cat] = cat_current.get(cat, 0) + row['current']
                d = row.get('delta')
                if pd.notna(d):
                    if cat in cat_delta and pd.notna(cat_delta.get(cat)):
                        cat_delta[cat] = cat_delta[cat] + d
                    else:
                        cat_delta[cat] = d

            bldg_total = sum(cat_current.values())

            valid_deltas = [d for d in cat_delta.values() if pd.notna(d)]
            total_delta = sum(valid_deltas) if valid_deltas else None

            total_str = str(bldg_total)
            if pd.notna(total_delta):
                delta_int = int(total_delta)
                delta_sign = f"({delta_int:+d})" if delta_int != 0 else ""
                delta_color = "green" if delta_int < 0 else "red"
                total_str += f" <span style='font-size:0.9rem; color:{delta_color};'>{delta_sign}</span>"

            bar_data = []
            for cat in category_order:
                val = cat_current.get(cat, 0)
                if val > 0:
                    bar_data.append({
                        "Category": CAT_LABELS.get(cat, cat),
                        "Count": val,
                        "Color": CAT_COLORS.get(cat, "#7f8c8d")
                    })

            grad_idx = (start_idx + i) % len(GRADIENTS)
            g1, g2 = GRADIENTS[grad_idx]

            # build the list html (white text for gradient bg)
            list_html = "<div style='margin-top:4px; font-size:0.82rem; line-height:1.25; color:#f8fafc;'>"
            for cat in category_order:
                label = CAT_LABELS.get(cat, cat)
                val = cat_current.get(cat, 0)
                d = cat_delta.get(cat)
                color = CAT_COLORS.get(cat, "#7f8c8d")

                delta_html = ""
                if pd.notna(d):
                    d_int = int(d)
                    delta_color = "green" if d_int < 0 else "red"
                    delta_str = f"({d_int:+d})"
                    delta_html = f" <span style='color:{delta_color}; font-size:0.75rem;'>{delta_str}</span>"

                list_html += f"<span style='color:{color}; font-weight:600'>■</span> {label}: <b>{val}</b>{delta_html}<br>"
            list_html += "</div>"

            with cols[i]:
                st.markdown(f'<div class="hsg17-pg-card" style="background: linear-gradient(135deg, {g1}, {g2}); color: white; border: none; box-shadow: 0 10px 15px -3px rgb(0 0 0 / 0.3), 0 4px 6px -4px rgb(0 0 0 / 0.3);">', unsafe_allow_html=True)

                st.markdown(f'<span class="pg-pill">{bldg}</span>', unsafe_allow_html=True)

                st.markdown(f"<div style='font-size:1.9rem; font-weight:700; line-height:1.1; margin-bottom:6px; color:#f8fafc;'>{total_str}</div>", unsafe_allow_html=True)

                if bar_data:
                    total_for_bar = sum(d['Count'] for d in bar_data)
                    bar_html = '<div style="height:16px; background: rgba(255,255,255,0.15); border-radius:999px; overflow:hidden; display:flex; margin:4px 0 6px; border:1px solid rgba(255,255,255,0.3);">'
                    for d in bar_data:
                        pct = (d['Count'] / total_for_bar * 100) if total_for_bar > 0 else 0
                        bar_html += f'<div style="width:{pct}%; background:{d["Color"]}; height:100%;"></div>'
                    bar_html += '</div>'
                    st.markdown(bar_html, unsafe_allow_html=True)

                st.markdown('<div style="background: rgba(0,0,0,0.25); border-radius: 6px; padding: 4px; margin-top: 4px;">', unsafe_allow_html=True)
                st.markdown(list_html, unsafe_allow_html=True)
                st.markdown('</div>', unsafe_allow_html=True)

                st.markdown('</div>', unsafe_allow_html=True)
else:
    st.info("No matching Placement Group data after filters. Try broadening your sidebar selections.")

st.divider()

# --- GPU-specific cards (distinct look from regular PG cards) ---
st.markdown('<div class="section-header">GPU Rack Breakdown </div>', unsafe_allow_html=True)

GPU_GRADIENTS = [
    ("#14b8a6", "#134e4b"),  # teal
    ("#4ade80", "#166534"),  # lime green
    ("#67e8f9", "#164e63"),  # cyan
    ("#a5f3fc", "#0e7490"),  # light cyan
    ("#5eead4", "#134e4b"),  # turquoise
]

if not current.empty:
    gpu_deltas = current_with_deltas[current_with_deltas['rack'].apply(is_gpu_rack)].copy() if 'rack' in current_with_deltas.columns else pd.DataFrame()
    if not gpu_deltas.empty:
        gpu_racks = sorted(gpu_deltas['rack'].dropna().unique())
        CARDS_PER_ROW = 5

        category_order = [c for c in CAT_LABELS.keys() if c in gpu_deltas['error_category'].unique()]

        # Assign a consistent colour per PG (building) for the card border and the bottom fill strip.
        # Refreshed palette: vibrant jewel tones (pinks, purples, blues, teals, greens) that work well
        # on dark slate backgrounds. No muddy yellows or clashing salmon tones.
        unique_buildings = sorted(gpu_deltas['building'].dropna().unique())
        PG_PILL_COLORS = [
            "#f472b6", "#c084fc", "#60a5fa", "#67e8f9",
            "#5eead4", "#4ade80", "#a78bfa", "#fb7185",
            "#38bdf8", "#34d399", "#818cf8", "#22d3ee"
        ]
        bldg_pill_color = {}
        for idx, b in enumerate(unique_buildings):
            bldg_pill_color[b] = PG_PILL_COLORS[idx % len(PG_PILL_COLORS)]

        for start_idx in range(0, len(gpu_racks), CARDS_PER_ROW):
            row_racks = gpu_racks[start_idx : start_idx + CARDS_PER_ROW]
            cols = st.columns(CARDS_PER_ROW)

            for i, rack in enumerate(row_racks):
                rack_deltas = gpu_deltas[gpu_deltas['rack'] == rack]
                bldg = rack_deltas['building'].iloc[0] if len(rack_deltas) > 0 else ""

                rack_type = get_rack_type(rack)

                # GPU cards: errors by rack only. No summing (even for racks in the same PG).
                # rack_deltas is already filtered to this one rack, so direct assignment per category.
                cat_current = {}
                cat_delta = {}
                for _, row in rack_deltas.iterrows():
                    cat = row['error_category']
                    cat_current[cat] = row['current']
                    cat_delta[cat] = row['delta']

                rack_total = sum(cat_current.values())

                valid_deltas = [d for d in cat_delta.values() if pd.notna(d)]
                total_delta = sum(valid_deltas) if valid_deltas else None

                total_str = str(rack_total)
                if pd.notna(total_delta):
                    delta_int = int(total_delta)
                    delta_sign = f"({delta_int:+d})" if delta_int != 0 else ""
                    delta_color = "green" if delta_int < 0 else "red"
                    total_str += f" <span style='font-size:0.9rem; color:{delta_color};'>{delta_sign}</span>"

                bar_data = []
                for cat in category_order:
                    val = cat_current.get(cat, 0)
                    if val > 0:
                        bar_data.append({
                            "Category": CAT_LABELS.get(cat, cat),
                            "Count": val,
                            "Color": CAT_COLORS.get(cat, "#7f8c8d")
                        })

                grad_idx = (start_idx + i) % len(GPU_GRADIENTS)
                g1, g2 = GPU_GRADIENTS[grad_idx]

                list_html = "<div style='margin-top:4px; font-size:0.82rem; line-height:1.25; color:#f8fafc;'>"
                for cat in category_order:
                    label = CAT_LABELS.get(cat, cat)
                    val = cat_current.get(cat, 0)
                    d = cat_delta.get(cat)
                    color = CAT_COLORS.get(cat, "#7f8c8d")

                    delta_html = ""
                    if pd.notna(d):
                        d_int = int(d)
                        delta_color = "green" if d_int < 0 else "red"
                        delta_str = f"({d_int:+d})"
                        delta_html = f" <span style='color:{delta_color}; font-size:0.75rem;'>{delta_str}</span>"

                    list_html += f"<span style='color:{color}; font-weight:600'>■</span> {label}: <b>{val}</b>{delta_html}<br>"
                list_html += "</div>"

                with cols[i]:
                    pill_text = f"🖥️ {rack_type} {rack}" if rack_type != "GPU" else f"🖥️ {rack}"
                    pill_color = bldg_pill_color.get(bldg, "#67e8f9")
                    # Outer: solid dark, per-PG border ONLY on left/right/bottom (top:none + padding:0 kills any top empty/yellow/coloured strip or cutoff above the pill). Dark pill (rack name + single emoji) sits flush at very top, fully inside the bordered area as requested.
                    st.markdown(f'<div class="hsg17-pg-card gpu-card" style="background: #0f172a; color: white; border-left: 2px solid {pill_color}; border-right: 2px solid {pill_color}; border-bottom: 2px solid {pill_color}; border-top: none; padding: 0; box-shadow: 0 10px 15px -3px rgb(0 0 0 / 0.3), 0 4px 6px -4px rgb(0 0 0 / 0.3); border-radius: 12px; overflow: hidden;">', unsafe_allow_html=True)

                    # Top dark rounded pill INSIDE the bordered card, rack name here (single 🖥️), flush top with 0 top margin + outer no top pad. 1.25rem font for the pill text.
                    st.markdown(f'<div style="background: #1e2937; border-radius: 9999px; padding: 8px 16px; margin: 0 8px 8px; display: flex; align-items: center; gap: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.3);"> <span style="font-size:1.25rem; font-weight:700; color: white;">{pill_text}</span> </div>', unsafe_allow_html=True)

                    # Main content area (sides padded here)
                    st.markdown('<div style="padding: 0 12px 8px;">', unsafe_allow_html=True)

                    if bldg:
                        st.markdown(f'<div style="font-size:0.7rem; opacity:0.6; margin-bottom:2px;">{bldg}</div>', unsafe_allow_html=True)

                    st.markdown(f"<div style='font-size:1.9rem; font-weight:700; line-height:1.1; margin-bottom:8px; color:#f8fafc;'>{total_str}</div>", unsafe_allow_html=True)

                    if bar_data:
                        total_for_bar = sum(d['Count'] for d in bar_data)
                        bar_html = '<div style="height:16px; background: rgba(255,255,255,0.15); border-radius:999px; overflow:hidden; display:flex; margin:4px 0 6px; border:1px solid rgba(255,255,255,0.3);">'
                        for d in bar_data:
                            pct = (d['Count'] / total_for_bar * 100) if total_for_bar > 0 else 0
                            bar_html += f'<div style="width:{pct}%; background:{d["Color"]}; height:100%;"></div>'
                        bar_html += '</div>'
                        st.markdown(bar_html, unsafe_allow_html=True)

                    st.markdown('<div style="background: rgba(0,0,0,0.25); border-radius: 6px; padding: 4px; margin-top: 4px; font-size:0.7rem;">', unsafe_allow_html=True)
                    st.markdown(list_html, unsafe_allow_html=True)
                    st.markdown('</div>', unsafe_allow_html=True)

                    # Bottom fill strip (per-PG color, matches the border/outline)
                    st.markdown(f'<div style="height: 18px; background: linear-gradient(90deg, {pill_color}, #0f172a); border-radius: 0 0 8px 8px; margin: 8px -12px -8px;"></div>', unsafe_allow_html=True)

                    st.markdown('</div>', unsafe_allow_html=True)  # close main padding
                    st.markdown('</div>', unsafe_allow_html=True)  # close outer card
else:
    st.info("No GPU rack data after filters.")

st.divider()

# Progress Trend moved to bottom as requested (full width, pie removed)
st.markdown('<div class="section-header">Progress Trend (Total Open Issues Over Time)</div>', unsafe_allow_html=True)

if not filtered_df.empty:
    # Compute the *total open issues at the time of each upload*.
    # For each historical run time, take all data up to that point, apply the same
    # "latest per (hall, building, rack, error_category)" logic used for the cards,
    # and sum the counts. This shows the actual current open issues as of that upload.
    run_times = sorted(filtered_df['timestamp'].unique())

    records = []
    for rt in run_times:
        hist = filtered_df[filtered_df['timestamp'] <= rt]
        latest_at_time = get_latest_snapshot(hist)
        total_open = int(latest_at_time['count'].sum()) if not latest_at_time.empty else 0
        records.append({
            'Run Time': rt,
            'Total Open Issues': total_open
        })

    trend = pd.DataFrame(records)
    trend = trend.sort_values('Run Time')
    trend['Run Time Str'] = trend['Run Time'].dt.strftime('%Y-%m-%d %H:%M')

    # Full-width trend panel
    st.markdown('<div class="dashboard-panel">', unsafe_allow_html=True)
    st.markdown("<div style='font-size:0.85rem; font-weight:600; color:#94a3b8; margin-bottom:4px;'>Total Open Issues at Time of Each Upload (filtered)</div>", unsafe_allow_html=True)
    # Trend chart - line + area showing cumulative current state
    fig = go.Figure()
    # Line + area for the total open issues (current state) as of each upload
    fig.add_trace(go.Scatter(
        x=trend['Run Time Str'],
        y=trend['Total Open Issues'],
        mode='lines+markers',
        line=dict(color='#22d3ee', width=3, shape='spline'),
        fill='tozeroy',
        fillcolor='rgba(34, 211, 238, 0.15)',
        marker=dict(size=6, color='#67e8f9', line=dict(width=1, color='#0b1120')),
        name='Total Open Issues'
    ))
    fig.update_layout(
        height=340,
        margin=dict(l=0, r=0, t=8, b=0),
        xaxis_title='Run Timestamp',
        yaxis_title='Total Open Issues',
        plot_bgcolor='#0b1120',
        paper_bgcolor='#0b1120',
        font_color='#e2e8f0',
        xaxis=dict(gridcolor='#1e2937', zerolinecolor='#334155'),
        yaxis=dict(gridcolor='#1e2937', zerolinecolor='#334155'),
        hovermode='x unified',
        showlegend=False
    )
    fig.update_traces(hovertemplate='%{x}<br>%{y} open issues<extra></extra>')

    # Export / Print pills for this panel
    tcols = st.columns([0.75, 0.25])
    with tcols[1]:
        bb1, bb2 = st.columns(2)
        with bb1:
            csv = trend[['Run Time Str', 'Total Open Issues']].rename(columns={'Run Time Str': 'Run Time'}).to_csv(index=False).encode('utf-8')
            st.download_button("⤵", data=csv, file_name="hsg17_trend.csv", mime="text/csv", key="trend_csv_dl", use_container_width=True, help="Export trend data")
        with bb2:
            if st.button("🖨", key="trend_print", use_container_width=True, help="Print chart"):
                st.toast("Use browser Print (Ctrl+P)")
    st.plotly_chart(fig, width="stretch", key="hsg17_trend", config={"displayModeBar": False})
    st.markdown('</div>', unsafe_allow_html=True)

    st.caption("Total open issues (current state) over time. For each upload we take all data up to that point, apply the latest-per-rack logic used by the cards, and sum the counts. This shows how the overall open issues evolved.")
else:
    st.info("No data for trend chart with current filters.")

# Moved the reset all the way to the bottom to reduce accidental clicks.
# Only relevant for testing.
st.divider()
st.markdown("### ⚠️ Danger Zone")
st.caption("Download Error Log here (the main data file for backup/restore). Reset (testing only) at the bottom.")

# Just the Error Log download - that's all that's needed for the "just in case" flow.
with open(DATA_FILE, "rb") as f:
    st.download_button(
        "📥 Download Error Log",
        data=f,
        file_name="HSG17_validation_error_log.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width="stretch",
        help="The full validation error log. Copy this file over the main one in data/ to restore everything."
    )

# Reset at the very bottom
confirm = st.checkbox("I confirm I want to permanently remove all HSG17 entries from the log.", key="confirm_reset")
if st.button("🗑️ Reset HSG17 Data", type="secondary", width="stretch", disabled=not confirm, key="reset_hsg17_data",
             help="Removes all HSG17 entries from the log so you can start fresh with real data. This only affects the dashboard feed."):
    try:
        if DATA_FILE.exists():
            backup_log()  # safety backup before any destructive action
            df = pd.read_excel(DATA_FILE)
            if 'hall' in df.columns:
                df_clean = df[df['hall'] != "HSG17"].copy()
            else:
                df_clean = df.copy()
            if df_clean.empty:
                # Recreate with proper headers
                df_clean = pd.DataFrame(columns=[
                    "timestamp", "hall", "rack_type", "building", 
                    "error_category", "count", "source_file", "processed_by"
                ])
            df_clean.to_excel(DATA_FILE, index=False)
            st.success("HSG17 dashboard data has been cleared. The page will refresh with empty state.")
            st.rerun()
        else:
            st.info("No data file found to clear.")
    except Exception as e:
        st.error(f"Failed to clear data: {e}")
