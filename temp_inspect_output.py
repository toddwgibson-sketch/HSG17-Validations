#!/usr/bin/env python3
"""Inspect the current formatted output sample to identify what needs to be made good."""

from openpyxl import load_workbook
from pathlib import Path

wb_path = Path('HSG17_T0_Host_sample_review.xlsx')
wb = load_workbook(wb_path)

print('=' * 60)
print('FORMATTED OUTPUT INSPECTION')
print('=' * 60)
print()

print('SHEETS:', wb.sheetnames)
print()

# === SUMMARY ===
print('--- SUMMARY TAB ---')
ws = wb['Summary']
print(f'A1: {ws["A1"].value}')
print(f'A2: {ws["A2"].value}')
print()
print('Content (from row 3):')
for idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
    print(f'  Row {idx}: {row}')
print()
print(f'Freeze: {ws.freeze_panes} | Filter: {ws.auto_filter.ref}')
print(f'Dimensions: {ws.dimensions}')
print()

# === LLDP ===
print('--- LLDP MISMATCH + LINK DOWN ---')
ws = wb['LLDP Mismatch + Link Down']
print(f'A1: {ws["A1"].value}')
print(f'A2: {ws["A2"].value}')
headers = [cell.value for cell in ws[3]]
print(f'Headers: {headers}')
print()

# Show first 10 data rows with key columns
print('Sample data rows (Block | Cluster | PP_Enriched | A-Name | A-Port | B-Name | B-Port | Expected-B):')
for r in range(4, min(14, ws.max_row + 1)):
    row = [ws.cell(row=r, column=c).value for c in range(1, 10)]
    block = row[0]
    cluster = row[1]
    pp = str(row[2])[:35] + '...' if row[2] and len(str(row[2])) > 35 else row[2]
    a_name = row[4]
    a_port = row[5]
    b_name = row[6]
    b_port = row[7]
    exp_b = row[8]
    print(f'  {block} | C{cluster} | {pp} | {a_name}:{a_port} -> {b_name}:{b_port} (exp {exp_b})')

print()
# Cluster grouping check
clusters = []
for r in range(4, ws.max_row + 1):
    c = ws.cell(row=r, column=2).value
    if c is not None:
        clusters.append(c)
grouped = all(clusters[i] <= clusters[i+1] for i in range(len(clusters)-1))
print(f'Cluster values (all): {clusters}')
print(f'Clusters are grouped (sorted non-decreasing): {grouped}')
print(f'Freeze: {ws.freeze_panes} | Filter: {ws.auto_filter.ref}')
print(f'Data rows: {ws.max_row - 3}')
print()

# === Other tabs quick check ===
for sheet in ['Optic Errors', 'FEC_BER Errors', 'Interface Down Errors']:
    ws = wb[sheet]
    headers = [cell.value for cell in ws[3]]
    print(f'--- {sheet} ---')
    print(f'Headers (first 6): {headers[:6]} ...')
    print(f'Has Block: {"Block" in headers}, PP_Enriched: {"PP_Enriched" in headers}')
    print(f'Freeze: {ws.freeze_panes} | Filter: {ws.auto_filter.ref}')
    print(f'Data rows: {ws.max_row - 3}')
    print()

print('=' * 60)
print('INSPECTION COMPLETE')
print('=' * 60)
