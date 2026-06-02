#!/usr/bin/env python3
r"""
HSG17 processor tests using fully synthetic data.

Run with:
    cd C:\Users\toddy\Documents\GitHub\HSG17-Validations
    python -m tests.test_hsg17_processor
    # or
    python tests/test_hsg17_processor.py

This exercises the full 6-stage pipeline + block derivation + clustering + enrichment
without needing any real Batam input files.
"""

import sys
from pathlib import Path
from io import BytesIO
from datetime import datetime

import pandas as pd
from openpyxl import Workbook

# Make imports work when run as script or module
sys.path.insert(0, str(Path(__file__).parent.parent))

from utils.hsg17_models import normalize_block, PG_TO_DH, RACK_TO_BLOCK, CANONICAL_BLOCKS
from utils.hsg17_t0_host import (
    process_hsg17_t0_host,
    extract_counts_for_logging,
    derive_block,
    cluster_mismatches,
    _derive_block_from_pp,
    _derive_block_from_rack,
    ingest,
    normalize,
    enrich_and_analyze,
)


def make_allconnections_df() -> pd.DataFrame:
    """Synthetic allconnections with PP labels and device/port info for enrichment + block derivation."""
    rows = [
        # DH-002 via PP label
        {"DeviceA Name": "hsg17-q2-p3-cpu-01", "DeviceA Port": "Eth1/1", "DeviceB Name": "hsg17-q2-p3-t0-a", "DeviceB Port": "Eth1/1",
         "Full Label": "PP.HSG17.3.DH2.R03.01-02", "EasyMark+ --- Patch Panels": "DH-002 / Rack 0807", "T0 Switch": "t0-leaf-01", "Cable Type": "OM4", "Cable Info": "cable-foo",
         "Rack A": "1009", "Elevation A": "3", "DMARC1": "DM1", "DMARC2": "DM2", "Z Rack": "9706", "Z Elevation": "9", "T1 Rack": "9706", "T1 Port": "swp13s0", "Interface": "swp31s0", "History": "initial", "Z Interface": "z-intf-foo", "T0 Switch Port": "t0p1"},
        {"DeviceA Name": "hsg17-q2-p3-cpu-02", "DeviceA Port": "Eth1/2", "DeviceB Name": "hsg17-q2-p3-t0-b", "DeviceB Port": "Eth1/2",
         "Full Label": "SomeLabel DH-002 extra", "EasyMark+ --- Patch Panels": "", "T0 Switch": "t0-leaf-02"},
        # DH-102 spines via rack in name + PP
        {"DeviceA Name": "hsg17-q2-p1-gpu-tray-03", "DeviceA Port": "Eth1/3", "DeviceB Name": "hsg17-q2-p1-t1-spine", "DeviceB Port": "Eth1/3",
         "Full Label": "DH102-PP-01", "EasyMark+ --- Patch Panels": "QFABT1 DH-102"},
        # DH-004 via rack prefix in allc (for secondary rule)
        {"DeviceA Name": "hsg17-q2-p4-cpu-09", "DeviceA Port": "Eth1/9", "DeviceB Name": "hsg17-q2-p4-t0-x", "DeviceB Port": "Eth1/9",
         "Full Label": "rack 2814 label", "EasyMark+ --- Patch Panels": ""},
        # IPR
        {"DeviceA Name": "hsg17-q2-ip-r01", "DeviceA Port": "Eth1/10", "DeviceB Name": "ipr-switch", "DeviceB Port": "1",
         "Full Label": "IPR-PG201-001", "EasyMark+ --- Patch Panels": ""},
    ]
    return pd.DataFrame(rows)


def make_cutsheet_error_sheet(name: str, num_rows: int = 6) -> pd.DataFrame:
    """Create a small synthetic error sheet with columns the processor expects."""
    base = {
        "Patch Panel Matrix": "",
        "Device A Name": "",
        "Device A Port": "",
        "Device B Name": "",
        "Device B Port": "",
        "Expected Device B Name": "",
        "Expected Device B Port": "",
        "Device A Rack": "",
        "Source Device Location": "",
    }
    rows = []
    for i in range(num_rows):
        r = base.copy()
        if name == "LLDP Mismatch + Link Down":
            # Mix of good block signals + some for clustering (swapped pairs)
            if i == 0:
                r.update({
                    "Patch Panel Matrix": "HSG17 DH-002 PP-03",
                    "Device A Name": "hsg17-q2-p3-cpu-01", "Device A Port": "Eth1/1",
                    "Device B Name": "hsg17-q2-p3-t0-a", "Device B Port": "Eth1/1",
                    "Expected Device B Name": "hsg17-q2-p3-t0-b", "Expected Device B Port": "Eth1/1",
                    "Device A Rack": "hsg17:0807:1",
                })
            elif i == 1:
                # pair of above for cluster (the swap side)
                r.update({
                    "Patch Panel Matrix": "HSG17 DH-002 PP-03",
                    "Device A Name": "hsg17-q2-p3-cpu-01", "Device A Port": "Eth1/1",
                    "Device B Name": "hsg17-q2-p3-t0-b", "Device B Port": "Eth1/1",
                    "Expected Device B Name": "hsg17-q2-p3-t0-a", "Expected Device B Port": "Eth1/1",
                    "Device A Rack": "hsg17:0807:1",
                })
            elif i == 2:
                r.update({
                    "Patch Panel Matrix": "DH-004 zone",
                    "Device A Name": "hsg17-q2-p4-cpu-09", "Device A Port": "Eth1/9",
                    "Device B Name": "hsg17-q2-p4-t0-x", "Device B Port": "Eth1/9",
                    "Expected Device B Name": "", "Expected Device B Port": "",
                    "Device A Rack": "hsg17:2814:9",
                })
            else:
                r.update({
                    "Patch Panel Matrix": f"DH-00{(i%3)+1}",
                    "Device A Name": f"comp-dh00{(i%3)+1}-{i}", "Device A Port": f"Eth1/{i}",
                    "Device B Name": f"t0-dh00{(i%3)+1}", "Device B Port": f"Eth1/{i}",
                    "Device A Rack": "hsg17:0907:3",
                })
        elif name == "Optic Errors":
            r.update({
                "Patch Panel Matrix": "DH-003 optic zone" if i % 2 == 0 else "DH-101 label",
                "Device A Name": f"gpu-tray-{i}", "Device A Port": f"Opt{i}",
                "Device B Name": f"t0-optic-{i}", "Device B Port": "1",
                "Device A Rack": "hsg17:0907:3" if i % 2 == 0 else "hsg17:2502:1",
            })
        elif name == "FEC_BER Errors":
            r.update({
                "Patch Panel Matrix": "DH-005 BER rack",
                "Device A Name": f"comp-ber-{i}", "Device A Port": f"Eth{i}",
                "Device B Name": "t0-ber", "Device B Port": "2",
                "Device A Rack": "hsg17:3010:4",
            })
        else:  # Interface Down Errors
            r.update({
                "Patch Panel Matrix": "DH-102 (Spines) down" if i == 0 else "AUX rack down",
                "Device A Name": "spine-if-01" if i == 0 else "aux-dev",
                "Device A Port": f"Eth1/{i}",
                "Device B Name": "down-t0",
                "Device B Port": "3",
                "Device A Rack": "hsg17:5710:2" if i == 0 else "5920",
            })
        rows.append(r)
    return pd.DataFrame(rows)


def make_cutsheet_xlsx_bytes() -> bytes:
    """Build an in-memory multi-sheet cutsheet xlsx matching what the real pre-classified file provides."""
    wb = Workbook()
    # Remove default
    if wb.active:
        wb.remove(wb.active)

    for sheet_name in ["LLDP Mismatch + Link Down", "Optic Errors", "FEC_BER Errors", "Interface Down Errors"]:
        ws = wb.create_sheet(sheet_name)
        df = make_cutsheet_error_sheet(sheet_name)
        # write header + data
        for r_idx, row in enumerate([list(df.columns)] + df.values.tolist(), start=1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val if pd.notna(val) else "")

    # Also a tiny Summary sheet (ignored for processing)
    ws_sum = wb.create_sheet("Summary")
    ws_sum["A1"] = "Pre-classified summary (ignored by processor)"
    ws_sum["A2"] = "LLDP"
    ws_sum["B2"] = 12

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def make_allconnections_xlsx_bytes() -> bytes:
    df = make_allconnections_df()
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="Sheet1")
    buf.seek(0)
    return buf.getvalue()


def test_block_derivation_standalone():
    print("== test_block_derivation_standalone ==")
    # Direct from PP
    assert _derive_block_from_pp("PP.HSG17.3.DH2.foo") == "DH-002"
    assert _derive_block_from_pp("DH-102-PP") == "DH-102 (Spines)"
    assert _derive_block_from_pp("DH4 label") == "DH-004"
    assert _derive_block_from_pp("random") is None

    # From rack
    assert _derive_block_from_rack("hsg17:5710:2") == "DH-102 (Spines)"
    assert _derive_block_from_rack("0807") == "DH-002"
    assert _derive_block_from_rack("42xx-foo") is None  # not in RACK_TO_BLOCK (falls back to PP or UNKNOWN)

    # Full derive_block priority
    row_pp = pd.Series({"Patch Panel Matrix": "DH-005 zone", "Device A Rack": "hsg17:0807:9"})
    assert derive_block(row_pp, ["Patch Panel Matrix"], ["Device A Rack"]) == "DH-005"

    row_rack = pd.Series({"Patch Panel Matrix": "", "Device A Rack": "hsg17:5807:1"})
    assert derive_block(row_rack, ["Patch Panel Matrix"], ["Device A Rack"]) == "DH-102 (Spines)"

    row_name = pd.Series({"Device A Name": "hsg17-q2-p*-t1-foo", "Patch Panel Matrix": ""})
    assert derive_block(row_name, ["Patch Panel Matrix"], ["Device A Rack"]) == "DH-102 (Spines)"

    row_ipr = pd.Series({"Device A Name": "hsg17-q2-ip-r5", "Patch Panel Matrix": ""})
    assert derive_block(row_ipr, ["Patch Panel Matrix"], ["Device A Rack"]) == "IPR (PG-201)"

    print("  block derivation: OK")


def test_normalize_block():
    print("== test_normalize_block ==")
    assert normalize_block("DH2") == "DH-002"
    assert normalize_block("dh-102") == "DH-102 (Spines)"
    assert normalize_block("DH.004") == "DH-004"
    assert normalize_block("hsg17:5708:3") is None  # this func is PP/name focused; rack handled elsewhere
    assert normalize_block(None) is None
    print("  normalize_block: OK")


def test_cluster_mismatches():
    print("== test_cluster_mismatches ==")
    df = pd.DataFrame([
        {"Device A Name": "c1", "Device A Port": "1", "Device B Name": "s1", "Device B Port": "p1",
         "Expected Device B Name": "s2", "Expected Device B Port": "p2"},
        {"Device A Name": "c1", "Device A Port": "1", "Device B Name": "s2", "Device B Port": "p2",
         "Expected Device B Name": "s1", "Expected Device B Port": "p1"},
        {"Device A Name": "c2", "Device A Port": "3", "Device B Name": "s3", "Device B Port": "p3",
         "Expected Device B Name": "", "Expected Device B Port": ""},
    ])
    clustered = cluster_mismatches(df)
    assert "Cluster" in clustered.columns
    # first two should share a cluster (connected via the expected swap)
    assert clustered.loc[0, "Cluster"] == clustered.loc[1, "Cluster"]
    # third is separate
    assert clustered.loc[2, "Cluster"] != clustered.loc[0, "Cluster"]
    print("  cluster_mismatches: OK (swap pair grouped)")


def test_full_pipeline_roundtrip():
    print("== test_full_pipeline_roundtrip ==")
    all_bytes = make_allconnections_xlsx_bytes()
    cuts_bytes = make_cutsheet_xlsx_bytes()

    result_bytes, filename = process_hsg17_t0_host(all_bytes, cuts_bytes, source_filename="test_run.xlsx")
    assert result_bytes, "expected non-empty bytes"
    assert "HSG17_T0_Host" in filename
    assert filename.endswith(".xlsx")

    # Load and inspect
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(result_bytes))
    sheets = wb.sheetnames
    assert "Summary" in sheets
    for cat in ["LLDP Mismatch + Link Down", "Optic Errors", "FEC_BER Errors", "Interface Down Errors"]:
        assert cat in sheets, f"missing sheet {cat}"

    # Check Summary has blocks and totals
    ws = wb["Summary"]
    # row 3 is header, row 4+ data, last TOTAL
    blocks_seen = set()
    total_row = None
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row and row[0]:
            if str(row[0]).upper() == "TOTAL":
                total_row = row
            else:
                blocks_seen.add(str(row[0]))

    assert len(blocks_seen) >= 1, "Summary should have at least one DH block"
    assert total_row is not None

    # Check one detailed sheet has Block promoted and for LLDP has Cluster + PP_Enriched possibly
    lldp = wb["LLDP Mismatch + Link Down"]
    # header is row 3
    headers = [cell.value for cell in lldp[3]]
    assert "Block" in headers
    assert "Cluster" in headers
    # New actionable columns added to make the report provide value (grouping + suggestion)
    assert "Cluster Size" in headers
    assert "Notes" in headers
    # The info the user needs in each sheet: Source_port, Rack, Elevation, DMARC1/2, Z Rack/Elev, T1 info, Interface, History, plus AllConn_*
    desired = ["Source_port", "Destination_port", "Rack", "Elevation", "Z_Interface", "T1_Rack", "Interface", "History", "T0 Switch Port", "Cable Info"]
    for d in desired:
        variants = [d, d.replace(' ', '_'), d.replace('_', ' ')]
        found = any(v in headers or any(v in str(h) for h in headers) for v in variants)
        assert found, f"Report must include {d} (or similar) from allconnections enrichment"
    assert any(h.startswith("AllConn_") for h in headers), "Final report xlsx must include AllConn_* columns with device info from allconnections"
    # PP_Enriched may or may not appear depending on join success; we check it was attempted
    # (in our synthetic it should succeed for some rows)
    if "PP_Enriched" not in headers:
        # still ok if join didn't trigger for this run, but we will assert in enrichment-specific test
        pass

    # Interface Down should have the full switch detection flag (value-add processing)
    iface = wb["Interface Down Errors"]
    iface_headers = [cell.value for cell in iface[3]]
    assert "Full Switch Down?" in iface_headers or "A Device Down Count" in iface_headers

    # Formatter polish: LLDP (and all error tabs) sorted by Block (primary for usability per sector), Cluster within block for pairs, plus freeze/filter
    assert lldp.freeze_panes == "D4"
    assert lldp.auto_filter.ref is not None and "A3:" in str(lldp.auto_filter.ref)
    # Verify sorted by Block primarily (clusters grouped within blocks is secondary and tested in cluster_mismatches)
    block_vals = []
    for row in lldp.iter_rows(min_row=4, max_col=1, values_only=True):
        if row[0] is not None:
            block_vals.append(row[0])
    assert all(block_vals[i] <= block_vals[i+1] or block_vals[i] == block_vals[i+1] for i in range(len(block_vals)-1) if block_vals[i] and block_vals[i+1]), "LLDP should be sorted by Block for report usability"

    # extract counts
    counts = extract_counts_for_logging(result_bytes)
    assert len(counts) >= 1
    for c in counts:
        assert "block" in c
        assert "LLDP Mismatch + Link Down" in c

    print(f"  full pipeline: OK (blocks={sorted(blocks_seen)}, filename={filename})")


def test_enrichment_and_pp_join():
    print("== test_enrichment_and_pp_join ==")
    all_bytes = make_allconnections_xlsx_bytes()
    cuts_bytes = make_cutsheet_xlsx_bytes()

    # Use lower level stages so we can inspect
    ingested = ingest(all_bytes, cuts_bytes)
    normalized = normalize(ingested)
    enriched = enrich_and_analyze(normalized)

    lldp = enriched.get("LLDP Mismatch + Link Down")
    assert lldp is not None and not lldp.empty
    assert "Block" in lldp.columns
    assert "Cluster" in lldp.columns
    assert "PP_Enriched" in lldp.columns

    # The info the user needs: Source_port, Rack, Elevation, Z/T1 info, History etc from allconnections
    desired = ["Source_port", "Destination_port", "Rack", "Elevation", "Z_Interface", "T1_Rack", "Interface", "History", "T0 Switch Port", "Cable Info"]
    for d in desired:
        variants = [d, d.replace(' ', '_'), d.replace('_', ' ')]
        found = any(v in lldp.columns or any(v in str(c) for c in lldp.columns) for v in variants)
        assert found, f"Expected {d} (or similar) from allconnections enrichment"
    assert any(c.startswith("AllConn_") for c in lldp.columns), "Expected AllConn_* device info columns from allconnections enrichment"

    # At least one row should have gotten a non-empty PP_Enriched from our synthetic allc
    non_empty = lldp["PP_Enriched"].astype(str).str.len() > 0
    assert non_empty.any(), "Expected at least one PP_Enriched value from the join"

    # Blocks should include DH-002 etc from our data
    blocks = set(lldp["Block"].dropna().astype(str).unique())
    assert any("DH-00" in b or "DH-102" in b for b in blocks), f"unexpected blocks: {blocks}"

    print("  enrichment + PP join: OK")


def test_ingest_error_on_missing_cutsheet():
    print("== test_ingest_error_on_missing_cutsheet ==")
    all_bytes = make_allconnections_xlsx_bytes()
    try:
        ingest(all_bytes, b"")
        assert False, "should have raised for empty cuts"
    except ValueError as e:
        assert "required" in str(e).lower() or "cutsheet" in str(e).lower()
    print("  ingest requires cutsheet: OK")


def main():
    print("HSG17 synthetic processor test suite")
    print(f"Python: {sys.version.split()[0]}  |  pandas: {pd.__version__}")
    print("-" * 50)

    test_block_derivation_standalone()
    test_normalize_block()
    test_cluster_mismatches()
    test_ingest_error_on_missing_cutsheet()
    test_enrichment_and_pp_join()
    test_full_pipeline_roundtrip()

    print("-" * 50)
    print("ALL TESTS PASSED ✓")


if __name__ == "__main__":
    main()
