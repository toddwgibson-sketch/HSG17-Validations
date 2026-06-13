import pandas as pd
from datetime import datetime, timezone
import time
import shutil
from pathlib import Path

LOG_FILE = Path(__file__).parent.parent / "data" / "validation_error_log.xlsx"

BACKUP_DIR = LOG_FILE.parent / "backups"
BACKUP_DIR.mkdir(parents=True, exist_ok=True)
MAX_BACKUPS = 30

SNAPSHOT_DIR = LOG_FILE.parent / "snapshots"
SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)

def _get_abs_log_path():
    return LOG_FILE.resolve()

def ensure_log_exists():
    LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    abs_path = _get_abs_log_path()

    if not LOG_FILE.exists():
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Error Log"
        headers = ["timestamp", "hall", "rack_type", "building", "rack",
                   "error_category", "count", "source_file", "processed_by"]
        ws.append(headers)
        wb.save(LOG_FILE)
        wb.close()
        print(f"[DATA_LOGGER] Created new error log at: {abs_path}")
    else:
        # Upgrade schema if 'rack' column missing (for backward compat with old logs)
        try:
            df = pd.read_excel(LOG_FILE)
            if 'rack' not in df.columns:
                # insert after 'building'
                cols = list(df.columns)
                idx = cols.index('building') + 1 if 'building' in cols else len(cols)
                df.insert(idx, 'rack', '')
                df['rack'] = df['rack'].astype('object').fillna('').astype(str)
                df.to_excel(LOG_FILE, index=False)
                print(f"[DATA_LOGGER] Upgraded log schema with 'rack' column at: {abs_path}")
        except Exception as e:
            print(f"[DATA_LOGGER] Schema upgrade check failed: {e}")
        print(f"[DATA_LOGGER] Using existing error log at: {abs_path}")

def backup_log():
    """Create a timestamped full backup of the log before any write.
    Prunes old backups to keep disk usage reasonable.
    Returns the backup_path on success, None on failure (e.g. if log doesn't exist yet).
    """
    if not LOG_FILE.exists():
        return None
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d_%H%M%S")
    backup_path = BACKUP_DIR / f"validation_error_log_{ts}.xlsx"
    try:
        shutil.copy2(LOG_FILE, backup_path)
        # Prune old backups
        backups = sorted(BACKUP_DIR.glob("validation_error_log_*.xlsx"))
        while len(backups) > MAX_BACKUPS:
            old = backups.pop(0)
            try:
                old.unlink()
            except Exception:
                pass
        print(f"[DATA_LOGGER] Backed up current log to {backup_path}")
        return backup_path
    except Exception as e:
        print(f"[DATA_LOGGER] Backup failed: {e}")
        return None


def save_daily_snapshot(full_hsg17_df: pd.DataFrame, latest_df: pd.DataFrame = None):
    """Save a once-per-day snapshot of the full log + the computed 'current state'
    (latest per rack/PG/category). This gives you a restore point for the end-of-day view.
    Called automatically from the dashboard.
    """
    if full_hsg17_df is None or full_hsg17_df.empty:
        return
    today = datetime.now().date().isoformat()

    # 1. Full log snapshot for the day (if not already saved today)
    full_log_snap = SNAPSHOT_DIR / f"validation_error_log_full_{today}.xlsx"
    if not full_log_snap.exists() and LOG_FILE.exists():
        try:
            shutil.copy2(LOG_FILE, full_log_snap)
            print(f"[SNAPSHOT] Saved full log snapshot for {today}")
        except Exception as e:
            print(f"[SNAPSHOT] Full log snapshot failed: {e}")

    # 2. Current state (deduped latest) as a simple table
    if latest_df is not None and not latest_df.empty:
        state_snap = SNAPSHOT_DIR / f"current_state_{today}.xlsx"
        if not state_snap.exists():
            try:
                latest_df.to_excel(state_snap, index=False)
                print(f"[SNAPSHOT] Saved current state snapshot for {today} ({len(latest_df)} rows)")
            except Exception as e:
                print(f"[SNAPSHOT] Current state snapshot failed: {e}")


def log_errors(hall: str, rack_type: str, building: str, rack: str, error_category: str, count: int,
               source_file: str = "", processed_by: str = "system"):
    ensure_log_exists()
    abs_path = _get_abs_log_path()
    print(f"[DATA_LOGGER] Logging: hall={hall}, rack_type={rack_type}, building={building}, rack={rack}, category={error_category}, count={count}")

    # Always backup the current log before appending new data
    backup_log()

    from openpyxl import load_workbook, Workbook

    max_retries = 5
    for attempt in range(max_retries):
        wb = None
        try:
            if LOG_FILE.exists():
                wb = load_workbook(LOG_FILE)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Error Log"
                headers = ["timestamp", "hall", "rack_type", "building", "rack",
                           "error_category", "count", "source_file", "processed_by"]
                ws.append(headers)

            ws.append([
                datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
                hall, rack_type, building, rack, error_category, count, source_file, processed_by
            ])
            wb.save(LOG_FILE)
            print(f"[DATA_LOGGER] SUCCESS - Logged to: {abs_path}")
            return True

        except PermissionError:
            if wb:
                try: wb.close()
                except: pass
            print(f"[DATA_LOGGER] Log file locked, retrying... ({attempt + 1}/{max_retries})")
            time.sleep(0.8 * (attempt + 1))
        except Exception as e:
            if wb:
                try: wb.close()
                except: pass
            print(f"[DATA_LOGGER] Failed: {e}")
            return False

    print("[DATA_LOGGER] Failed after multiple retries.")
    return False
