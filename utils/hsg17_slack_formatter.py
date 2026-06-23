#!/usr/bin/env python3
r"""
HSG17 T1-to-T0 Slack Formatter — core logic (GFAB CODE19 reference).

Imported by pages/02_HSG17_T1_to_T0_Slack.py. Formatting only — no tkinter/Streamlit.
Synced from GFAB_CODE19_T1_T0 reference script.

Supports legacy slack reports, integrated normalized exports, and T1 Optics /
T1 combined_fec tabs built from t1_t0_* source sheets.
"""

import sys
import os
import re
import shutil
from collections import Counter

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Tab-name resolution ──────────────────────────────────────────────────────
# The source-report generator periodically renames tabs (e.g. adding a
# "_with_pp" suffix). To survive these renames, look up tabs by ANY of their
# known aliases instead of hard-coding a single name.

TAB_ALIASES = {
    'lldp':         ('lldp_sp', 'full_path_lldp_with_int_down'),
    'optics':       ('optics_rx_tx_threshold', 'optics_rx_tx_threshold_with_pp'),
    'interfaces':   ('interfaces_sp', 'interfaces_sp_with_pp'),
    'combined_fec': ('combined_fec', 'combined_fec_with_pp'),
}

# Tabs that should always be dropped from the output (case-insensitive match).
# Add or remove names here as the source report evolves.
TABS_TO_REMOVE = (
    'device_reporting_failure',
    'bgp_sp',
    'spectrum_health',
    'sp_power',
    'sp_fans',
    'optics_temp',
    'pre_fec_ber_threshold_with_pp',
    'cert_sp',
    'cannot_decode_sp',
    'unknown_test_sp',
)

# Only these tabs survive in the final output. Anything else is dropped,
# regardless of TABS_TO_REMOVE. Case-insensitive match.
TABS_TO_KEEP = (
    'Summary',
    'Downlinks',
    'Mismatches',
    'Optics',
    'combined_fec',
    'T1 Optics',
    'T1 combined_fec',
)

# Output tabs built from the t1_t0_* source tabs. They mirror Optics /
# combined_fec but describe the T1-side of each T1↔T0 link. Kept in a set so
# the generic A-side enrichment loops can skip them (they get their own
# Z-side enrichment in step 3c).
T1_TABS = ('T1 Optics', 'T1 combined_fec')

# Columns to strip from every tab in the final output. Exact-match,
# case-sensitive. Edit this list to drop more columns globally.
COLUMNS_TO_REMOVE = (
    'Building',
    'Act. Building',
    'Exp. Building',
    # Expected-side columns (dropped per new-format request; harmless no-op
    # on old-format files that may not have all of these):
    'Expected Hostname',
    'Exp. Interface',
    'Exp. L/R',
    'Exp. Rack',
    'Exp. Elevation',
    'PP_A',
    'PP_Z',
    # combined_fec source noise:
    'Lock Status',
    'Pre-FEC BER',
    'Remote Host',
    'Remote Interface',
    'Mapped Remote Host',
    'Mapped Remote Interface',
    'Mapped Remote Rack',
    'Mapped Remote Elevation',
    'Remote Host Match',
    'Remote Interface Match',
    'Remote End Match',
    'Z_end_host',
    'Z_end_intf',
    'rack_z',
    'Z_Rack',
    'Z_Elevation',
    'Index',
    'Source Sheet',
    'Placement Group',
    'Cable Info',
    'A Transceiver SKU',
    'Z Transceiver SKU',
)

# Tabs that should receive Z-side info (Z Hostname, Z Interface,
# Z Rack, Z Elevation) pulled from the cutsheet by Hostname+Interface match.
# Columns are inserted right after Destination_port. Add more tab names
# here if you want Z-side info in additional tabs.
Z_FILL_TABS = ('Optics', 'combined_fec', 'Downlinks', 'Mismatches')

def formatted_output_path(input_path, output_path) -> str:
    """Stable output name: ``<input-stem>_FORMATTED.xlsx`` (filesystem-safe)."""
    stem = os.path.splitext(os.path.basename(input_path))[0].strip()
    stem = re.sub(r'[<>:"/\\|?*]', '_', stem).rstrip('. ')
    if stem.upper().endswith('_FORMATTED'):
        stem = stem[: -len('_FORMATTED')]
    return os.path.join(os.path.dirname(output_path) or '.', f"{stem}_FORMATTED.xlsx")


def finalize_output(input_path, output_path, log) -> str:
    """Rename saved workbook to the input-derived FORMATTED filename."""
    try:
        final = formatted_output_path(input_path, output_path)
        if os.path.abspath(final) != os.path.abspath(output_path):
            load_workbook(output_path).save(final)
            os.remove(output_path)
        log(f"  ✓ Saved → {os.path.basename(final)}")
        return final
    except Exception as exc:
        log(f"  ⚠ Could not rename output: {exc}")
        log(f"  ✓ Saved → {os.path.basename(output_path)}")
        return output_path


def find_tab(wb_or_sheetnames, key):
    """Return the actual tab name in the workbook for the given logical key,
    or None if no alias is present."""
    names = (wb_or_sheetnames.sheetnames
             if hasattr(wb_or_sheetnames, 'sheetnames')
             else list(wb_or_sheetnames))
    for alias in TAB_ALIASES[key]:
        if alias in names:
            return alias
    return None


# ── Style helpers ────────────────────────────────────────────────────────────

PINK   = 'FFB6C1'
YELLOW = 'FFFF00'

def thin_border():
    s = Side(style='thin', color='000000')
    return Border(left=s, right=s, top=s, bottom=s)


def clear_and_border(ws, pink_cols=None):
    """Remove all fills (preserve pink cols); yellow-highlight row 1; apply black border everywhere."""
    bd        = thin_border()
    no_fill   = PatternFill(fill_type=None)
    pink_fill = PatternFill('solid', start_color=PINK)
    yellow_fill = PatternFill('solid', start_color=YELLOW)
    pink_cols = set(pink_cols or [])
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                             min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.row == 1:
                cell.fill = yellow_fill
            elif cell.column in pink_cols:
                cell.fill = pink_fill
            else:
                cell.fill = no_fill
            cell.border = bd
            if cell.font:
                cell.font = Font(
                    bold=cell.font.bold,
                    name=cell.font.name or 'Arial',
                    size=cell.font.size or 10,
                    color='FF000000'
                )


def header_cell(cell, value, fill=None):
    cell.value     = value
    cell.font      = Font(bold=True, name='Arial', size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
    cell.border    = thin_border()
    cell.fill      = fill if fill else PatternFill('solid', start_color=YELLOW)


def data_cell(cell, value, fill=None):
    cell.value  = value
    cell.border = thin_border()
    if fill:
        cell.fill = fill


def autofit_sheet(ws, header_row_height=24, data_row_height=20, max_col_width=80):
    """Expand columns to fit content and give rows a comfortable height.

    - Column width = longest cell content in that column (+ padding), capped
      at `max_col_width` so a single huge value can't blow up the layout.
    - Cells inside a merged range are ignored when measuring column width.
    - Row 1 gets `header_row_height`; remaining rows get `data_row_height`.
    """
    # Cells that are part of a merge — skip them when measuring widths
    merged = set()
    for mrange in ws.merged_cells.ranges:
        for r in range(mrange.min_row, mrange.max_row + 1):
            for c in range(mrange.min_col, mrange.max_col + 1):
                merged.add((r, c))

    col_max = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                             min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is None:
                continue
            if (cell.row, cell.column) in merged:
                continue
            # Handle multi-line values: width = longest line
            longest_line = max(
                (len(line) for line in str(cell.value).splitlines()),
                default=0
            )
            letter = get_column_letter(cell.column)
            if longest_line > col_max.get(letter, 0):
                col_max[letter] = longest_line

    for letter, length in col_max.items():
        ws.column_dimensions[letter].width = min(length + 4, max_col_width)

    # Row heights
    if ws.max_row >= 1:
        ws.row_dimensions[1].height = header_row_height
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = data_row_height


def write_sheet(wb, name, df):
    ws = wb.create_sheet(name)
    bd = thin_border()
    for c, col in enumerate(df.columns, 1):
        header_cell(ws.cell(row=1, column=c), col)
    for r, (_, row) in enumerate(df.iterrows(), 2):
        for c, col in enumerate(df.columns, 1):
            val = row[col]
            cell = ws.cell(row=r, column=c, value=None if pd.isna(val) else val)
            cell.border = bd
    for c, col in enumerate(df.columns, 1):
        mx = max([len(str(col))] + [len(str(v)) for v in df[col].dropna()])
        ws.column_dimensions[get_column_letter(c)].width = min(mx + 2, 40)
    ws.freeze_panes = 'A2'
    return ws


# ── Reference-file loaders ───────────────────────────────────────────────────

def _load_cutsheet_deviceab(path):
    """QFAB allconnections (Sheet1 / DeviceA+DeviceB) — same normalisation as before."""
    xl = pd.ExcelFile(path)
    df = pd.read_excel(path, sheet_name=xl.sheet_names[0])

    def _split_device(col):
        parts = df[col].str.rsplit(' ', n=1, expand=True)
        return parts[0].str.strip(), parts[1].str.strip()

    df['Hostname'],   df['Interface']   = _split_device('DeviceA')
    df['Z Hostname'], df['Z Interface'] = _split_device('DeviceB')

    def _lr(col):
        return df[col].apply(
            lambda v: str(v).strip()[-1] if pd.notna(v) and str(v).strip() else '')

    df['L/R']   = _lr('DeviceA Physical Port')
    df['Z L/R'] = _lr('DeviceB Physical Port')

    def _rack(col):
        return (df[col].str.extract(r'Rack\s+(\d+)')[0]
                .astype(float).fillna(0).astype(int))

    def _elev(col):
        return (df[col].str.extract(r'U(\d+)')[0]
                .astype(float).fillna(0).astype(int))

    df['Rack']        = _rack('RackA')
    df['Elevation']   = _elev('RackA')
    df['Z Rack']      = _rack('RackB')
    df['Z Elevation'] = _elev('RackB')

    return df


def load_cutsheet(path):
    """Load cutsheet — identical to GFAB reference when Installation Sheet exists."""
    try:
        return pd.read_excel(path, sheet_name='Installation Sheet')
    except ValueError:
        return _load_cutsheet_deviceab(path)


def build_cutsheet_lookup(cut_df):
    """Key: (Hostname, Interface) → row dict for Source_port etc.

    Adapts to schema changes: only carries forward the fill columns that
    actually exist in the cutsheet. DMARC1/DMARC2 were dropped in later
    cutsheet revisions; the lookup still works without them.
    """
    candidate_cols = ['Source_port', 'DMARC1', 'DMARC2', 'Destination_port']
    fill_cols = [c for c in candidate_cols if c in cut_df.columns]
    lookup = {}
    for _, row in cut_df.iterrows():
        key = (
            str(row['Hostname']).strip(),
            str(row['Interface']).strip(),
        )
        lookup[key] = {c: row[c] for c in fill_cols}
    # Stash which columns were actually available so step 6 can use the same set
    lookup['__fill_cols__'] = fill_cols
    return lookup


def build_z_lookup(cut_df):
    """Key: (Z Hostname, Z Interface) → full row. Z-side is 1:1 unique in
    the cutsheet (same as A-side), so the simpler 2-tuple key is reliable
    and matches the (Hostname, Interface) pattern used elsewhere."""
    lookup = {}
    for _, row in cut_df.iterrows():
        key = (
            str(row['Z Hostname']).strip(),
            str(row['Z Interface']).strip(),
        )
        lookup[key] = row
    return lookup


def remap_optics_interface_by_channel(df):
    """Remap optic-lane interface suffixes from the Channel value.

    Some optics exports use fine-grained lane labels (s0–s3). When s2/s3 are
    present, the s0/s1 labels are unreliable too, so EVERY interface suffix is
    rewritten purely from its Channel value to match the cutsheet (which only
    defines s0/s1, each covering four channels):
        Channel 1-4 → s0
        Channel 5-8 → s1
    This only activates when at least one s2 or s3 interface is present; files
    that only ever use s0/s1 are left untouched. Returns rows changed.
    """
    if 'Interface' not in df.columns or 'Channel' not in df.columns:
        return 0
    # Only activate when fine-grained labels are present.
    has_s2s3 = df['Interface'].astype(str).str.contains(r's[23]$', regex=True).any()
    if not has_s2s3:
        return 0
    changed = 0
    for i in df.index:
        iface = str(df.at[i, 'Interface'])
        m = re.search(r's([0-3])$', iface)
        if not m:
            continue
        try:
            ch = int(float(df.at[i, 'Channel']))
        except (ValueError, TypeError):
            continue
        if 1 <= ch <= 4:
            new_suffix = 's0'
        elif 5 <= ch <= 8:
            new_suffix = 's1'
        else:
            continue
        new_iface = iface[:m.start()] + new_suffix
        if new_iface != iface:
            df.at[i, 'Interface'] = new_iface
            changed += 1
    return changed


def paired_subport(iface):
    """Return the paired sub-port interface name.

    s0 ↔ s1 are a pair, s2 ↔ s3 are a pair.
    e.g. 'swp4s0' → 'swp4s1', 'swp15s3' → 'swp15s2'.
    Returns None if iface doesn't end in s0/s1/s2/s3.
    """
    pairs = {'s0': 's1', 's1': 's0', 's2': 's3', 's3': 's2'}
    for suffix, mate in pairs.items():
        if iface.endswith(suffix):
            return iface[:-len(suffix)] + mate
    return None


def enforce_block_order(ws, block, anchor_after='Elevation'):
    """Force `block` columns to appear contiguously, in the given order,
    immediately after the `anchor_after` column.

    Handles the new input format where some of these columns (e.g.
    Source_port / Destination_port) already exist out of order: those
    pre-existing columns are repositioned rather than left in place.
    Values, cell styles and column widths are preserved. Columns not
    present are skipped, so this is a no-op on the old format where the
    block was already inserted in order.
    """
    from copy import copy as _copy
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    present = [c for c in block if c in header]
    if len(present) < 2 or anchor_after not in header:
        return
    # Already correct? (contiguous, in order, right after anchor) → skip.
    a = header.index(anchor_after)
    if header[a + 1:a + 1 + len(present)] == present:
        return

    # Snapshot every column: per-cell (value, style) + width.
    cols = []
    for c in range(1, ws.max_column + 1):
        cells = [(ws.cell(row=r, column=c).value, _copy(ws.cell(row=r, column=c)._style))
                 for r in range(1, ws.max_row + 1)]
        cols.append({'name': ws.cell(row=1, column=c).value,
                     'cells': cells,
                     'width': ws.column_dimensions[get_column_letter(c)].width})

    rest       = [d for d in cols if d['name'] not in present]
    block_cols = [next(d for d in cols if d['name'] == b)
                  for b in block if b in present]
    out, inserted = [], False
    for d in rest:
        out.append(d)
        if d['name'] == anchor_after:
            out.extend(block_cols)
            inserted = True
    if not inserted:
        out.extend(block_cols)

    for ci, d in enumerate(out, start=1):
        for ri, (val, style) in enumerate(d['cells'], start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell._style = style
        if d['width']:
            ws.column_dimensions[get_column_letter(ci)].width = d['width']


# ── Integrated ("long") format support ──────────────────────────────────────
# A newer source report exports an already-enriched, normalized layout where
# every issue tab shares one schema:
#   Category, Metric, Direction, T0 Host, T0 Interface, T1 Host, T1 Interface,
#   Rack, Z Rack, Sources, Source Directions, Values, Source_port,
#   Destination_port, Details
# There's no lldp_sp / combined_fec / Measured (dBm) here, and Source_port /
# Destination_port are pre-populated — so the cutsheet-matching pipeline used
# for the old format has nothing to bind to. These files get their own handler.
#
# Tab mapping requested:
#   interfaces_sp                              -> Downlinks
#   optics_rx_tx_threshold + optics_temp       -> Optics
#   fec_bin_issues + pre_fec_ber_threshold     -> Fec Errors
#   all_issues, source_audit                   -> dropped

INTEGRATED_OUTPUT = [
    ('Downlinks',  ['interfaces_sp']),
    ('Optics',     ['optics_rx_tx_threshold', 'optics_temp']),
    ('Fec Errors', ['fec_bin_issues', 'pre_fec_ber_threshold']),
]

# Integrated columns to keep per output tab (before cutsheet enrichment is
# appended). Host / Interface are created by coalescing the T0/T1 ends.
INTEGRATED_KEEP = {
    'Downlinks':  ['Direction', 'Metric', 'Host', 'Interface'],
    'Optics':     ['Direction', 'Metric', 'Values', 'Host', 'Interface'],
    'Fec Errors': ['Direction', 'Metric', 'Host', 'Interface'],
}

# Cutsheet fields appended to every tab (oriented so the matched Host/Interface
# is the local end), in this order, right after Interface.
ENRICH_COLS = ['L/R', 'Rack', 'Elevation',
               'Source_port', 'DMARC1', 'DMARC2', 'Destination_port',
               'Z Hostname', 'Z Interface', 'Z L/R', 'Z Rack', 'Z Elevation']


def is_integrated_format(input_path):
    """True when the input is the new normalized/integrated export.

    Detection: no old-format lldp/combined_fec tab is present AND an issue tab
    carries the integrated schema (T0 Host + T1 Host columns).
    """
    try:
        names = pd.ExcelFile(input_path).sheet_names
    except Exception:
        return False
    if any(a in names for a in TAB_ALIASES['lldp']):
        return False
    if any(a in names for a in TAB_ALIASES['combined_fec']):
        return False
    for t in ('all_issues', 'optics_rx_tx_threshold', 'interfaces_sp',
              'fec_bin_issues', 'pre_fec_ber_threshold'):
        if t in names:
            try:
                hdr = pd.read_excel(input_path, sheet_name=t, nrows=0).columns.tolist()
            except Exception:
                continue
            if 'T0 Host' in hdr and 'T1 Host' in hdr:
                return True
    return False


def _canon_link_keys(ws):
    """Set of orientation-independent endpoint-pair keys for a worksheet."""
    hdr = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(hdr)}
    keys = set()
    if 'Host' not in idx or 'Interface' not in idx:
        return keys, idx
    for r in range(2, ws.max_row + 1):
        def g(name):
            if name not in idx:
                return ''
            v = ws.cell(row=r, column=idx[name]).value
            return '' if v is None else str(v).strip()
        keys.add(tuple(sorted([(g('Host'), g('Interface')),
                               (g('Z Hostname'), g('Z Interface'))])))
    return keys, idx


def grey_rows_matching(wb, target_tab, source_tab, log):
    """Light-grey the font of every row in target_tab whose link (orientation-
    independent endpoint pair) also appears in source_tab."""
    if target_tab not in wb.sheetnames or source_tab not in wb.sheetnames:
        return
    GREY = 'FFD3D3D3'
    src_keys, _ = _canon_link_keys(wb[source_tab])
    if not src_keys:
        log(f"  · Greyed 0 {target_tab} row(s) also present in {source_tab}")
        return
    ws = wb[target_tab]
    _, idx = _canon_link_keys(ws)
    if 'Host' not in idx:
        return
    cnt = 0
    for r in range(2, ws.max_row + 1):
        def g(name):
            if name not in idx:
                return ''
            v = ws.cell(row=r, column=idx[name]).value
            return '' if v is None else str(v).strip()
        key = tuple(sorted([(g('Host'), g('Interface')),
                            (g('Z Hostname'), g('Z Interface'))]))
        if key in src_keys:
            cnt += 1
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                f = cell.font
                cell.font = Font(bold=f.bold if f else False,
                                 name=(f.name if f else None) or 'Arial',
                                 size=(f.size if f else None) or 10,
                                 color=GREY)
    log(f"  · Greyed {cnt} {target_tab} row(s) also present in {source_tab}")


def _dedup_integrated(df, out_name, log):
    """Remove duplicate rows for Optics / Fec Errors using an orientation-
    independent link key, so a row enriched via the Z-side (flipped) collapses
    with the same physical link seen in the normal direction.

    The canonical link key is the *unordered* pair of endpoints
    {(Host, Interface), (Z Hostname, Z Interface)} — identical for a link and
    its reciprocal.

    Optics     : one row per canonical link, keeping the highest 'Values'.
    Fec Errors : one row per (Metric, canonical link), keep first — so the
                 same link reported as both fec_bin and pre_fec_ber is kept,
                 but a true reciprocal duplicate of the same metric is dropped.
    """
    if out_name not in ('Optics', 'Fec Errors') or df.empty:
        return df

    def col(name):
        return (df[name].astype(str).str.strip()
                if name in df.columns else pd.Series([''] * len(df), index=df.index))

    h, i  = col('Host'), col('Interface')
    zh, zi = col('Z Hostname'), col('Z Interface')
    canon = [tuple(sorted([(str(a), str(b)), (str(c), str(d))]))
             for a, b, c, d in zip(h, i, zh, zi)]

    n0 = len(df)
    df = df.copy()
    if out_name == 'Optics':
        def _num(v):
            try:
                return float(str(v).split(':')[-1].strip())
            except (ValueError, AttributeError):
                return float('-inf')
        df['__k'] = canon
        df['__v'] = df['Values'].map(_num) if 'Values' in df.columns else 0
        df['__o'] = range(len(df))
        df = (df.sort_values('__v', ascending=False)
                .drop_duplicates('__k', keep='first')
                .sort_values('__o')
                .drop(columns=['__k', '__v', '__o'])
                .reset_index(drop=True))
    else:  # Fec Errors
        metric = col('Metric')
        df['__k'] = [(m,) + c for m, c in zip(metric, canon)]
        df = df.drop_duplicates('__k', keep='first').drop(columns='__k').reset_index(drop=True)

    removed = n0 - len(df)
    if removed:
        log(f"    · Removed {removed} duplicate row(s) from {out_name}")
    return df


def build_oriented_lookup(cut_df):
    """(host, interface) → cutsheet fields oriented so the matched end is local.

    The cutsheet is one-directional (A-side = T0 via Hostname/Interface,
    Z-side = T1 via Z Hostname/Z Interface). To let a single Host/Interface
    match either end, every cutsheet row produces two keyed entries:

      • A-side key → fields as-is (local = A/T0).
      • Z-side key → fields flipped (local = Z/T1): local L/R/Rack/Elevation
        come from the Z columns, Source_port/Destination_port and DMARC1/DMARC2
        are swapped, and the Z* output columns describe the A-side far end.
    """
    def v(r, c):
        x = r[c] if c in r else None
        return None if (isinstance(x, float) and pd.isna(x)) else x

    L = {}
    for _, r in cut_df.iterrows():
        L[(str(r['Hostname']).strip(), str(r['Interface']).strip())] = {
            'L/R': v(r, 'L/R'), 'Rack': v(r, 'Rack'), 'Elevation': v(r, 'Elevation'),
            'Source_port': v(r, 'Source_port'), 'DMARC1': v(r, 'DMARC1'),
            'DMARC2': v(r, 'DMARC2'), 'Destination_port': v(r, 'Destination_port'),
            'Z Hostname': v(r, 'Z Hostname'), 'Z Interface': v(r, 'Z Interface'),
            'Z L/R': v(r, 'Z L/R'), 'Z Rack': v(r, 'Z Rack'), 'Z Elevation': v(r, 'Z Elevation'),
        }
        L[(str(r['Z Hostname']).strip(), str(r['Z Interface']).strip())] = {
            'L/R': v(r, 'Z L/R'), 'Rack': v(r, 'Z Rack'), 'Elevation': v(r, 'Z Elevation'),
            'Source_port': v(r, 'Destination_port'), 'DMARC1': v(r, 'DMARC2'),
            'DMARC2': v(r, 'DMARC1'), 'Destination_port': v(r, 'Source_port'),
            'Z Hostname': v(r, 'Hostname'), 'Z Interface': v(r, 'Interface'),
            'Z L/R': v(r, 'L/R'), 'Z Rack': v(r, 'Rack'), 'Z Elevation': v(r, 'Elevation'),
        }
    return L


def enrich_full_by_host(df, cut_df, log):
    """Append ENRICH_COLS to df, matching (Host, Interface) against the cutsheet
    (either end), with s0↔s1 / s2↔s3 paired-subport fallback."""
    if cut_df is None or 'Host' not in df.columns or 'Interface' not in df.columns:
        return df
    L = build_oriented_lookup(cut_df)
    df = df.copy()
    for c in ENRICH_COLS:
        df[c] = None
    hits = 0
    for i in df.index:
        h  = '' if df.at[i, 'Host'] is None else str(df.at[i, 'Host']).strip()
        it = '' if df.at[i, 'Interface'] is None else str(df.at[i, 'Interface']).strip()
        if not h or not it or h.lower() == 'nan' or it.lower() == 'nan':
            continue
        m = L.get((h, it))
        if m is None:
            mate = paired_subport(it)
            if mate:
                m = L.get((h, mate))
        if m is None:
            continue
        hits += 1
        for c in ENRICH_COLS:
            df.at[i, c] = m[c]
    log(f"    · Enriched {hits}/{len(df)} row(s) from cutsheet by Host/Interface")
    return df


def process_integrated_file(input_path, output_path, cut_df, log):
    """Build a formatted workbook from the new integrated-format input."""
    available = set(pd.ExcelFile(input_path).sheet_names)

    # Build a clean workbook — we emit only the mapped tabs + Summary.
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    # Capture rack counts per tab BEFORE columns are dropped, so the Summary
    # tab and the rack-based filename still work after 'Rack' is removed.
    rack_by_tab = {}

    for out_name, src_tabs in INTEGRATED_OUTPUT:
        frames = []
        used   = []
        for src in src_tabs:
            if src not in available:
                continue
            # Rack / Z Rack are zero-padded 4-digit text (e.g. '0810'); read
            # them as strings so pandas doesn't coerce to int and drop padding.
            df = pd.read_excel(input_path, sheet_name=src,
                               dtype={'Rack': str, 'Z Rack': str})
            if not df.empty:
                frames.append(df)
            used.append(src)
        if not frames:
            if used:
                log(f"  · {out_name}: source tab(s) {', '.join(used)} present but empty — skipped")
            else:
                log(f"  · {out_name}: no source tab found — skipped")
            continue
        combined = pd.concat(frames, ignore_index=True, sort=False)

        # Per-row: keep only the source-end host/interface for the Direction.
        #   T0 -> T1  → blank T1 Host / T1 Interface
        #   T1 -> T0  → blank T0 Host / T0 Interface
        if 'Direction' in combined.columns:
            for i in combined.index:
                d = str(combined.at[i, 'Direction']).strip()
                if d == 'T0 -> T1':
                    blanks = ('T1 Host', 'T1 Interface')
                elif d == 'T1 -> T0':
                    blanks = ('T0 Host', 'T0 Interface')
                else:
                    continue
                for c in blanks:
                    if c in combined.columns:
                        combined.at[i, c] = None

        # Collapse T0/T1 Host & Interface into single Host / Interface columns.
        host_cols = ['T0 Host', 'T1 Host']
        int_cols  = ['T0 Interface', 'T1 Interface']
        present   = [c for c in host_cols + int_cols if c in combined.columns]
        if present:
            def _coalesce(row, cols):
                for c in cols:
                    if c in combined.columns:
                        v = row.get(c)
                        if v is not None and not (isinstance(v, float) and pd.isna(v)) \
                                and str(v).strip() != '':
                            return v
                return None
            host_s = combined.apply(lambda r: _coalesce(r, host_cols), axis=1)
            int_s  = combined.apply(lambda r: _coalesce(r, int_cols), axis=1)
            anchor = min(list(combined.columns).index(c) for c in present)
            combined = combined.drop(columns=present)
            anchor = min(anchor, len(combined.columns))
            combined.insert(anchor, 'Host', host_s.values)
            combined.insert(anchor + 1, 'Interface', int_s.values)

        # Snapshot rack values (Rack + Z Rack) from the integrated columns
        # BEFORE trimming, so the Summary tab + filename still work.
        racks = {'Rack': [], 'Z Rack': []}
        for col in ('Rack', 'Z Rack'):
            if col in combined.columns:
                racks[col] = [str(v).strip() for v in combined[col].dropna()
                              if str(v).strip() != '']
        rack_by_tab[out_name] = racks

        # Keep only the wanted integrated columns, then append the full
        # cutsheet enrichment (matched by Host/Interface).
        keep = [c for c in INTEGRATED_KEEP.get(out_name, []) if c in combined.columns]
        combined = combined[keep]
        combined = enrich_full_by_host(combined, cut_df, log)
        combined = _dedup_integrated(combined, out_name, log)

        # Sort by Direction ascending (stable), so T0 -> T1 rows precede T1 -> T0.
        if 'Direction' in combined.columns:
            combined = (combined.sort_values('Direction', kind='stable')
                                .reset_index(drop=True))

        log(f"  · {out_name} ← {' + '.join(used)} ({len(combined)} row(s))")
        write_sheet(wb, out_name, combined)

    if not wb.sheetnames:
        raise RuntimeError("No mapped source tabs found in integrated input.")

    # ── Standard formatting (mirrors the old-format path) ───────────────────
    log("  · Removing fills and applying borders")
    for sheet_name in wb.sheetnames:
        clear_and_border(wb[sheet_name])

    log("  · Aligning all cells to middle-centre")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=False)
    for sheet_name in wb.sheetnames:
        for row in wb[sheet_name].iter_rows():
            for cell in row:
                cell.alignment = center_align

    log("  · Adding NOTE column and filters to all tabs")
    no_fill     = PatternFill(fill_type=None)
    yellow_fill = PatternFill('solid', start_color=YELLOW)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        col_idx = ws.max_column + 1
        hdr = ws.cell(row=1, column=col_idx, value='NOTE')
        hdr.font      = Font(bold=True, name='Arial', size=10)
        hdr.alignment = center_align
        hdr.fill      = yellow_fill
        hdr.border    = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = 16
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            cell.fill   = no_fill
            cell.border = thin_border()
        if ws.max_row > 1 and ws.max_column > 0:
            ws.auto_filter.ref = ws.dimensions

    # Grey-font Optics rows present in Downlinks, then Fec Errors rows present
    # in Optics (orientation-independent link match).
    grey_rows_matching(wb, 'Optics', 'Downlinks', log)
    grey_rows_matching(wb, 'Fec Errors', 'Optics', log)

    log("  · Expanding all columns and rows")
    for sheet_name in wb.sheetnames:
        autofit_sheet(wb[sheet_name])

    # ── Summary tab (per Rack breakdown; Rack kept as text e.g. '0810') ─────
    log("  · Creating Summary tab")
    center_s      = center_align
    bd_s          = thin_border()
    no_fill_s     = PatternFill(fill_type=None)
    yellow_fill_s = PatternFill('solid', start_color=YELLOW)

    def _s(cell, value, bold=False, header=False):
        cell.value     = value
        cell.font      = Font(bold=bold, name='Arial', size=10)
        cell.alignment = center_s
        cell.border    = bd_s
        cell.fill      = yellow_fill_s if header else no_fill_s

    tab_rack  = {}
    all_racks = set()
    for sname in wb.sheetnames:
        counts = {}
        for k in rack_by_tab.get(sname, {}).get('Rack', []):
            counts[k] = counts.get(k, 0) + 1
            all_racks.add(k)
        tab_rack[sname] = counts

    racks      = sorted(all_racks)
    tabs_order = list(wb.sheetnames)
    total_cols = 1 + len(racks) + 1

    wb.create_sheet('Summary', 0)
    ws_s = wb['Summary']
    ws_s.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(total_cols, 1))
    title_c = ws_s.cell(row=1, column=1, value='Tab Summary by Rack')
    title_c.font = Font(bold=True, name='Arial', size=13)
    title_c.alignment = center_s
    title_c.border    = bd_s
    title_c.fill      = yellow_fill_s
    ws_s.row_dimensions[1].height = 28

    _s(ws_s.cell(row=2, column=1), 'Tab Name', bold=True, header=True)
    for c, rack in enumerate(racks, start=2):
        _s(ws_s.cell(row=2, column=c), str(rack), bold=True, header=True)
    _s(ws_s.cell(row=2, column=total_cols), 'Total', bold=True, header=True)

    rack_totals = {r: 0 for r in racks}
    data_tabs   = [n for n in tabs_order if n.lower() != 'summary']
    for i, tab_name in enumerate(data_tabs, start=3):
        _s(ws_s.cell(row=i, column=1), tab_name)
        row_total = 0
        for c, rack in enumerate(racks, start=2):
            count = tab_rack.get(tab_name, {}).get(rack, 0)
            _s(ws_s.cell(row=i, column=c), count if count > 0 else '')
            rack_totals[rack] += count
            row_total += count
        _s(ws_s.cell(row=i, column=total_cols), row_total, bold=True)

    tot_r = 3 + len(data_tabs)
    _s(ws_s.cell(row=tot_r, column=1), 'TOTAL', bold=True)
    grand = 0
    for c, rack in enumerate(racks, start=2):
        _s(ws_s.cell(row=tot_r, column=c), rack_totals[rack], bold=True)
        grand += rack_totals[rack]
    _s(ws_s.cell(row=tot_r, column=total_cols), grand, bold=True)

    ws_s.column_dimensions['A'].width = 20
    for c in range(2, total_cols + 1):
        ws_s.column_dimensions[get_column_letter(c)].width = 14

    # Order: Summary, Downlinks, Optics, Fec Errors
    desired  = ['Summary', 'Downlinks', 'Optics', 'Fec Errors']
    ordered  = [n for n in desired if n in wb.sheetnames]
    leftover = [n for n in wb.sheetnames if n not in desired]
    wb._sheets = [wb[n] for n in ordered + leftover]

    wb.save(output_path)
    return finalize_output(input_path, output_path, log)


# ── Core processor ───────────────────────────────────────────────────────────

def process_file(input_path, output_path, cut_df, log):
    # Route the new normalized/integrated export to its own handler; the
    # old-format pipeline below is left exactly as-is.
    if is_integrated_format(input_path):
        log("  · Detected integrated format — using integrated handler")
        shutil.copy2(input_path, output_path)
        return process_integrated_file(input_path, output_path, cut_df, log)

    shutil.copy2(input_path, output_path)
    wb = load_workbook(output_path)

    cutsheet_lookup = build_cutsheet_lookup(cut_df)
    z_lookup        = build_z_lookup(cut_df)

    # ── 1. Split lldp tab → Downlinks / Mismatches ──────────────────────────
    mis_orig_df = None   # keep for step 6b
    lldp_tab = find_tab(wb, 'lldp')
    if lldp_tab:
        log(f"  · Splitting {lldp_tab} → Downlinks / Mismatches")
        df = pd.read_excel(input_path, sheet_name=lldp_tab)
        down_df     = df[df['Act. Interface'] == 'interface down'].copy()
        mis_orig_df = df[df['Act. Interface'].str.startswith('swp', na=False)].copy()
        # Downlinks: drop only the Active-side columns; keep the Expected-side
        # columns and order them right after Elevation. Source_port and
        # Destination_port will be inserted between Elevation and Expected*
        # by step 6, giving the final order:
        #   Hostname, Interface, Rack, Elevation,
        #   Source_port, Destination_port,
        #   Expected Hostname, Exp. Interface, Exp. Rack, Exp. Elevation
        drop = ['Active Host', 'Act. Interface', 'Act. Rack', 'Act. Elevation']
        down_df.drop(columns=[c for c in drop if c in down_df.columns], inplace=True)
        # Pull Expected* to the end (relative order preserved) so they sit
        # after Elevation in the written sheet.
        exp_cols = ['Expected Hostname', 'Exp. Interface', 'Exp. Rack', 'Exp. Elevation']
        present_exp = [c for c in exp_cols if c in down_df.columns]
        if present_exp:
            rest = [c for c in down_df.columns if c not in present_exp]
            down_df = down_df[rest + present_exp]
        del wb[lldp_tab]
        write_sheet(wb, 'Downlinks', down_df)
        write_sheet(wb, 'Mismatches', mis_orig_df.drop(
            columns=[c for c in [] if c in mis_orig_df.columns]))  # keep all cols for now

    # ── 2. Optics tab ───────────────────────────────────────────────────────
    optics_src = find_tab(wb, 'optics')
    if optics_src:
        log(f"  · Processing Optics tab ({optics_src})")
        # Extra cols introduced by the *_with_pp variant — drop them too.
        drop_cols = {'Transceiver', 'Channel',
                     'Min Threshold (dBm)', 'Max Threshold (dBm)',
                     'PP_A', 'PP_Z', 'Z_end_host', 'Z_end_intf',
                     'rack_z', 'Z_Rack', 'Z_Elevation', 'Index',
                     'Status', 'Placement Group',
                     'Breakout Source', 'Breakout Mode', 'Breakout Warning'}
        optics_df = pd.read_excel(input_path, sheet_name=optics_src)
        # New-format inputs label optic lanes s0–s3, but the cutsheet only
        # defines s0/s1. When s2/s3 are present the s0/s1 labels are also
        # unreliable, so remap EVERY interface by Channel (1-4→s0, 5-8→s1) to
        # match the cutsheet. Must run before 'Channel' is dropped.
        n_remap = remap_optics_interface_by_channel(optics_df)
        if n_remap:
            log(f"    · Remapped {n_remap} optic interface(s) by Channel "
                f"(1-4→s0, 5-8→s1)")
        optics_df.drop(columns=[c for c in drop_cols if c in optics_df.columns], inplace=True)
        # Put 'Metric' first, then 'Measured (dBm)', so both columns can be
        # frozen and stay visible while scrolling horizontally.
        leading = [c for c in ('Metric', 'Measured (dBm)') if c in optics_df.columns]
        if leading:
            rest = [c for c in optics_df.columns if c not in leading]
            optics_df = optics_df[leading + rest]
        del wb[optics_src]
        write_sheet(wb, 'Optics', optics_df)
        # Freeze row 1 + the two leading columns (Metric, Measured (dBm)).
        wb['Optics'].freeze_panes = 'C2' if len(leading) >= 2 else 'B2'

    # ── 3. Remove interfaces tab ────────────────────────────────────────────
    interfaces_tab = find_tab(wb, 'interfaces')
    if interfaces_tab:
        log(f"  · Removing {interfaces_tab}")
        del wb[interfaces_tab]

    # ── 3a. Remove unwanted source tabs ─────────────────────────────────────
    drop_lower = {t.lower() for t in TABS_TO_REMOVE}
    for existing in list(wb.sheetnames):
        if existing.lower() in drop_lower:
            log(f"  · Removing {existing}")
            del wb[existing]

    # ── 3b. combined_fec: move Lock Status + Pre-FEC BER to the front ───────
    fec_tab = find_tab(wb, 'combined_fec')
    if fec_tab:
        log(f"  · Reordering {fec_tab} (Lock Status, Pre-FEC BER first)")
        fec_df = pd.read_excel(input_path, sheet_name=fec_tab)

        def _norm(s):
            # Normalize hyphen variants and whitespace for tolerant matching
            return (str(s)
                    .replace('\u2011', '-')   # non-breaking hyphen
                    .replace('\u2013', '-')   # en dash
                    .replace('\u2014', '-')   # em dash
                    .strip()
                    .lower())

        wanted = ['lock status', 'pre-fec ber']
        front = []
        for target in wanted:
            for col in fec_df.columns:
                if _norm(col) == target and col not in front:
                    front.append(col)
                    break

        if front:
            rest = [c for c in fec_df.columns if c not in front]
            fec_df = fec_df[front + rest]
            del wb[fec_tab]
            write_sheet(wb, 'combined_fec', fec_df)
        else:
            log("    ⚠ Lock Status / Pre-FEC BER not found — leaving combined_fec as-is")

    # ── 3c. Build T1 Optics / T1 combined_fec from the t1_t0_* source tabs ───
    # These mirror the standard Optics / combined_fec tabs but describe the
    # T1-side of each T1↔T0 link. T1 devices live on the *Z-side* of the
    # cutsheet, so enrichment matches each row's (Hostname, Interface) against
    # the cutsheet Z-side and flips orientation: the local (T1) end becomes
    # Source_port / L/R and the remote (T0) end becomes the Z block — yielding
    # exactly the same column layout as the standard tabs.
    src_names = pd.ExcelFile(input_path).sheet_names
    has_dmarc = 'DMARC1' in cut_df.columns and 'DMARC2' in cut_df.columns

    # Full-row cutsheet lookups for BOTH orientations. The standard tabs only
    # ever match the A-side; T1 devices normally live on the Z-side, but we
    # match either side so enrichment never depends on which way a link was
    # catalogued. z_lookup (A-side keyed by Z Hostname/Interface) is built at
    # the top of process_file; build the A-side keyed lookup here to match it.
    a_lookup = {}
    for _, _row in cut_df.iterrows():
        a_lookup[(str(_row['Hostname']).strip(),
                  str(_row['Interface']).strip())] = _row

    def _link_enrich_row(host, iface):
        """Orientation-independent enrichment, always presenting the LOCAL
        (T1) end as the source.

        Match (host, iface) against the cutsheet on the Z-side first, then the
        A-side, with a paired sub-port (s0↔s1, s2↔s3) fallback on each. When
        the match lands on the Z-side, flip the row so the local end's patch
        becomes Source_port and the far (A-side) end becomes the Z block. When
        it lands on the A-side, the local end is already the source, so the
        cutsheet fields are used as-is. Either way the local end reads as the
        source — i.e. the layout is independent of how the link was stored.
        """
        flip = True
        m = z_lookup.get((host, iface))
        if m is None:
            mate = paired_subport(iface)
            if mate:
                m = z_lookup.get((host, mate))
        if m is None:                       # not on the Z-side — try A-side
            flip = False
            m = a_lookup.get((host, iface))
            if m is None:
                mate = paired_subport(iface)
                if mate:
                    m = a_lookup.get((host, mate))

        out = {k: None for k in
               ['L/R', 'Rack', 'Elevation',
                'Source_port', 'DMARC1', 'DMARC2', 'Destination_port',
                'Z Hostname', 'Z Interface', 'Z L/R', 'Z Rack', 'Z Elevation']}
        if m is None:
            return out

        def g(col):
            v = m.get(col)
            return None if (v is None or (isinstance(v, float) and pd.isna(v))) else v

        if flip:
            # Local matched the cutsheet Z-side: flip so local end = source,
            # far (cutsheet A-side) end = Z block.
            out['L/R']              = g('Z L/R')
            out['Rack']             = g('Z Rack')
            out['Elevation']        = g('Z Elevation')
            out['Source_port']      = g('Destination_port')
            out['Destination_port'] = g('Source_port')
            if has_dmarc:
                out['DMARC1'] = g('DMARC2')
                out['DMARC2'] = g('DMARC1')
            out['Z Hostname']  = g('Hostname')
            out['Z Interface'] = g('Interface')
            out['Z L/R']       = g('L/R')
            out['Z Rack']      = g('Rack')
            out['Z Elevation'] = g('Elevation')
        else:
            # Local matched the cutsheet A-side: already the source end.
            out['L/R']              = g('L/R')
            out['Rack']             = g('Rack')
            out['Elevation']        = g('Elevation')
            out['Source_port']      = g('Source_port')
            out['Destination_port'] = g('Destination_port')
            if has_dmarc:
                out['DMARC1'] = g('DMARC1')
                out['DMARC2'] = g('DMARC2')
            out['Z Hostname']  = g('Z Hostname')
            out['Z Interface'] = g('Z Interface')
            out['Z L/R']       = g('Z L/R')
            out['Z Rack']      = g('Z Rack')
            out['Z Elevation'] = g('Z Elevation')
        return out

    def _build_t1(src_tab, out_name, leading):
        if src_tab not in src_names:
            return
        sdf = pd.read_excel(input_path, sheet_name=src_tab)
        if not all(c in sdf.columns for c in ('Hostname', 'Interface')):
            return
        log(f"  · Building {out_name} from {src_tab}")
        recs = []
        for _, srow in sdf.iterrows():
            host  = str(srow.get('Hostname', '') or '').strip()
            iface = str(srow.get('Interface', '') or '').strip()
            enr = _link_enrich_row(host, iface)
            # Local Rack/Elevation come from the cutsheet (oriented to the
            # local end); the t1_t0 report's own Rack/Elevation describe the
            # far T0 end, so only fall back to them when the cutsheet misses.
            rec = {
                'Hostname':  srow.get('Hostname'),
                'Interface': srow.get('Interface'),
                'Rack':      enr['Rack'] if enr['Rack'] is not None else srow.get('Rack'),
                'Elevation': enr['Elevation'] if enr['Elevation'] is not None else srow.get('Elevation'),
            }
            for c in leading:
                rec[c] = srow.get(c)
            enr.pop('Rack', None)
            enr.pop('Elevation', None)
            rec.update(enr)
            recs.append(rec)

        port_cols = (['Source_port']
                     + (['DMARC1', 'DMARC2'] if has_dmarc else [])
                     + ['Destination_port'])
        col_order = (list(leading)
                     + ['Hostname', 'Interface', 'L/R', 'Rack', 'Elevation']
                     + port_cols
                     + ['Z Hostname', 'Z Interface', 'Z L/R', 'Z Rack', 'Z Elevation'])
        out_df = pd.DataFrame(recs)
        if 'Measured (dBm)' in out_df.columns:
            out_df['Measured (dBm)'] = pd.to_numeric(out_df['Measured (dBm)'],
                                                     errors='coerce')
        out_df = out_df.reindex(columns=col_order)
        write_sheet(wb, out_name, out_df)
        # Freeze row 1 + the leading columns (Metric, Measured (dBm)) like Optics.
        if leading:
            wb[out_name].freeze_panes = 'C2' if len(leading) >= 2 else 'B2'

    _build_t1('t1_t0_optics_rx_tx', 'T1 Optics', ['Metric', 'Measured (dBm)'])
    _build_t1('t1_t0_combined_fec', 'T1 combined_fec', [])

    # Drop the raw t1_t0_* source tabs (consumed above / unused link_ports).
    for t in ('t1_t0_optics_rx_tx', 't1_t0_combined_fec', 't1_t0_link_ports'):
        if t in wb.sheetnames:
            del wb[t]

    # ── 4. Reorder tabs ─────────────────────────────────────────────────────
    # Put the four primary output tabs at the END in the documented order,
    # everything else stays in its current relative order. Rewriting _sheets
    # directly is simpler and more robust than chained move_sheet() calls
    # whose offset arithmetic depended on the count of intermediate tabs.
    desired  = ['Downlinks', 'Mismatches', 'Optics', 'combined_fec',
                'T1 Optics', 'T1 combined_fec']
    existing = [s for s in desired if s in wb.sheetnames]
    others   = [s for s in wb.sheetnames if s not in desired]
    wb._sheets = [wb[n] for n in others + existing]

    # ── 5. Insert L/R columns (sourced from cutsheet) ───────────────────────
    log("  · Adding L/R mapped columns (from cutsheet)")
    # Build interface→L/R map from both A-side and Z-side cutsheet columns.
    lr_from_cut = {}
    for _, row in cut_df.iterrows():
        iface = str(row.get('Interface', '') or '').strip()
        lr    = str(row.get('L/R', '')    or '').strip()
        if iface and lr:
            lr_from_cut[iface] = lr
        z_iface = str(row.get('Z Interface', '') or '').strip()
        z_lr    = str(row.get('Z L/R', '')      or '').strip()
        if z_iface and z_lr:
            lr_from_cut[z_iface] = z_lr

    lr_name_for = {
        'Interface':      'L/R',
        'Z Interface':    'Z L/R',
        'Exp. Interface': 'Exp. L/R',
    }
    for sheet_name in wb.sheetnames:
        if sheet_name in T1_TABS:
            continue          # already enriched with L/R in step 3c
        ws = wb[sheet_name]
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        targets = [(i+1, h) for i, h in enumerate(header) if h in lr_name_for]
        for col_idx, col_name in sorted(targets, reverse=True):
            new_name = lr_name_for[col_name]
            ws.insert_cols(col_idx + 1)
            header_cell(ws.cell(row=1, column=col_idx + 1), new_name)
            for r in range(2, ws.max_row + 1):
                val = str(ws.cell(row=r, column=col_idx).value or '').strip()
                ws.cell(row=r, column=col_idx + 1, value=lr_from_cut.get(val, ''))
                ws.cell(row=r, column=col_idx + 1).border = thin_border()
            ws.column_dimensions[get_column_letter(col_idx + 1)].width = 10

    # ── 5b. Populate Source_port / DMARC1 / DMARC2 / Destination_port ────────
    # Match each row's (Hostname, Interface) against the cutsheet and fill
    # the available cutsheet columns. New fill columns get inserted right
    # after Elevation when the target sheet doesn't already have them.
    fill_cols = cutsheet_lookup.get('__fill_cols__',
                                    ['Source_port', 'DMARC1', 'DMARC2', 'Destination_port'])
    log(f"  · Filling {', '.join(fill_cols) or '(no cutsheet fill cols available)'} (match on Hostname + Interface)")
    for sheet_name in wb.sheetnames:
        if sheet_name in T1_TABS:
            continue          # already enriched (Z-side) in step 3c
        ws = wb[sheet_name]
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        if not all(c in header for c in ['Hostname', 'Interface']):
            continue
        # Append any fill columns this sheet doesn't already have, right after
        # Elevation if present, otherwise at the end.
        anchor = (header.index('Elevation') + 1) if 'Elevation' in header else len(header)
        insert_at = anchor + 1
        for col_name in fill_cols:
            if col_name in header:
                continue
            ws.insert_cols(insert_at)
            header_cell(ws.cell(row=1, column=insert_at), col_name)
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=insert_at).border = thin_border()
            ws.column_dimensions[get_column_letter(insert_at)].width = max(len(col_name)+2, 14)
            insert_at += 1
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        host_c, int_c = header.index('Hostname')+1, header.index('Interface')+1
        fill_idx = {c: header.index(c)+1 for c in fill_cols}
        for r in range(2, ws.max_row + 1):
            host  = str(ws.cell(row=r, column=host_c).value or '').strip()
            iface = str(ws.cell(row=r, column=int_c).value or '').strip()
            match = cutsheet_lookup.get((host, iface))
            if match:
                for col_name, col_idx in fill_idx.items():
                    val = match.get(col_name)
                    if val is not None and not (isinstance(val, float) and pd.isna(val)):
                        ws.cell(row=r, column=col_idx, value=val)

        # Force canonical contiguous order right after Elevation, repositioning
        # any of these that already existed in the input (new-format files).
        enforce_block_order(ws, ['Source_port', 'DMARC1', 'DMARC2', 'Destination_port'])

    # ── 6c. Fill Z-side info in designated tabs (default: Optics) ───────────
    Z_COLS = ['Z Hostname', 'Z Interface', 'Z L/R', 'Z Rack', 'Z Elevation']
    z_available = [c for c in Z_COLS if c in cut_df.columns]
    # Build unconditionally — also used by step 6b Mismatches act_lookup for Z L/R
    z_by_host_int = {}
    for _, row in cut_df.iterrows():
        k = (str(row['Hostname']).strip(), str(row['Interface']).strip())
        z_by_host_int[k] = {c: row[c] for c in z_available}

    if z_available and Z_FILL_TABS:
        for tab in Z_FILL_TABS:
            if tab not in wb.sheetnames:
                continue
            ws_z = wb[tab]
            header = [ws_z.cell(row=1, column=c).value
                      for c in range(1, ws_z.max_column + 1)]
            if not all(c in header for c in ['Hostname', 'Interface']):
                continue
            log(f"  · Filling Z-side info in {tab}: {', '.join(z_available)}")
            # Anchor: right after Destination_port if present, else after
            # Elevation, else at the end of the sheet.
            if 'Destination_port' in header:
                anchor = header.index('Destination_port') + 1
            elif 'Elevation' in header:
                anchor = header.index('Elevation') + 1
            else:
                anchor = len(header)
            insert_at = anchor + 1
            for col_name in z_available:
                if col_name in header:
                    continue
                ws_z.insert_cols(insert_at)
                header_cell(ws_z.cell(row=1, column=insert_at), col_name)
                for r in range(2, ws_z.max_row + 1):
                    ws_z.cell(row=r, column=insert_at).border = thin_border()
                ws_z.column_dimensions[get_column_letter(insert_at)].width = max(len(col_name)+2, 14)
                insert_at += 1
            header = [ws_z.cell(row=1, column=c).value
                      for c in range(1, ws_z.max_column + 1)]
            host_c, int_c = header.index('Hostname')+1, header.index('Interface')+1
            fill_idx = {c: header.index(c)+1 for c in z_available}
            for r in range(2, ws_z.max_row + 1):
                host  = str(ws_z.cell(row=r, column=host_c).value or '').strip()
                iface = str(ws_z.cell(row=r, column=int_c).value or '').strip()
                match = z_by_host_int.get((host, iface))
                if match:
                    for col_name, col_idx in fill_idx.items():
                        val = match.get(col_name)
                        if val is not None and not (isinstance(val, float) and pd.isna(val)):
                            ws_z.cell(row=r, column=col_idx, value=val)

    # ── 6b. Mismatches: Possible columns + Active Z columns (pink) ───────────
    if 'Mismatches' in wb.sheetnames:
        log("  · Building Possible + Active Z columns in Mismatches")
        pink_fill   = PatternFill('solid', start_color=PINK)
        yellow_fill = PatternFill('solid', start_color=YELLOW)
        bd          = thin_border()

        # Build act_lookup from original lldp tab (alias-aware)
        # z_by_host_int built earlier for Z-fill is reused here to get Z L/R from cutsheet
        act_lookup = {}
        src_sheets = pd.ExcelFile(input_path).sheet_names
        src_lldp = find_tab(src_sheets, 'lldp')
        if src_lldp:
            orig_df  = pd.read_excel(input_path, sheet_name=src_lldp)
            mis_rows = orig_df[orig_df['Act. Interface'].str.startswith('swp', na=False)]
            for _, row in mis_rows.iterrows():
                key       = (str(row['Hostname']).strip(), str(row['Interface']).strip())
                act_host  = str(row['Active Host']).strip()
                act_iface = str(row['Act. Interface']).strip()
                # Pull Z L/R from cutsheet: Active Host/Interface is the Z-side,
                # so look it up in z_lookup (keyed by Z Hostname, Z Interface).
                # Fallback to paired sub-port (s0↔s1, s2↔s3) if exact miss.
                cut_z_row = z_lookup.get((act_host, act_iface))
                if cut_z_row is None:
                    mate = paired_subport(act_iface)
                    if mate:
                        cut_z_row = z_lookup.get((act_host, mate))
                z_lr_val = ''
                if cut_z_row is not None:
                    raw = cut_z_row.get('Z L/R')
                    if raw is not None and not (isinstance(raw, float) and pd.isna(raw)):
                        z_lr_val = str(raw).strip()
                act_lookup[key] = {
                    'Z Hostname' : act_host,
                    'Z Interface': act_iface,
                    'Z L/R'      : z_lr_val,
                    'Z Rack'     : int(float(str(row['Act. Rack']))),
                    'Z Elevation': int(float(str(row['Act. Elevation']))),
                }

        ws_m   = wb['Mismatches']
        header = [ws_m.cell(row=1, column=c).value for c in range(1, ws_m.max_column + 1)]

        # Drop only the Active-side columns. Expected-side columns are kept
        # so they can sit between Destination_port and the Possible/Z blocks
        # in the final layout.
        act_drop = {'Active Host', 'Act. Interface', 'Act. Rack', 'Act. Elevation'}
        for idx in sorted([i+1 for i, h in enumerate(header) if h in act_drop], reverse=True):
            ws_m.delete_cols(idx)
        header = [ws_m.cell(row=1, column=c).value for c in range(1, ws_m.max_column + 1)]

        h_idx = header.index('Hostname') + 1
        i_idx = header.index('Interface') + 1

        # Collect act data per row
        act_rows = []
        for r in range(2, ws_m.max_row + 1):
            hn    = str(ws_m.cell(row=r, column=h_idx).value or '').strip()
            iface = str(ws_m.cell(row=r, column=i_idx).value or '').strip()
            act_rows.append(act_lookup.get((hn, iface), {}))

        # Possible columns: match act Z key against cutsheet Z side.
        # Filter out ones whose source column doesn't exist in this cutsheet
        # (e.g. DMARC1 / DMARC2 were dropped in later cutsheet revisions).
        possible_cols_all = [
            ('Possible Hostname',         'Hostname'),
            ('Possible Interface',        'Interface'),
            ('Possible L/R',             'L/R'),
            ('Possible Rack',             'Rack'),
            ('Possible Elevation',        'Elevation'),
            ('Possible Source_port',      'Source_port'),
            ('Possible DMARC1',           'DMARC1'),
            ('Possible DMARC2',           'DMARC2'),
            ('Possible Destination_port', 'Destination_port'),
        ]
        cut_cols = set(cut_df.columns)
        possible_cols = [
            (out_col, src) for out_col, src in possible_cols_all
            if src in cut_cols
        ]

        possible_data = {col: [] for col, _ in possible_cols}
        for act in act_rows:
            zh   = act.get('Z Hostname', '')
            zi   = act.get('Z Interface', '')
            match = z_lookup.get((zh, zi)) if zh else None
            # Fallback: if exact sub-port not in cutsheet, try its pair.
            # s0↔s1 are a pair, s2↔s3 are a pair.
            if match is None and zh and zi:
                mate = paired_subport(zi)
                if mate:
                    match = z_lookup.get((zh, mate))
            for col, src in possible_cols:
                if match is not None:
                    val = match.get(src, '')
                else:
                    val = ''
                possible_data[col].append(val)

        # Write Possible columns
        pink_col_indices = []
        start = ws_m.max_column + 1
        for c_off, (col_name, _) in enumerate(possible_cols):
            col_idx = start + c_off
            pink_col_indices.append(col_idx)
            hdr = ws_m.cell(row=1, column=col_idx, value=col_name)
            hdr.font = Font(bold=True, name='Arial', size=10)
            hdr.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            hdr.fill  = yellow_fill
            hdr.border = bd
            ws_m.column_dimensions[get_column_letter(col_idx)].width = max(len(col_name)+3, 14)
            for r_off, val in enumerate(possible_data[col_name]):
                cell = ws_m.cell(row=r_off+2, column=col_idx,
                                 value=val if val != '' else None)
                cell.fill   = pink_fill
                cell.border = bd

        # Write the active-neighbour Z columns as a second "Possible" block.
        # Renamed with a "Possible Z" prefix (so they don't collide with the
        # primary Possible block above) and highlighted pink like the rest.
        act_z_cols = [
            ('Possible Z Hostname',  'Z Hostname'),
            ('Possible Z Interface', 'Z Interface'),
            ('Possible Z L/R',       'Z L/R'),
            ('Possible Z Rack',      'Z Rack'),
            ('Possible Z Elevation', 'Z Elevation'),
        ]
        start2 = ws_m.max_column + 1
        for c_off, (disp_name, src_key) in enumerate(act_z_cols):
            col_idx = start2 + c_off
            pink_col_indices.append(col_idx)
            hdr = ws_m.cell(row=1, column=col_idx, value=disp_name)
            hdr.font = Font(bold=True, name='Arial', size=10)
            hdr.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            hdr.fill  = yellow_fill
            hdr.border = bd
            ws_m.column_dimensions[get_column_letter(col_idx)].width = max(len(disp_name)+3, 14)
            for r_off, act in enumerate(act_rows):
                val = act.get(src_key, '')
                cell = ws_m.cell(row=r_off+2, column=col_idx,
                                 value=val if val != '' else None)
                cell.fill   = pink_fill
                cell.border = bd

    # ── 6c. Strip unwanted columns across every tab ─────────────────────────
    if COLUMNS_TO_REMOVE:
        log(f"  · Stripping columns: {', '.join(COLUMNS_TO_REMOVE)}")

        def _norm_col(s):
            return (str(s)
                    .replace('\u2011', '-').replace('\u2013', '-')
                    .replace('\u2014', '-')
                    .strip().lower())

        drop_set = {_norm_col(c) for c in COLUMNS_TO_REMOVE}
        for sheet_name in wb.sheetnames:
            ws_x = wb[sheet_name]
            header = [ws_x.cell(row=1, column=c).value
                      for c in range(1, ws_x.max_column + 1)]
            # Find 1-based indices to drop, delete right-to-left so earlier
            # indices stay valid as columns shift left.
            to_drop = [i + 1 for i, h in enumerate(header)
                       if _norm_col(h) in drop_set]
            for idx in sorted(to_drop, reverse=True):
                ws_x.delete_cols(idx)

    # ── 6d. Remove duplicate rows (Downlinks / Optics / combined_fec) ────────
    # Runs after all population/enrichment so the comparison sees final values.
    # Downlinks & combined_fec: drop exact-duplicate rows (keep first).
    # Optics: collapse to one row per link (Hostname+Interface+Source_port+
    #         Destination_port), keeping the row with the highest Measured (dBm).
    def _dedup_exact(ws_d):
        seen, to_delete = set(), []
        for r in range(2, ws_d.max_row + 1):
            row_key = tuple(
                '' if ws_d.cell(row=r, column=c).value is None
                else str(ws_d.cell(row=r, column=c).value)
                for c in range(1, ws_d.max_column + 1)
            )
            if row_key in seen:
                to_delete.append(r)
            else:
                seen.add(row_key)
        for r in sorted(to_delete, reverse=True):
            ws_d.delete_rows(r)
        return len(to_delete)

    def _dedup_optics_keep_highest(ws_d):
        header = [ws_d.cell(row=1, column=c).value
                  for c in range(1, ws_d.max_column + 1)]
        link_names = ['Hostname', 'Interface', 'Source_port', 'Destination_port']
        if not all(n in header for n in link_names) or 'Measured (dBm)' not in header:
            return _dedup_exact(ws_d)   # fall back if expected columns absent
        key_idx  = [header.index(n) + 1 for n in link_names]
        meas_idx = header.index('Measured (dBm)') + 1
        # For each link key, find the row with the highest measured value.
        best_row = {}     # key -> (measured_float, row_number)
        for r in range(2, ws_d.max_row + 1):
            key = tuple(str(ws_d.cell(row=r, column=c).value or '').strip()
                        for c in key_idx)
            try:
                meas = float(ws_d.cell(row=r, column=meas_idx).value)
            except (TypeError, ValueError):
                meas = float('-inf')
            if key not in best_row or meas > best_row[key][0]:
                best_row[key] = (meas, r)
        keep = {row for _, row in best_row.values()}
        to_delete = [r for r in range(2, ws_d.max_row + 1) if r not in keep]
        for r in sorted(to_delete, reverse=True):
            ws_d.delete_rows(r)
        return len(to_delete)

    for tab in ('Downlinks', 'Optics', 'combined_fec',
                'T1 Optics', 'T1 combined_fec'):
        if tab not in wb.sheetnames:
            continue
        removed = (_dedup_optics_keep_highest(wb[tab])
                   if tab in ('Optics', 'T1 Optics')
                   else _dedup_exact(wb[tab]))
        if removed:
            log(f"  · Removed {removed} duplicate row(s) from {tab}")

    # ── 7. Summary tab (per Rack breakdown) ─────────────────────────────────
    log("  · Creating Summary tab")

    # Gather rack counts per sheet from the workbook data
    tab_rack  = {}
    all_racks = set()
    no_fill_s   = PatternFill(fill_type=None)
    yellow_fill_s = PatternFill('solid', start_color=YELLOW)
    center_s  = Alignment(horizontal='center', vertical='center', wrap_text=False)
    bd_s      = thin_border()

    def _s(cell, value, bold=False, header=False):
        cell.value     = value
        cell.font      = Font(bold=bold, name='Arial', size=10)
        cell.alignment = center_s
        cell.border    = bd_s
        cell.fill      = yellow_fill_s if header else no_fill_s

    for sname in wb.sheetnames:
        ws_tmp = wb[sname]
        hdr = [ws_tmp.cell(row=1, column=c).value for c in range(1, ws_tmp.max_column+1)]
        if 'Rack' not in hdr:
            tab_rack[sname] = {}
            continue
        rack_col = hdr.index('Rack') + 1
        counts = {}
        for r in range(2, ws_tmp.max_row + 1):
            val = ws_tmp.cell(row=r, column=rack_col).value
            if val is not None:
                try:
                    k = int(float(str(val)))
                    counts[k] = counts.get(k, 0) + 1
                    all_racks.add(k)
                except ValueError:
                    pass
        tab_rack[sname] = counts

    racks      = sorted(all_racks)
    tabs_order = [n for n in wb.sheetnames]
    total_cols = 1 + len(racks) + 1  # Tab Name + per rack + Total

    # Excel treats sheet names case-insensitively. The new report ships a
    # lowercase 'summary' tab; delete any case-variant before creating ours.
    for existing in list(wb.sheetnames):
        if existing.lower() == 'summary':
            del wb[existing]
    wb.create_sheet('Summary', 0)
    ws_s = wb['Summary']

    # Title
    ws_s.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_c = ws_s.cell(row=1, column=1, value='Tab Summary by Rack')
    title_c.font = Font(bold=True, name='Arial', size=13)
    title_c.alignment = center_s
    title_c.border    = bd_s
    title_c.fill      = yellow_fill_s
    ws_s.row_dimensions[1].height = 28

    # Header row
    _s(ws_s.cell(row=2, column=1), 'Tab Name', bold=True, header=True)
    for c, rack in enumerate(racks, start=2):
        _s(ws_s.cell(row=2, column=c), str(rack), bold=True, header=True)
    _s(ws_s.cell(row=2, column=total_cols), 'Total', bold=True, header=True)

    # Data rows (exclude the Summary tab itself — case-insensitive, since the
    # source report may ship a lowercase 'summary' tab that we delete later)
    rack_totals = {r: 0 for r in racks}
    data_tabs   = [n for n in tabs_order if n.lower() != 'summary']
    for i, tab_name in enumerate(data_tabs, start=3):
        _s(ws_s.cell(row=i, column=1), tab_name)
        row_total = 0
        for c, rack in enumerate(racks, start=2):
            count = tab_rack.get(tab_name, {}).get(rack, 0)
            _s(ws_s.cell(row=i, column=c), count if count > 0 else '')
            rack_totals[rack] += count
            row_total += count
        _s(ws_s.cell(row=i, column=total_cols), row_total, bold=True)

    # Grand total row
    tot_r = 3 + len(data_tabs)
    _s(ws_s.cell(row=tot_r, column=1), 'TOTAL', bold=True)
    grand = 0
    for c, rack in enumerate(racks, start=2):
        _s(ws_s.cell(row=tot_r, column=c), rack_totals[rack], bold=True)
        grand += rack_totals[rack]
    _s(ws_s.cell(row=tot_r, column=total_cols), grand, bold=True)

    # Column widths
    ws_s.column_dimensions['A'].width = 20
    for c in range(2, total_cols + 1):
        ws_s.column_dimensions[get_column_letter(c)].width = 14

    # ── 8. No fill + borders (preserve pink in Mismatches) ──────────────────
    log("  · Removing fills and applying borders")
    pink_col_indices = []
    # Recompute by name from the current header so every pink column gets
    # filled correctly regardless of how many columns were stripped.
    if 'Mismatches' in wb.sheetnames:
        ws_m = wb['Mismatches']
        m_header = [ws_m.cell(row=1, column=c).value
                    for c in range(1, ws_m.max_column + 1)]
        # Only the Possible columns are pink; Z columns are left unfilled.
        pink_col_indices = [
            i + 1 for i, h in enumerate(m_header)
            if (h and str(h).startswith('Possible '))
        ]

    for sheet_name in wb.sheetnames:
        pcols = pink_col_indices if sheet_name == 'Mismatches' else []
        clear_and_border(wb[sheet_name], pink_cols=pcols)

    # ── 8c. Centre-align all cells across all tabs ──────────────────────────
    log("  · Aligning all cells to middle-centre")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=False)
    for sheet_name in wb.sheetnames:
        for row in wb[sheet_name].iter_rows():
            for cell in row:
                cell.alignment = center_align

    # ── 8b. Add NOTE column + autofilter to all tabs ───────────────────────
    log("  · Adding NOTE column and filters to all tabs")
    no_fill     = PatternFill(fill_type=None)
    yellow_fill = PatternFill('solid', start_color=YELLOW)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for col_name in ['NOTE']:
            col_idx = ws.max_column + 1
            hdr = ws.cell(row=1, column=col_idx, value=col_name)
            hdr.font      = Font(bold=True, name='Arial', size=10)
            hdr.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            hdr.fill      = yellow_fill
            hdr.border    = thin_border()
            ws.column_dimensions[get_column_letter(col_idx)].width = 16
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=col_idx)
                cell.fill   = no_fill
                cell.border = thin_border()
        if ws.max_row > 1 and ws.max_column > 0:
            ws.auto_filter.ref = ws.dimensions

    # ── 8d. Grey-out rows whose link also appears in a reference tab ────────
    #   Optics          greyed where the link is interface-down (Downlinks)
    #   combined_fec    greyed where the link is also flagged in Optics
    #   T1 Optics       greyed where the same link is also flagged in Optics
    #   T1 combined_fec greyed where the same link is also flagged in T1 Optics
    #
    # The standard tabs are all T0-oriented, so a column-by-column key compare
    # (`_grey_matching`) recognises the same link. The T1 tabs are flipped
    # relative to the standard Optics tab (local end = T1, Z end = T0), so the
    # same physical link has mirrored columns there — those comparisons use an
    # orientation-independent endpoint-pair key (`_grey_matching_link`) instead.
    LINK_COLS = [
        'Hostname', 'Interface', 'L/R', 'Rack', 'Elevation',
        'Source_port', 'DMARC1', 'DMARC2', 'Destination_port',
        'Z Hostname', 'Z Interface', 'Z L/R', 'Z Rack', 'Z Elevation',
    ]
    GREY_FONT_COLOR = 'FFD3D3D3'  # light grey

    def _grey_row(ws, r):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(
                bold=cell.font.bold if cell.font else False,
                name=(cell.font.name if cell.font else None) or 'Arial',
                size=(cell.font.size if cell.font else None) or 10,
                color=GREY_FONT_COLOR,
            )

    def _grey_matching(target_tab, ref_tab):
        """Grey target rows whose link columns exactly match a reference row.
        Use only when both tabs share the same orientation."""
        if target_tab not in wb.sheetnames or ref_tab not in wb.sheetnames:
            return
        ws_t = wb[target_tab]
        ws_r = wb[ref_tab]
        t_hdr = [ws_t.cell(row=1, column=c).value
                 for c in range(1, ws_t.max_column + 1)]
        r_hdr = [ws_r.cell(row=1, column=c).value
                 for c in range(1, ws_r.max_column + 1)]
        common = [c for c in LINK_COLS if c in t_hdr and c in r_hdr]
        if not common:
            log(f"    ⚠ No shared link columns between {target_tab} and "
                f"{ref_tab} — skipped")
            return
        log(f"  · Greying out {target_tab} rows that match {ref_tab}")
        t_idx = {c: t_hdr.index(c) + 1 for c in common}
        r_idx = {c: r_hdr.index(c) + 1 for c in common}
        ref_keys = set()
        for r in range(2, ws_r.max_row + 1):
            ref_keys.add(tuple(
                str(ws_r.cell(row=r, column=r_idx[c]).value or '').strip()
                for c in common))
        for r in range(2, ws_t.max_row + 1):
            key = tuple(
                str(ws_t.cell(row=r, column=t_idx[c]).value or '').strip()
                for c in common)
            if key in ref_keys:
                _grey_row(ws_t, r)

    def _grey_matching_link(target_tab, ref_tab):
        """Grey target rows whose physical link also appears in the reference
        tab, matching on an orientation-independent endpoint pair so a flipped
        (T1-side) row still matches its standard (T0-side) counterpart.

        A link is the unordered pair of its two ends — (Hostname, Interface)
        and (Z Hostname, Z Interface) — so the same cable matches regardless of
        which end each tab calls 'local'."""
        if target_tab not in wb.sheetnames or ref_tab not in wb.sheetnames:
            return
        ws_t = wb[target_tab]
        ws_r = wb[ref_tab]
        need = ['Hostname', 'Interface', 'Z Hostname', 'Z Interface']
        t_hdr = [ws_t.cell(row=1, column=c).value
                 for c in range(1, ws_t.max_column + 1)]
        r_hdr = [ws_r.cell(row=1, column=c).value
                 for c in range(1, ws_r.max_column + 1)]
        if not all(c in t_hdr for c in need) or not all(c in r_hdr for c in need):
            log(f"    ⚠ Missing endpoint columns between {target_tab} and "
                f"{ref_tab} — skipped")
            return
        log(f"  · Greying out {target_tab} rows whose link is in {ref_tab}")
        ti = {c: t_hdr.index(c) + 1 for c in need}
        ri = {c: r_hdr.index(c) + 1 for c in need}

        def pair_key(ws, r, idx):
            a = (str(ws.cell(row=r, column=idx['Hostname']).value or '').strip(),
                 str(ws.cell(row=r, column=idx['Interface']).value or '').strip())
            z = (str(ws.cell(row=r, column=idx['Z Hostname']).value or '').strip(),
                 str(ws.cell(row=r, column=idx['Z Interface']).value or '').strip())
            if not a[0] or not z[0]:
                return None          # incomplete endpoint — don't match
            return tuple(sorted([a, z]))

        ref_keys = set()
        for r in range(2, ws_r.max_row + 1):
            k = pair_key(ws_r, r, ri)
            if k:
                ref_keys.add(k)
        for r in range(2, ws_t.max_row + 1):
            k = pair_key(ws_t, r, ti)
            if k and k in ref_keys:
                _grey_row(ws_t, r)

    _grey_matching('Optics',       'Downlinks')
    _grey_matching('combined_fec', 'Optics')
    _grey_matching_link('T1 Optics',       'Optics')
    _grey_matching_link('T1 combined_fec', 'T1 Optics')

    # ── 8e. Expand all columns and rows on every sheet ──────────────────────
    log("  · Expanding all columns and rows")
    for sheet_name in wb.sheetnames:
        autofit_sheet(wb[sheet_name])

    # ── 8f. Highlight reciprocal mismatch pairs orange ──────────────────────
    log("  · Highlighting reciprocal mismatch pairs")
    highlight_mismatch_pairs(wb, log)

    # ── 8g. Keep only the whitelisted output tabs ───────────────────────────
    # Anything not in this set is dropped, whatever it's called. Matching is
    # case-insensitive so 'combined_fec'/'Combined_Fec' etc. all survive.
    keep_lower = {t.lower() for t in TABS_TO_KEEP}
    for existing in list(wb.sheetnames):
        if existing.lower() not in keep_lower:
            log(f"  · Removing {existing} (not in keep list)")
            del wb[existing]

    wb.save(output_path)
    return finalize_output(input_path, output_path, log)


# ── Mismatch pair detection / highlight ──────────────────────────────────────

ORANGE = 'FFA500'   # reciprocal-swap pair highlight

def highlight_mismatch_pairs(wb, log=lambda *_: None):
    """Mismatches tab: find reciprocal swap pairs and highlight them orange.

    Two mismatch rows are a *pair* when one row's A-side block
    (Hostname, Interface, L/R, Rack, Elevation, Source_port, Destination_port)
    is EXACTLY equal to the other row's Possible block (the fall-back lookup of
    where that row's cable actually lands) — and vice versa. That reciprocal
    match means the two cables are simply swapped with each other (e.g. the
    s0/s1 strands of a pair are crossed).

    For every pair found, the partner row is moved directly beneath its match
    (inserted as the row under) and the 12-column block
    Hostname … Exp. Elevation of BOTH rows is filled orange. Unpaired rows are
    left where they are with no orange fill.
    """
    if 'Mismatches' not in wb.sheetnames:
        return
    ws = wb['Mismatches']
    if ws.max_row < 3:
        return

    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    def idxs(names):
        return [header.index(n) + 1 for n in names if n in header]

    a_cols = idxs(['Hostname', 'Interface', 'L/R', 'Rack', 'Elevation',
                   'Source_port', 'Destination_port'])
    p_cols = idxs(['Possible Hostname', 'Possible Interface', 'Possible L/R',
                   'Possible Rack', 'Possible Elevation',
                   'Possible Source_port', 'Possible Destination_port'])
    orange_cols = idxs(['Hostname', 'Interface', 'L/R', 'Rack', 'Elevation',
                        'Source_port', 'DMARC1', 'DMARC2', 'Destination_port',
                        'Z Hostname', 'Z Interface', 'Z L/R',
                        'Z Rack', 'Z Elevation'])
    if not a_cols or not p_cols:
        return

    ncol, nrow = ws.max_column, ws.max_row
    rows = [[ws.cell(row=r, column=c).value for c in range(1, ncol + 1)]
            for r in range(2, nrow + 1)]

    def key(rv, cols):
        return tuple(str(rv[c - 1] if rv[c - 1] is not None else '').strip()
                     for c in cols)

    a_keys = [key(rv, a_cols) for rv in rows]
    p_keys = [key(rv, p_cols) for rv in rows]

    n = len(rows)
    partner = [None] * n
    for i in range(n):
        if partner[i] is not None or not any(a_keys[i]):
            continue
        for j in range(i + 1, n):
            if partner[j] is not None:
                continue
            if a_keys[i] == p_keys[j] and a_keys[j] == p_keys[i]:
                partner[i] = j
                partner[j] = i
                break

    # New row order: all pairs first (each partner directly under its match,
    # in order of first appearance), then every unpaired row pushed to the end.
    placed = [False] * n
    order = []
    for i in range(n):
        if placed[i] or partner[i] is None:
            continue
        j = partner[i]
        order.append(i); placed[i] = True
        if not placed[j]:
            order.append(j); placed[j] = True
    for i in range(n):                       # leftover = unpaired rows
        if not placed[i]:
            order.append(i); placed[i] = True

    paired = {i for i in range(n) if partner[i] is not None}
    npairs = len(paired) // 2

    # Number the pairs in the order their first row appears, so the colour can
    # alternate: 1st pair orange, 2nd pair yellow, 3rd orange, 4th yellow, …
    pair_no = {}
    counter = 0
    for i in range(n):
        j = partner[i]
        if j is not None and i < j:
            pair_no[i] = counter
            pair_no[j] = counter
            counter += 1

    # Re-apply column-determined formatting (pink stays on Possible cols only)
    # plus the alternating pair highlight on the listed columns.
    bd          = thin_border()
    pink_fill   = PatternFill('solid', start_color=PINK)
    orange_fill = PatternFill('solid', start_color=ORANGE)
    yellow_fill = PatternFill('solid', start_color=YELLOW)
    pink_names  = {'Possible Hostname', 'Possible Interface', 'Possible L/R',
                   'Possible Rack', 'Possible Elevation', 'Possible Source_port',
                   'Possible DMARC1', 'Possible DMARC2', 'Possible Destination_port'}
    # Pink applies to every "Possible *" column (including the Possible Z block).
    pink_idx    = {i + 1 for i, h in enumerate(header)
                   if h and (h in pink_names or str(h).startswith('Possible '))}
    orange_set  = set(orange_cols)
    no_fill     = PatternFill(fill_type=None)

    for out_off, src_i in enumerate(order):
        r  = out_off + 2
        rv = rows[src_i]
        is_pair = src_i in paired
        pair_fill = (orange_fill if pair_no.get(src_i, 0) % 2 == 0
                     else yellow_fill)
        for c in range(1, ncol + 1):
            cell = ws.cell(row=r, column=c, value=rv[c - 1])
            cell.font = Font(name='Arial', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = bd
            if is_pair and c in orange_set:
                cell.fill = pair_fill
            elif c in pink_idx:
                cell.fill = pink_fill
            else:
                cell.fill = no_fill

    log(f"  · Highlighted {npairs} mismatch pair(s) (alternating orange/yellow)")
