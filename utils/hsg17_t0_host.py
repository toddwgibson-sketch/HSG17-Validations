"""
HSG17 T0-to-Host / T1-to-T0 clean processor (6-stage architecture).

Stages (as designed):
1. Ingest
2. Normalize (block derivation is here)
3. Enrich (attach ALL columns from matched allconnections row as AllConn_* + parsing for clean fields)
4. Analyze (light clustering / grouping for rich output)
5. Format (professional 5-tab workbook + Summary, with columns in target order + AllConn_ at end)
6. Log (extract counts per DH block → central logger)

This module is intentionally self-contained and testable.
All columns from the (new simpler) allconnections will be in the output as requested.
"""

from __future__ import annotations

import io
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Tuple, Optional, Dict, List, Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from utils.hsg17_models import normalize_block, PG_TO_DH, RACK_TO_BLOCK


def parse_label_fields(label_str: str) -> dict:
    """Parse the single-string Full Label or EasyMark from allconnections into structured fields.
    Why single string? The allconnections export stores rich path/patch panel info this way
    (for EasyMark labeling software or as a flat export field) – it packs source device/color,
    rack/elev, PP label, dest device, multiple paths into one \n-delimited blob.
    We parse it here so the report can have separate useful columns (Rack, Elevation, Source_port as the PP label, etc.)
    instead of one big unreadable string.
    """
    if not label_str:
        return {}
    text = str(label_str)
    result = {}
    # PP labels like PP.HSG17:4.DH4.R2809.U27.MPO1 or PP.HSG17.3.DH2... (synthetic compat)
    pps = re.findall(r'PP\.HSG17[:.][^\s\n|]+', text, re.IGNORECASE)
    if pps:
        result['pp_label'] = pps[0]
        if len(pps) > 1:
            result['pp_label2'] = pps[1]
    # Racks and U (elevation): Rack 2809 U1 or R2809.U1 or in PP R2809.U27
    rack_elevs = re.findall(r'(?:Rack\s*|R)(\d+)[.\s]*U(\d+)', text, re.IGNORECASE)
    if rack_elevs:
        result['rack'] = rack_elevs[0][0]
        result['elev'] = rack_elevs[0][1]
        if len(rack_elevs) > 1:
            result['rack2'] = rack_elevs[1][0]
            result['elev2'] = rack_elevs[1][1]
    # Also catch standalone Rxxxx.Uyy in PP labels even without preceding Rack word
    if 'rack' not in result:
        re_r = re.search(r'R(\d+)[.\s]*U(\d+)', text, re.IGNORECASE)
        if re_r:
            result['rack'] = re_r.group(1)
            result['elev'] = re_r.group(2)
    # Source device line e.g. hsg17-... color or device
    dev_match = re.search(r'(hsg17-[^\s\n|]+)', text, re.IGNORECASE)
    if dev_match:
        result['device'] = dev_match.group(1)
    # Color if present e.g. Blue, Orange
    color_match = re.search(r'\b(Blue|Orange|Green|Red|Yellow|Black|White)\b', text, re.IGNORECASE)
    if color_match:
        result['color'] = color_match.group(1)
    return result


# ================== Styles (matching the "pretty" vibe the user likes) ==================
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="1F4E79")
BODY_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)
ORANGE_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


ERROR_SHEETS = [
    "LLDP Mismatch + Link Down",
    "Optic Errors",
    "FEC_BER Errors",
    "Interface Down Errors",
]


def _safe_str(x: Any) -> str:
    return "" if pd.isna(x) else str(x).strip()


def _derive_block_from_pp(pp_text: str) -> Optional[str]:
    """Priority #1 rule from DESIGN: parse DH from Patch Panel Matrix / Full Label."""
    if not pp_text:
        return None
    # Look for DH2, DH-4, DH.7, PP.HSG17.4.DH4. etc.
    m = re.search(r"(?:PP\.[^.]+\.)?DH[\.-]?(\d+)", str(pp_text), re.IGNORECASE)
    if m:
        num = m.group(1)
        if num == "102":
            return "DH-102 (Spines)"
        if len(num) == 1:
            num = f"00{num}"
        elif len(num) == 2 and num[0] != "1":
            num = f"0{num}"
        candidate = f"DH-{num}"
        return candidate
    return None


def _derive_block_from_rack(rack: str) -> Optional[str]:
    """Secondary rule using rack number ranges (handles hsg17:5708:3 and bare numbers)."""
    if not rack:
        return None
    r = str(rack).strip()
    m = re.search(r"(\d{3,4})", r)
    if m:
        rack_num = m.group(1)
        prefix2 = rack_num[:2]
        if prefix2 in RACK_TO_BLOCK:
            return RACK_TO_BLOCK[prefix2]
        # 08xx / 09xx small racks etc. can be extended here later
    return None


def derive_block(row: pd.Series, pp_col_candidates: List[str], rack_col_candidates: List[str]) -> Optional[str]:
    """
    Full priority strategy from DESIGN.md:
    1. Patch Panel DH (strongest signal)
    2. hsg17: rack / numeric rack range
    3. Device name patterns (t1 → spines, etc.)
    4. Fallback None
    """
    # 1. PP columns (Patch Panel Matrix, Full Label, EasyMark...)
    for col in pp_col_candidates:
        if col in row and pd.notna(row[col]):
            blk = _derive_block_from_pp(_safe_str(row[col]))
            if blk:
                return blk

    # 2. Rack columns
    for col in rack_col_candidates:
        if col in row and pd.notna(row[col]):
            blk = _derive_block_from_rack(_safe_str(row[col]))
            if blk:
                return blk

    # 3. Device name heuristic (T1 spines + IPR)
    for col in ["Device A Name", "Device Name", "Source Device Name", "device_a", "device_b", "Device B Name"]:
        if col in row and pd.notna(row[col]):
            name = _safe_str(row[col]).lower()
            if "t1" in name or "-t1-" in name or "t1-r" in name:
                return "DH-102 (Spines)"
            if "ip-r" in name or "q2-ip" in name:
                return "IPR (PG-201)"

    return None


# ================== Stage 1: Ingest ==================
def ingest(allconnections_bytes: bytes, cutsheet_bytes: bytes) -> Dict[str, pd.DataFrame]:
    """Load both inputs. Returns dict of useful frames."""
    # All connections (large — keep all columns for strongest possible PP / cutsheet enrichment and future joins)
    # Use sheet_name=0 for robustness (first sheet) instead of hardcoding "Sheet1"
    allc = pd.read_excel(io.BytesIO(allconnections_bytes), sheet_name=0).copy()

    # Cutsheet — all 4 error sheets + Summary
    cuts = {}
    if not cutsheet_bytes:
        raise ValueError("Pre-classified cutsheet is required. Please upload the rack_validation_merged_*.xlsx file containing the error sheets.")
    try:
        with io.BytesIO(cutsheet_bytes) as bio:
            xl = pd.ExcelFile(bio)
            for sheet in ERROR_SHEETS:
                if sheet in xl.sheet_names:
                    cuts[sheet] = pd.read_excel(xl, sheet_name=sheet)
            if "Summary" in xl.sheet_names:
                cuts["Summary"] = pd.read_excel(xl, sheet_name="Summary")
    except Exception as e:
        raise ValueError(f"Failed to read cutsheet file as Excel (expected the pre-classified validation file): {e}")

    if not any(s in cuts for s in ERROR_SHEETS):
        raise ValueError(f"Cutsheet loaded but contained none of the expected error sheets {ERROR_SHEETS}. Check that you uploaded the correct pre-classified file.")

    return {"allconnections": allc, "cutsheet_sheets": cuts}


# ================== Stage 2: Normalize (block derivation lives here) ==================
def normalize(ingested: Dict[str, Any]) -> Dict[str, pd.DataFrame]:
    """Add canonical 'Block' column to every error sheet. Return enriched frames."""
    cuts = ingested["cutsheet_sheets"]
    allc = ingested["allconnections"]

    pp_candidates = ["Patch Panel Matrix", "Full Label", "EasyMark+ --- Patch Panels", "PP Matrix"]
    rack_candidates = ["Device A Rack", "Device B Rack", "Rack", "device_a_rack", "hsg17:"]

    normalized = {}

    for sheet_name, df in cuts.items():
        if sheet_name == "Summary":
            normalized[sheet_name] = df
            continue

        df = df.copy()
        df["Block"] = df.apply(lambda r: derive_block(r, pp_candidates, rack_candidates), axis=1)

        # Fallback: try a few more obvious columns that often contain DH in this data
        mask = df["Block"].isna()
        if mask.any():
            for col in ["Patch Panel Matrix", "Source Device Location", "Device A Rack"]:
                if col in df.columns:
                    df.loc[mask, "Block"] = df.loc[mask, col].apply(_derive_block_from_pp)
                    mask = df["Block"].isna()
                    if not mask.any():
                        break

        # Final safety: mark unknowns (they will still be logged under a sensible bucket later if needed)
        df["Block"] = df["Block"].fillna("UNKNOWN")

        normalized[sheet_name] = df

    normalized["allconnections"] = allc
    return normalized


# ================== Union-Find for Connected Components ==================
class UnionFind:
    def __init__(self):
        self.parent: Dict[str, str] = {}

    def find(self, x: str) -> str:
        if x not in self.parent:
            self.parent[x] = x
        if self.parent[x] != x:
            self.parent[x] = self.find(self.parent[x])
        return self.parent[x]

    def union(self, x: str, y: str):
        px, py = self.find(x), self.find(y)
        if px != py:
            self.parent[px] = py


def cluster_mismatches(df: pd.DataFrame) -> pd.DataFrame:
    """
    Rich connected-component clustering for LLDP Mismatch + Link Down.
    Nodes are 'device:port'. We union actual + expected mismatches so
    swapped pairs and multi-port issues form visible clusters.
    Returns df with new 'Cluster' column (integer group id).
    """
    if df.empty:
        df = df.copy()
        df["Cluster"] = -1
        return df

    uf = UnionFind()

    for _, row in df.iterrows():
        a = f"{_safe_str(row.get('Device A Name'))}:{_safe_str(row.get('Device A Port'))}"
        b = f"{_safe_str(row.get('Device B Name'))}:{_safe_str(row.get('Device B Port'))}"
        exp_b = f"{_safe_str(row.get('Expected Device B Name'))}:{_safe_str(row.get('Expected Device B Port'))}"

        uf.union(a, b)
        if exp_b and exp_b != b:
            uf.union(a, exp_b)

    # Assign dense cluster ids
    component_to_id: Dict[str, int] = {}
    next_id = 0
    cluster_col = []
    for _, row in df.iterrows():
        a = f"{_safe_str(row.get('Device A Name'))}:{_safe_str(row.get('Device A Port'))}"
        root = uf.find(a)
        if root not in component_to_id:
            component_to_id[root] = next_id
            next_id += 1
        cluster_col.append(component_to_id[root])

    df = df.copy()
    df["Cluster"] = cluster_col
    return df


# ================== Stage 3+4: Enrich + Analyze ==================
def enrich_and_analyze(normalized: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
    """Enrichment + real mismatch clustering + PP joins.

    Stronger joins (v2):
    - Index allconnections by BOTH ends (A and B device+port) so errors on either side can hit.
    - Try A-side key first on error rows, then B-side key if miss.
    - Support many common column name variants on both allc and cutsheet sides.
    - Capture richer PP info (Full Label + EasyMark combined when available).
    - Pull additional useful fields (t0 switch info etc.) if present in allc.
    """
    out = {}
    allc = normalized.get("allconnections")

    # Build fast lookup from allconnections for PP enrichment (index by (dev, port) for *both* sides of the link)
    pp_lookup: Dict[tuple, Dict] = {}
    if allc is not None and not allc.empty:
        # Flexible column discovery on the (now full) allconnections frame.
        # Use keyword heuristics because real files have varying names like "Dev A", "Source", "swp port", "Easy Mark + PP" etc.
        def _find_likely_dev_name_col(side: str) -> Optional[str]:
            """Prioritize exact known columns from this allconnections file."""
            if side == 'a':
                for cand in ['DeviceA Name', 'DeviceA', 'Device A Name']:
                    if cand in allc.columns:
                        return cand
            else:
                for cand in ['DeviceB Name', 'DeviceB', 'Device B Name']:
                    if cand in allc.columns:
                        return cand
            # fallback keyword (avoid bad ones)
            keywords = ["device", "dev", "name"]
            side_kw = ["a", "source", "from", "left"] if side == "a" else ["b", "dest", "to", "right", "destination"]
            avoid = ["transceiver", "oracle", "part", "sku", "type", "building", "floor"]
            best = None
            for c in allc.columns:
                cl = str(c).lower()
                if any(k in cl for k in keywords) and any(k in cl for k in side_kw) and "port" not in cl and "rack" not in cl and not any(a in cl for a in avoid):
                    if "name" in cl or side_kw[0] in cl:
                        return c
                    best = c
            return best or None

        def _find_likely_port_col(side: str) -> Optional[str]:
            if side == 'a':
                for cand in ['DeviceA Port', 'DeviceA Physical Port', 'Device A Port']:
                    if cand in allc.columns:
                        return cand
            else:
                for cand in ['DeviceB Port', 'DeviceB Physical Port', 'Device B Port']:
                    if cand in allc.columns:
                        return cand
            keywords = ["port", "swp", "interface", "intf"]
            side_kw = ["a", "source", "from"] if side == "a" else ["b", "dest", "to"]
            avoid = ["transceiver", "oracle", "part", "sku"]
            for c in allc.columns:
                cl = str(c).lower()
                if any(k in cl for k in keywords) and any(k in cl for k in side_kw) and not any(a in cl for a in avoid):
                    return c
            for c in allc.columns:
                cl = str(c).lower()
                if any(k in cl for k in keywords) and "rack" not in cl and not any(a in cl for a in avoid):
                    return c
            return None

        dev_a_col = _find_likely_dev_name_col("a")
        port_a_col = _find_likely_port_col("a")
        dev_b_col = _find_likely_dev_name_col("b")
        port_b_col = _find_likely_port_col("b")

        def _find_pp_col() -> Optional[str]:
            for c in allc.columns:
                cl = str(c).lower()
                if ("full" in cl and "label" in cl) or "easymark" in cl or ("patch" in cl and ("panel" in cl or "label" in cl or "mark" in cl)):
                    return c
            return None

        full_label_col = _find_pp_col()
        easymark_col = full_label_col  # often same or we take the best one
        t0_col = None
        for c in allc.columns:
            cl = str(c).lower()
            if "t0" in cl and ("switch" in cl or "port" in cl or "sw" in cl):
                t0_col = c
                break

        # Additional finders for the columns the user wants in the report (rack, elevation, dmarc, z/t1 interface, history, etc.)
        # These will be pulled from the matching allconnections row to enrich the error rows with full device/rack/elevation/PP info.
        def _find_col_kw(keywords, side_kws=None, avoid=None):
            avoid = avoid or []
            for c in allc.columns:
                cl = str(c).lower()
                if any(k in cl for k in keywords) and all(a not in cl for a in avoid):
                    if side_kws:
                        if any(sk in cl for sk in side_kws):
                            return c
                    else:
                        return c
            return None

        rack_a_col = _find_col_kw(["rack"], ["a", "source", "from", "left"])
        elev_a_col = _find_col_kw(["elev", "elevation"], ["a", "source", "from"])
        rack_b_col = _find_col_kw(["rack"], ["b", "dest", "to", "right"])
        elev_b_col = _find_col_kw(["elev", "elevation"], ["b", "dest", "to"])
        dmarc1_col = _find_col_kw(["dmarc1", "dmarc 1", "dmarc-1"])
        dmarc2_col = _find_col_kw(["dmarc2", "dmarc 2", "dmarc-2"])
        z_interface_col = _find_col_kw(["z interface", "z_interface", "z int"])
        z_rack_col = _find_col_kw(["z rack", "z_rack"])
        z_elev_col = _find_col_kw(["z elev", "z_elev", "z elevation"])
        t1_rack_col = _find_col_kw(["t1 rack", "t1_rack", "t1 rack"])
        t1_port_col = _find_col_kw(["t1 port", "t1_port", "t1 port"])
        interface_col = _find_col_kw(["interface", "^intf$"], avoid=["z "])
        history_col = _find_col_kw(["history"])

        def _get_conn_info(r: pd.Series, key_dev_col: Optional[str], key_port_col: Optional[str]) -> Dict[str, str]:
            """Pull PP + other useful device/connection info from this allconnections row for enrichment."""
            fl = _safe_str(r.get(full_label_col)) if full_label_col else ""
            em = _safe_str(r.get(easymark_col)) if easymark_col else ""
            t0 = _safe_str(r.get(t0_col)) if t0_col else ""
            if fl and em:
                pp = f"{fl} | {em}"
            else:
                pp = fl or em or ""

            info = {"PP": pp}
            if t0:
                info["T0"] = t0

            # Parse the single-string fields so we can split into structured columns (Rack, Elevation, Source_port as clean PP label, etc.)
            parsed = parse_label_fields(fl)
            for k, v in parsed.items():
                info[k] = v

            # Pull explicit short/structured cols from allc that are already split (Source_port etc.)
            for col in ['Source_port', 'Destination_port', 'T0 Switch Port', 'Cable Info', 'Cable Color',
                        'DeviceA Rack', 'DeviceB Rack', 'DeviceA RU', 'DeviceB RU', 'RackA', 'RackB']:
                val = _safe_str(r.get(col))
                if val:
                    info[col] = val

            # Pull "mate" (the other end of the connection) device info -- this is the key "device information from allconnects"
            if key_dev_col:
                other_dev_col = dev_b_col if key_dev_col == dev_a_col else dev_a_col
                other_port_col = port_b_col if key_dev_col == dev_a_col else port_a_col
                mate_name = _safe_str(r.get(other_dev_col)) if other_dev_col else ""
                mate_port = _safe_str(r.get(other_port_col)) if other_port_col else ""
                if mate_name or mate_port:
                    info["Mate"] = f"{mate_name}:{mate_port}" if mate_port else mate_name

            # Pull the key columns from allconnections that provide the "device information", rack, elevation, ports, labels etc.
            # These are what the user needs in the report sheets.
            key_allc_cols = ['Source_port', 'Destination_port', 'T0 Switch Port', 'DeviceA Rack', 'DeviceA RU', 'DeviceA Name', 'DeviceA Port', 'DeviceB Rack', 'DeviceB RU', 'DeviceB Name', 'DeviceB Port', 'Full Label', 'EasyMark+ --- Patch Panels', 'Cable Info', 'Cable Color', 'RackA', 'RackB', 'OHR', 'T0']
            for col in key_allc_cols:
                val = _safe_str(r.get(col))
                if val:
                    info[col] = val

            # Also pull any other potentially useful like dmarc, z, t1, interface, history, elev if present in this file
            for col in allc.columns:
                cl = str(col).lower()
                if any(k in cl for k in ['dmarc', 'z ', 't1', 'elev', 'interface', 'history', 'rack', 'port']) and col not in key_allc_cols and col not in (full_label_col, easymark_col, key_dev_col, key_port_col):
                    val = _safe_str(r.get(col))
                    if val and len(val) > 2:
                        nice = col.replace(' ', '_').replace('/', '_').replace('.', '_')
                        if nice not in info:
                            info[nice] = val

            return info

        if dev_a_col and port_a_col:
            for _, r in allc.iterrows():
                key = (_safe_str(r.get(dev_a_col)), _safe_str(r.get(port_a_col)))
                if key[0] or key[1]:  # avoid empty keys
                    # Store ALL columns from this allconnections row, as user wants all in output
                    all_info = {c: _safe_str(r.get(c)) for c in allc.columns}
                    pp_lookup[key] = all_info

        if dev_b_col and port_b_col:
            for _, r in allc.iterrows():
                key = (_safe_str(r.get(dev_b_col)), _safe_str(r.get(port_b_col)))
                if key[0] or key[1]:
                    existing = pp_lookup.get(key, {})
                    new_info = {c: _safe_str(r.get(c)) for c in allc.columns}
                    # Merge to keep all
                    merged = {**existing, **new_info}
                    pp_lookup[key] = merged

    # Build device-name-only lookup for fallback matching when exact port doesn't match in allc (e.g. different swp port for the T0 in error vs allc host connections)
    # Normalize dev name to first token (before space) to handle 'name port' vs 'name' in cuts vs allc
    dev_lookup = {}
    for key, inf in list(pp_lookup.items()):
        dev = key[0]
        if dev:
            norm_dev = dev.split()[0] if ' ' in dev else dev
            if norm_dev not in dev_lookup:
                dev_lookup[norm_dev] = inf

    for name, df in normalized.items():
        if name in ("allconnections", "Summary"):
            out[name] = df
            continue

        df = df.copy()

        # === PP Enrichment join (from allconnections) - stronger A then B side ===
        if pp_lookup and not df.empty:
            # Error-side column discovery (cutsheet side) - more flexible to handle variations in pre-classified files
            def _find_error_col(candidates):
                for cand in candidates:
                    for c in df.columns:
                        cl = str(c).lower().strip()
                        if cand.lower() in cl or cl == cand.lower():
                            return c
                return None

            dev_a_col = _find_error_col(["Device A Name", "Source Device Name", "Device Name", "device_a", "A Name", "From Device"])
            port_a_col = _find_error_col(["Device A Port", "Source Device Port", "Device A Port", "A Port", "From Port"])
            dev_b_col = _find_error_col(["Device B Name", "Destination Device Name", "Device B Name", "device_b", "B Name", "To Device"])
            port_b_col = _find_error_col(["Device B Port", "Destination Device Port", "Device B Port", "B Port", "To Port"])

            if dev_a_col and port_a_col:
                def _enrich_row(r):
                    # Try A side
                    key = (_safe_str(r.get(dev_a_col)), _safe_str(r.get(port_a_col)))
                    info = pp_lookup.get(key, {})
                    if not info and dev_b_col and port_b_col:
                        # Fallback to B side on this error row
                        key_b = (_safe_str(r.get(dev_b_col)), _safe_str(r.get(port_b_col)))
                        info = pp_lookup.get(key_b, {})

                    if not info:
                        # Fallback to device name only (when exact port doesn't match in allc, e.g. swp3s0 in error vs swp1s0 in allc for same t0 device)
                        dev = _safe_str(r.get(dev_b_col)) if dev_b_col else _safe_str(r.get(dev_a_col))
                        norm_dev = dev.split()[0] if ' ' in dev else dev
                        if norm_dev in dev_lookup:
                            info = dev_lookup[norm_dev]

                    # Attach the device/connection info from the matched allconnections row.
                    # This is the absolute minimum the user asked for: pull real device details / path from the big connections file
                    # to give context on what the error row "should" be connected to.
                    enriched = {}
                    if info:
                        # Attach ALL columns from the matched allconnections row, prefixed with AllConn_ 
                        # so they appear in the output report as user requested "all the columns in the output"
                        for k, v in info.items():
                            nice_k = f"AllConn_{k.replace(' ', '_').replace('+', 'Plus').replace('/', '_')}"
                            enriched[nice_k] = str(v) if v else ""  # include even empty, so column exists; drop later only if all empty in sheet

                        # For backward compat / key fields, also set some top-level like before if wanted, but since all via AllConn, ok
                        # But keep PP_Enriched as clean if possible
                        pp_label_clean = info.get("Source_port") or info.get("EasyMark+") or ""
                        if pp_label_clean:
                            # take first PP if multi
                            import re
                            pps = re.findall(r'PP\.HSG17[:.][^\s\n|]+', pp_label_clean, re.IGNORECASE)
                            if pps:
                                pp_label_clean = pps[0]
                        enriched["PP_Enriched"] = str(pp_label_clean) if pp_label_clean else ""
                    else:
                        enriched["PP_Enriched"] = ""
                    return pd.Series(enriched)

                extra = df.apply(_enrich_row, axis=1)
                for col in extra.columns:
                    df[col] = extra[col]

                # Do not drop AllConn_* as user wants all columns from the new allconnections in output.
                # Only drop other optionals if completely empty.
                cols_to_check = [c for c in list(df.columns) if not c.startswith("AllConn_") and c in ["DMARC1", "DMARC2", "Z Interface", "Z Rack", "Z Elevation", "Possible T1 Rack / U", "T1 Rack", "Possible T1 Port", "Interface", "History", "T0 Switch Port", "Cable Info", "Cable Color", "Source_port", "Destination_port", "Rack", "Elevation"] or any(x in c for x in ["Z ", "T1 "])]
                for col in cols_to_check:
                    if col in df.columns:
                        non_empty = df[col].fillna("").astype(str).str.strip().ne("")
                        if not non_empty.any():
                            df = df.drop(columns=[col], errors="ignore")

            # Fallback for PP_Enriched if the allconnections join gave nothing for the row
            # Parse the label string from cutsheet's orig_col (which may be the big single string) and split into columns
            if "PP_Enriched" in df.columns:
                empty_pp = df["PP_Enriched"].astype(str).str.len() == 0
                if empty_pp.any():
                    for orig_col in ["Patch Panel Matrix", "Full Label", "EasyMark+ --- Patch Panels", "PP Matrix", "Patch Panel"]:
                        if orig_col in df.columns:
                            fallbacks = df.loc[empty_pp, orig_col].apply(_safe_str)
                            # For each, parse and set split columns, and clean PP_Enriched
                            for idx in empty_pp[empty_pp].index:
                                val = fallbacks.get(idx, "")
                                parsed = parse_label_fields(val)
                                if parsed.get('pp_label'):
                                    df.loc[idx, "PP_Enriched"] = parsed['pp_label']
                                elif val:
                                    df.loc[idx, "PP_Enriched"] = val  # fallback to original if no parse
                                if parsed.get('rack'):
                                    df.loc[idx, "Rack"] = parsed['rack']
                                if parsed.get('elev'):
                                    df.loc[idx, "Elevation"] = parsed['elev']
                                if parsed.get('rack2'):
                                    df.loc[idx, "Z Rack"] = parsed['rack2']
                                if parsed.get('elev2'):
                                    df.loc[idx, "Z Elevation"] = parsed['elev2']
                                if parsed.get('pp_label2'):
                                    df.loc[idx, "Destination_port"] = parsed['pp_label2']
                            break

            # For Interface Down, add simple "full switch down" flag - this makes the report actually help prioritize (if one device has many downs, likely the whole switch/rack is affected)
            if name == "Interface Down Errors" and "Device A Name" in df.columns:
                dev_counts = df["Device A Name"].value_counts().to_dict()
                df["A Device Down Count"] = df["Device A Name"].map(dev_counts)
                df["Full Switch Down?"] = df["A Device Down Count"].apply(lambda c: "LIKELY - inspect entire device/switch" if c >= 3 else "")

        # Promote to prioritize target columns (from user's CSV), then our additions including all AllConn_ from the new allconnections
        target_order = [
            'Interface', 'L&R', 'Rack', 'Elevation', 'Source_port', 'Destination_port', 'Z Interface', 'L&R.1', 'Z Rack', 'Z Elevation',
            'Possible Device A', 'Possible Rack / U', 'Possible Source Port', 'Possible DMARC1', 'Possible DMARC2', 'Possible Dest Port',
            'Possible T1 Rack / U', 'Possible T1 Port', 'Act. Interface', 'Act. Rack', 'Act. Elevation', 'Exp. Interface', 'Exp. Rack', 'Exp. Elevation', 'History'
        ]
        front = [c for c in target_order if c in df.columns]
        for extra in ["Block", "Cluster", "Cluster Size", "Notes", "PP_Enriched"]:
            if extra in df.columns and extra not in front:
                front.append(extra)
        # Add ALL AllConn_ columns (from the new allconnections file)
        for c in list(df.columns):
            if c.startswith("AllConn_") and c not in front:
                front.append(c)
        other = [c for c in df.columns if c not in front]
        df = df[front + other]

        # === Real connected-component clustering on LLDP sheet ===
        if name == "LLDP Mismatch + Link Down":
            df = cluster_mismatches(df)

            # Add actionable value to the report (so it "does something" beyond repackaging input)
            if "Cluster" in df.columns and not df.empty:
                cluster_sizes = df["Cluster"].value_counts().to_dict()
                df["Cluster Size"] = df["Cluster"].map(lambda c: cluster_sizes.get(c, 0) if c != -1 else 0)
                df["Notes"] = df.apply(
                    lambda r: "Review group for cable swap / mislabel (multiple mismatches connected via expected/actual)"
                    if r.get("Cluster Size", 0) > 1 else "",
                    axis=1
                )

        out[name] = df

    out["allconnections"] = allc
    return out


# ================== Stage 5: Format ==================
def build_workbook(enriched: Dict[str, pd.DataFrame], source_basename: str) -> Tuple[bytes, str]:
    """Create the professional 5-tab Excel the user expects."""
    wb = Workbook()

    # Remove default sheet
    default = wb.active
    wb.remove(default)

    # --- Summary (counts per Block per category) ---
    summary_ws = wb.create_sheet("Summary", 0)
    summary_ws["A1"] = f"HSG17 T0-to-Host Validation Summary — {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    summary_ws["A1"].font = TITLE_FONT
    summary_ws.merge_cells("A1:F1")
    summary_ws["A2"] = f"Source: {source_basename} | Blocks derived using PP labels + rack ranges + device patterns (see DESIGN.md)"
    summary_ws["A2"].font = Font(name="Arial", size=9, italic=True, color="666666")
    summary_ws.merge_cells("A2:F2")

    # Build summary data from the 4 error sheets
    summary_rows: List[List[Any]] = [["Block", "LLDP Mismatch + Link Down", "Optic Errors", "FEC_BER Errors", "Interface Down Errors", "Total"]]

    block_counts: Dict[str, Counter] = defaultdict(Counter)

    for sheet_name in ERROR_SHEETS:
        if sheet_name in enriched:
            df = enriched[sheet_name]
            if "Block" in df.columns:
                for blk, cnt in df["Block"].value_counts().items():
                    block_counts[blk][sheet_name] = cnt

    # Sort blocks nicely (DH-001 first, then spines, then specials)
    def block_sort_key(b: str) -> tuple:
        if b.startswith("DH-00"):
            return (0, int(b.split("-")[1]))
        if b.startswith("DH-10"):
            return (1, int(b.split("-")[1][:3]))
        if "Spines" in b:
            return (2, 102)
        if "IPR" in b:
            return (3, 0)
        if "AUX" in b:
            return (4, 0)
        return (5, 0)

    all_blocks = sorted(block_counts.keys(), key=block_sort_key)

    for blk in all_blocks:
        c = block_counts[blk]
        row = [
            blk,
            c.get("LLDP Mismatch + Link Down", 0),
            c.get("Optic Errors", 0),
            c.get("FEC_BER Errors", 0),
            c.get("Interface Down Errors", 0),
        ]
        row.append(sum(row[1:]))
        summary_rows.append(row)

    # Total row
    totals = [sum(r[i] for r in summary_rows[1:]) for i in range(1, 6)]
    summary_rows.append(["TOTAL"] + totals)

    total_row_idx = 3 + len(summary_rows) - 1

    for r_idx, row in enumerate(summary_rows, start=3):
        for c_idx, val in enumerate(row, start=1):
            cell = summary_ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if r_idx == 3:  # header
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            if r_idx == total_row_idx:  # TOTAL row
                cell.font = Font(name="Arial", bold=True, size=11)
                cell.fill = LIGHT_GREEN_FILL

    # Column widths + usability
    for col in range(1, 7):
        summary_ws.column_dimensions[get_column_letter(col)].width = 28 if col == 1 else 18

    # Freeze header and add filter for the summary table too (header at row 3)
    summary_ws.freeze_panes = "A4"
    last_col_sum = get_column_letter(6)
    last_row_sum = total_row_idx
    summary_ws.auto_filter.ref = f"A3:{last_col_sum}{last_row_sum}"

    # --- The 4 detailed error sheets (with Block + Cluster promoted) ---
    for sheet_name in ERROR_SHEETS:
        if sheet_name not in enriched:
            continue
        df = enriched[sheet_name].copy()

        # Drop any remaining internal columns
        for drop in ["_pair_group", "raw_row"]:
            if drop in df.columns:
                df = df.drop(columns=[drop], errors="ignore")

        # Sanitize PP_Enriched (ensure no None/NaN leak into the xlsx - use robust check)
        if "PP_Enriched" in df.columns:
            def _clean_pp(v):
                if v is None:
                    return ""
                try:
                    if pd.isna(v):
                        return ""
                except Exception:
                    pass
                s = str(v).strip()
                if s.lower() in ("", "nan", "none", "null"):
                    return ""
                return s
            df["PP_Enriched"] = df["PP_Enriched"].apply(_clean_pp)

        # Reorder to prioritize target columns (from user's CSV), then our additions including all AllConn_ from the new allconnections
        target_order = [
            'Interface', 'L&R', 'Rack', 'Elevation', 'Source_port', 'Destination_port', 'Z Interface', 'L&R.1', 'Z Rack', 'Z Elevation',
            'Possible Device A', 'Possible Rack / U', 'Possible Source Port', 'Possible DMARC1', 'Possible DMARC2', 'Possible Dest Port',
            'Possible T1 Rack / U', 'Possible T1 Port', 'Act. Interface', 'Act. Rack', 'Act. Elevation', 'Exp. Interface', 'Exp. Rack', 'Exp. Elevation', 'History'
        ]
        front = [c for c in target_order if c in df.columns]
        for extra in ["Block", "Cluster", "Cluster Size", "Notes", "PP_Enriched"]:
            if extra in df.columns and extra not in front:
                front.append(extra)
        # Add ALL AllConn_ columns (from the new allconnections file)
        for c in list(df.columns):
            if c.startswith("AllConn_") and c not in front:
                front.append(c)
        other = [c for c in df.columns if c not in front]
        df = df[front + other]

        # For the LLDP sheet, sort rows by Cluster (stable sort) so that mismatch pairs / connected components
        # are grouped together consecutively. This makes the orange/yellow highlighting dramatically more useful
        # for visual pairing review (stronger mismatch pairing visuals).
        if sheet_name == "LLDP Mismatch + Link Down" and "Cluster" in df.columns:
            df = df.sort_values(by=["Cluster"], kind="stable").reset_index(drop=True)

        # Always sort the error sheet by Block so same-rack / same-sector issues are grouped in the report. Makes it actually usable for fixing per Block.
        if "Block" in df.columns:
            sort_keys = ["Block"]
            if "Cluster" in df.columns:
                sort_keys.append("Cluster")
            df = df.sort_values(by=sort_keys, kind="stable").reset_index(drop=True)

        # Reorder: target matched cols first, then remaining cuts cols, then key added, then ALL AllConn_ at end
        target_order = [
            'Interface', 'L&R', 'Rack', 'Elevation', 'Source_port', 'Destination_port', 'Z Interface', 'L&R.1', 'Z Rack', 'Z Elevation',
            'Possible Device A', 'Possible Rack / U', 'Possible Source Port', 'Possible DMARC1', 'Possible DMARC2', 'Possible Dest Port',
            'Possible T1 Rack / U', 'Possible T1 Port', 'Act. Interface', 'Act. Rack', 'Act. Elevation', 'Exp. Interface', 'Exp. Rack', 'Exp. Elevation', 'History'
        ]
        existing_target = [c for c in target_order if c in df.columns]
        cuts_remaining = [c for c in df.columns if not c.startswith('AllConn_') and c not in existing_target and c not in ['Block','Cluster','Cluster Size','Notes','PP_Enriched']]
        key_added = [c for c in ['Block','Cluster','Cluster Size','Notes','PP_Enriched'] if c in df.columns]
        allconn_cols = [c for c in df.columns if c.startswith('AllConn_')]
        df = df[existing_target + cuts_remaining + key_added + allconn_cols]

        ws = wb.create_sheet(sheet_name)

        # Title + source note (makes the report more self-documenting)
        num_cols = len(df.columns)
        ws["A1"] = f"HSG17 — {sheet_name}"
        ws["A1"].font = TITLE_FONT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        ws["A2"] = f"Source: {source_basename}  |  Enriched with PP from allconnections + DH Block derivation  |  Cluster = connected-component grouping on mismatches (swap pairs share id)"
        ws["A2"].font = Font(name="Arial", size=9, italic=True, color="666666")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)

        is_lldp = sheet_name == "LLDP Mismatch + Link Down"

        # Write data with rich coloring for LLDP clusters (only on the "pair" columns so the report stays readable)
        # We deliberately do NOT color the left structural columns (Block/Cluster/PP_Enriched) or the A-side context.
        lldp_pair_cols = {"Device B Name", "Device B Port", "Expected Device B Name", "Expected Device B Port"}
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=3):
            cluster_val = None
            if is_lldp and "Cluster" in df.columns:
                try:
                    cluster_idx_in_df = list(df.columns).index("Cluster")
                    cluster_val = row[cluster_idx_in_df] if cluster_idx_in_df < len(row) else None
                except:
                    cluster_val = None

            for c_idx, val in enumerate(row, start=1):
                col_name = df.columns[c_idx-1]
                # Extra safety: never write None/NaN into cells (especially for string cols like PP_Enriched)
                if val is None:
                    val = ""
                else:
                    try:
                        if pd.isna(val):
                            val = ""
                    except Exception:
                        pass
                # Force str for the enriched info columns the user wants (to avoid 1009.0 etc in Excel for rack/elev values)
                if col_name in ["Rack", "Elevation", "DMARC1", "DMARC2", "Z Rack", "Z Elevation", "T1 Rack", "Possible T1 Rack / U", "Possible T1 Port", "Z Interface", "T1 Rack", "Interface", "History", "T0 Switch Port", "Cable Info", "Cable Color"] or col_name.startswith("AllConn_") or col_name in ["Source_port", "Destination_port"] or any(x in col_name for x in ["Z ", "T1 "]):
                    val = str(val) if val is not None else ""
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = BODY_FONT
                cell.border = THIN_BORDER

                if r_idx == 3:  # header
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                else:
                    # Highlight Block (always) - structural
                    if col_name == "Block":
                        cell.fill = LIGHT_BLUE_FILL
                        cell.font = Font(name="Arial", bold=True, size=10)

                    # Cluster coloring ONLY on the actual mismatch pair columns (makes swapped rows pop without washing the whole sheet)
                    if (is_lldp and cluster_val is not None and
                            col_name != "Block" and col_name != "Cluster" and col_name != "PP_Enriched" and
                            (col_name in lldp_pair_cols or any(k in str(col_name).lower() for k in ["expected", "b name", "b port", "lldp", "mismatch"]))):
                        try:
                            cid = int(cluster_val)
                            if cid % 2 == 0:
                                cell.fill = ORANGE_FILL
                            else:
                                cell.fill = YELLOW_FILL
                            cell.font = Font(name="Arial", size=10, bold=True)
                        except (ValueError, TypeError):
                            pass

        # Auto width — sample a few data rows for better sizing on real data
        for c_idx, col in enumerate(df.columns, start=1):
            w = len(str(col)) + 2
            # sample up to 5 data rows
            for r in range(4, min(4 + 5, ws.max_row + 1)):
                cell_val = ws.cell(row=r, column=c_idx).value
                if cell_val:
                    w = max(w, min(60, len(str(cell_val)) + 2))
            ws.column_dimensions[get_column_letter(c_idx)].width = max(10, min(55, w))

        # Usability: freeze header + first 3 cols (Block/Cluster/PP stay visible when scrolling)
        ws.freeze_panes = "D4"
        # AutoFilter on the whole table (header at row 3)
        last_col = get_column_letter(len(df.columns))
        ws.auto_filter.ref = f"A3:{last_col}{ws.max_row}"

    # Filename
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"HSG17_T0_Host_{source_basename}_{ts}.xlsx"

    # Write to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), filename


# ================== Public entry point ==================
def process_hsg17_t0_host(
    allconnections_bytes: bytes,
    cutsheet_bytes: bytes,
    source_filename: str = "HSG17",
) -> Tuple[bytes, str]:
    """
    Full clean pipeline. Returns (excel_bytes, suggested_filename).
    This is what the Streamlit page will call.
    """
    base = Path(source_filename).stem.replace(" ", "_")[:40]

    # 1. Ingest
    ingested = ingest(allconnections_bytes, cutsheet_bytes)

    # 2. Normalize (block derivation)
    normalized = normalize(ingested)

    # 3+4. Enrich + Analyze (light v1)
    enriched = enrich_and_analyze(normalized)

    # 5. Format
    xlsx_bytes, filename = build_workbook(enriched, base)

    return xlsx_bytes, filename


# Convenience for the page: extract the counts we will log (after we have the workbook)
def extract_counts_for_logging(wb_bytes: bytes) -> List[Dict[str, Any]]:
    """Read the Summary tab we just created and return list of log rows."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(wb_bytes))
    if "Summary" not in wb.sheetnames:
        return []

    ws = wb["Summary"]
    rows = list(ws.iter_rows(min_row=4, values_only=True))  # skip title + header + blank

    results = []
    for row in rows:
        if not row or not row[0]:
            continue
        blk = str(row[0]).strip()
        if blk.upper() == "TOTAL":
            continue
        results.append({
            "block": blk,
            "LLDP Mismatch + Link Down": row[1] or 0,
            "Optic Errors": row[2] or 0,
            "FEC_BER Errors": row[3] or 0,
            "Interface Down Errors": row[4] or 0,
        })
    return results
